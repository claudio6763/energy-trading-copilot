# -*- coding: utf-8 -*-
"""Construtor do MODELO EXCEL — o calculo vive na planilha, nao no Python.

PRINCIPIO DE ARQUITETURA
------------------------
    Python  = CARREGADOR de dado observado (abas DADOS_*) e do MANIFESTO.
    Excel   = MOTOR DE CALCULO. Sazonalidade EWMA, curva, cenarios, VaR, PnL e
              VPL sao FORMULAS nativas. Nenhum resultado e colado.

Consequencia pratica: mudar a meia-vida, o lambda, o nivel de confianca ou o
tamanho da posicao em PARAMETROS reprecifica a planilha inteira sem rodar
Python. O Python so precisa rodar de novo quando se quer DADO novo.

O que e valor colado, por definicao, e legitimo:
    - abas DADOS_* : serie observada da CCEE/ONS (fonte, nao calculo)
    - aba MANIFESTO: proveniencia e hash
Todo o resto e formula.
"""
from __future__ import annotations

import calendar
import re
from datetime import date
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.formatting.rule import CellIsRule, ColorScaleRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.workbook.defined_name import DefinedName

# ----------------------------------------------------------------- estilo
ROXO, ROSA = "1F1235", "C4007A"
F_OBS, F_CALC, F_PREM, F_PROXY, F_FV = "E8F5E9", "E3F2FD", "FFF8E1", "FCE4EC", "ECEFF1"
F_HDR, F_TIT = "1F1235", "FAFAFC"
BORDA = Border(*[Side(style="thin", color="D9D9E0")] * 4)


def _hdr(ws, cols, linha=1, larg=16):
    for j, c in enumerate(cols, 1):
        cel = ws.cell(row=linha, column=j, value=c)
        cel.fill = PatternFill("solid", fgColor=F_HDR)
        cel.font = Font(color="FFFFFF", bold=True, size=9)
        cel.alignment = Alignment(horizontal="center", wrap_text=True, vertical="center")
        cel.border = BORDA
        ws.column_dimensions[get_column_letter(j)].width = larg
    ws.freeze_panes = ws.cell(row=linha + 1, column=1)


def _titulo(ws, texto, sub=""):
    ws["A1"] = texto
    ws["A1"].font = Font(bold=True, size=14, color=ROSA)
    if sub:
        ws["A2"] = sub
        ws["A2"].font = Font(italic=True, size=9, color="666677")


def _nome(wb, nome, ref):
    wb.defined_names.add(DefinedName(nome, attr_text=ref))


def _n(ws, r, c, v, fmt=None, fill=None, bold=False):
    cel = ws.cell(row=r, column=c, value=v)
    if fmt:
        cel.number_format = fmt
    if fill:
        cel.fill = PatternFill("solid", fgColor=fill)
    if bold:
        cel.font = Font(bold=True)
    return cel


# =============================================================== construtor
def construir(caminho: Path, dados: dict, prem, meta: dict) -> dict:
    """dados: {'pld_diario': df[data,pld], 'hidro_mensal': df[mes_ref,ena_pct,ear_pct],
               'mve': df, 'manifesto': df}"""
    wb = Workbook()
    pld_d = dados["pld_diario"].reset_index(drop=True)
    hid = dados["hidro_mensal"].reset_index(drop=True)
    mve = dados["mve"]
    man = dados["manifesto"]
    alvo = meta["alvo"]

    # =================================================== 1. DADOS (observado)
    ws = wb.active
    ws.title = "DADOS_PLD_D"
    _hdr(ws, ["data", "pld_diario_R$/MWh", "ano", "mes"], larg=14)
    for i, r in enumerate(pld_d.itertuples(index=False), start=2):
        _n(ws, i, 1, pd.Timestamp(r.data).date(), "yyyy-mm-dd", F_OBS)
        _n(ws, i, 2, float(r.pld), "#,##0.00", F_OBS)
        ws.cell(row=i, column=3, value=f"=YEAR(A{i})")
        ws.cell(row=i, column=4, value=f"=MONTH(A{i})")
    n_d = len(pld_d) + 1
    _nome(wb, "PLD_DATA", f"DADOS_PLD_D!$A$2:$A${n_d}")
    _nome(wb, "PLD_VAL", f"DADOS_PLD_D!$B$2:$B${n_d}")

    ws = wb.create_sheet("DADOS_HIDRO_M")
    _hdr(ws, ["mes_ref", "ENA_%MLT", "EAR_%max"], larg=16)
    for i, r in enumerate(hid.itertuples(index=False), start=2):
        _n(ws, i, 1, pd.Timestamp(r.mes_ref).date(), "yyyy-mm", F_OBS)
        _n(ws, i, 2, None if pd.isna(r.ena_pct) else float(r.ena_pct), "#,##0.0", F_OBS)
        _n(ws, i, 3, None if pd.isna(r.ear_pct) else float(r.ear_pct), "#,##0.0", F_OBS)
    n_h = len(hid) + 1

    ws = wb.create_sheet("DADOS_MVE")
    cols_mve = ["data_negociacao", "preco", "montante", "submercado", "vigencia_ini", "vigencia_fim"]
    _hdr(ws, cols_mve, larg=18)
    if len(mve):
        for i, r in enumerate(mve.itertuples(index=False), start=2):
            for j, cn in enumerate(cols_mve, 1):
                v = getattr(r, cn, None)
                if isinstance(v, pd.Timestamp):
                    v = v.date()
                _n(ws, i, j, None if (v is None or (isinstance(v, float) and np.isnan(v))) else v,
                   fill=F_PROXY)
    else:
        ws["A2"] = "SEM NEGOCIO PUBLICO COMPARAVEL NA DATA DE CORTE"
        ws["A2"].font = Font(italic=True, color="B00020")
    n_mve = max(len(mve), 1) + 1
    _nome(wb, "MVE_PRECO", f"DADOS_MVE!$B$2:$B${n_mve}")
    _nome(wb, "MVE_MONT", f"DADOS_MVE!$C$2:$C${n_mve}")

    # =================================================== 2. PARAMETROS
    ws = wb.create_sheet("PARAMETROS")
    _titulo(ws, "PARAMETROS DO MODELO",
            "Celulas amarelas sao editaveis. Alterar qualquer uma reprecifica a planilha inteira.")
    _hdr(ws, ["parametro", "valor", "unidade", "rotulo", "justificativa"], linha=4, larg=26)
    P = [
        ("Data de corte", meta["data_corte"], "data", "PREMISSA", "definida pelo case"),
        ("Mes de referencia (ultimo fechado)", meta["ref_mes"], "data", "CALCULADO",
         "ultimo mes com cobertura completa"),
        ("Meia-vida EWMA sazonal", meta.get("meia_vida", prem.meias_vidas_teste[1]), "dias", "CALCULADO",
         "escolhida na aba BACKTEST; pode ser sobrescrita aqui"),
        ("Sigma do kernel de calendario", prem.kernel_calendario_dias / 30.0, "meses", "PREMISSA",
         "vizinhanca circular; sem ela 24 meses dao 2 obs por mes"),
        ("Janela sazonal", prem.janela_sazonal_meses, "meses", "PREMISSA",
         "regime pos-DESSEM (2021) e pos-entrada de renovavel"),
        ("Winsor inferior", prem.winsor[0], "quantil", "PREMISSA", "robustez a spike"),
        ("Winsor superior", prem.winsor[1], "quantil", "PREMISSA", "robustez a spike"),
        ("Lambda EWMA do VaR", prem.var_lambda_ewma, "-", "PREMISSA", "padrao RiskMetrics"),
        ("Confianca do VaR", prem.var_confianca, "-", "PREMISSA", "escolha declarada"),
        ("Horizonte do VaR", prem.var_horizonte_du, "dias uteis", "PREMISSA",
         "prazo realista de desmontagem"),
        ("Limite de VaR", prem.limite_var_brl, "R$", "PREMISSA", "teto do case, nao meta"),
        ("Fracao do limite utilizavel", prem.var_orcamento_frac, "-", "PREMISSA", "buffer de risco"),
        ("Edge minimo", prem.edge_minimo_rs_mwh, "R$/MWh", "PREMISSA", "gatilho de conviccao"),
        ("Taxa de desconto", prem.taxa_desconto_aa, "a.a.", "PREMISSA", "Selic vigente"),
        ("Quantil regime seco", prem.quantil_seco, "-", "PREMISSA", "corte do indice hidro"),
        ("Quantil regime umido", prem.quantil_umido, "-", "PREMISSA", "corte do indice hidro"),
        ("Prob. cenario Base", prem.prob_cenarios["base"], "-", "PREMISSA", "premissa declarada"),
        ("Prob. cenario Seco", prem.prob_cenarios["seco"], "-", "PREMISSA", "premissa declarada"),
        ("Prob. cenario Umido", prem.prob_cenarios["umido"], "-", "PREMISSA", "premissa declarada"),
        ("Piso do PLD", meta["piso"], "R$/MWh", "OBSERVADO", "ANEEL, ano do horizonte"),
        ("Teto estrutural do PLD", meta["teto"], "R$/MWh", "OBSERVADO", "ANEEL, ano do horizonte"),
        ("Peso do componente fundamental", meta["peso_fund"], "-", "CALCULADO",
         "backtest; 0 quando nao ha CMO projetado para o horizonte"),
        ("Meia-vida do nivel (meses)", prem.meia_vida_nivel_meses, "meses", "PREMISSA",
         "EWMA do log flat dessazonalizado; media movel de 12m e lisa demais"),
        ("Alfa do nivel", 1 - 0.5 ** (1 / prem.meia_vida_nivel_meses), "-", "CALCULADO",
         "=1-0.5^(1/meia_vida)"),
        ("Lambda EWMA do forward", prem.lambda_forward, "-", "PREMISSA",
         "serie mensal do termo; menor que o 0,94 diario"),
        ("Horizonte do VaR (meses)", prem.var_horizonte_meses, "meses", "PREMISSA",
         "cadencia de reavaliacao"),
        ("Teto operacional de tamanho", prem.mwm_maximo_operacional, "MWm", "PREMISSA",
         "limite de LIQUIDEZ: acima disso a desmontagem move o proprio preco"),
        ("Fracao do limite para stress", prem.frac_stress, "-", "PREMISSA",
         "a perda no cenario adverso carregado ate a entrega tambem cabe no limite"),
        ("Peso da ancora do InfoPLD", meta.get("peso_ancora", 0.0), "-", "PREMISSA",
         "quanto a projecao oficial da CCEE pesa contra o componente estatistico"),
        ("Premio/basis a termo", meta.get("premio_basis", 0.0), "R$/MWh", "PREMISSA",
         "ZERO por premissa: projecao de PLD nao e preco bilateral, e nao ha "
         "referencia publica para estimar o premio"),
        ("Anos de amostra do regime", 6, "anos", "PREMISSA",
         "janela do estimador de cenario; alinhada com a implementacao de referencia"),
    ]
    nomes = ["DATA_CORTE", "REF_MES", "MEIA_VIDA", "SIGMA_K", "JANELA_M", "WINS_LO", "WINS_HI",
             "LAMBDA_EWMA", "CONF", "HORIZ_DU", "LIMITE_VAR", "FRAC_ORC", "EDGE_MIN", "TX_DESC",
             "Q_SECO", "Q_UMIDO", "PROB_BASE", "PROB_SECO", "PROB_UMIDO", "PISO", "TETO",
             "W_FUND", "MV_NIVEL", "ALFA_NIVEL", "LAMBDA_FWD", "HORIZ_M",
             "MWM_TETO", "FRAC_STRESS", "PESO_ANCORA", "PREMIO_BASIS", "ANOS_REG"]
    for i, ((nm, val, un, rot, jus), chave) in enumerate(zip(P, nomes), start=5):
        _n(ws, i, 1, nm, bold=True)
        cel = _n(ws, i, 2, val, fill=F_PREM if rot == "PREMISSA" else F_CALC)
        if isinstance(val, float) and abs(val) < 1:
            cel.number_format = "0.0000"
        _n(ws, i, 3, un); _n(ws, i, 4, rot); _n(ws, i, 5, jus)
        _nome(wb, chave, f"PARAMETROS!$B${i}")
    ws.column_dimensions["E"].width = 62

    # =================================================== 3. CALC_MENSAL
    ws = wb.create_sheet("CALC_MENSAL")
    _titulo(ws, "FLAT MENSAL, NIVEL E DESVIO SAZONAL",
            "Tudo calculado a partir da aba DADOS_PLD_D. Nada colado.")
    _hdr(ws, ["mes_ref", "flat_mensal", "n_dias", "dias_mes", "cobertura", "ln(flat)",
              "nivel_trailing_12m", "desvio", "idade_dias", "mes_cal", "na_janela",
              "desvio_winsor", "resid_pos_sazonal", "cobertura_mes_fechado",
              "x_deseason", "NIVEL_EWMA", "idx_mes", "desvio_num"], linha=4, larg=15)
    meses = pd.date_range(pd.Timestamp(pld_d.data.min()).to_period("M").to_timestamp(),
                          pd.Timestamp(pld_d.data.max()).to_period("M").to_timestamp(), freq="MS")
    r0 = 5
    for i, m in enumerate(meses):
        r = r0 + i
        _n(ws, r, 1, m.date(), "yyyy-mm")
        ws.cell(row=r, column=2, value=(
            f'=IFERROR(AVERAGEIFS(PLD_VAL,PLD_DATA,">="&A{r},PLD_DATA,"<="&EOMONTH(A{r},0)),"")'))
        ws.cell(row=r, column=3, value=f'=COUNTIFS(PLD_DATA,">="&A{r},PLD_DATA,"<="&EOMONTH(A{r},0))')
        ws.cell(row=r, column=4, value=f"=DAY(EOMONTH(A{r},0))")
        ws.cell(row=r, column=5, value=f"=IFERROR(C{r}/D{r},0)")
        ws.cell(row=r, column=6, value=f'=IF(OR(B{r}="",E{r}<0.98),"",LN(B{r}))')
        if i >= 11:
            ws.cell(row=r, column=7, value=f'=IF(COUNT(F{r-11}:F{r})=12,AVERAGE(F{r-11}:F{r}),"")')
        else:
            ws.cell(row=r, column=7, value='=""')
        ws.cell(row=r, column=8, value=f'=IF(OR(F{r}="",G{r}=""),"",F{r}-G{r})')
        ws.cell(row=r, column=9, value=f"=REF_MES-A{r}")
        ws.cell(row=r, column=10, value=f"=MONTH(A{r})")
        ws.cell(row=r, column=11, value=f'=IF(AND(H{r}<>"",A{r}>EDATE(REF_MES,-JANELA_M),A{r}<=REF_MES),1,0)')
        ws.cell(row=r, column=12, value=f'=IF(K{r}=1,MAX($T$5,MIN($U$5,H{r})),0)')
        ws.cell(row=r, column=14, value=f'=IF(A{r}<=REF_MES,E{r},1)')
        # x = ln(flat) - fator sazonal do mes;  nivel = EWMA recursivo de x
        ws.cell(row=r, column=15, value=(
            f'=IF(F{r}="","",F{r}-INDEX(CALC_SAZONAL!$F$6:$F$17,MONTH(A{r})))'))
        if i == 0:
            ws.cell(row=r, column=16, value=f'=IF(O{r}="","",O{r})')
        else:
            ws.cell(row=r, column=16, value=(
                f'=IF(O{r}="",P{r-1},IF(P{r-1}="",O{r},ALFA_NIVEL*O{r}+(1-ALFA_NIVEL)*P{r-1}))'))
        ws.cell(row=r, column=17, value=i)          # indice inteiro do mes
        # desvio numerico de TODA a serie. A coluna L (desvio_winsor) e zerada fora
        # da janela de 24 meses, o que impedia o backcast de reestimar fatores em
        # origens antigas — todas voltavam vazias.
        ws.cell(row=r, column=19, value=f'=IF(H{r}="",0,H{r})')
        ws.cell(row=r, column=15).number_format = "0.0000"
        ws.cell(row=r, column=16).number_format = "0.0000"
        # resid usa TODA a serie com nivel valido (nao so a janela sazonal de 24m),
        # porque o estimador de regime tem janela propria (ANOS_REG) em CALC_REGIME
        ws.cell(row=r, column=13, value=(
            f'=IF(H{r}="",0,H{r}-INDEX(CALC_SAZONAL!$F$6:$F$17,MONTH(A{r})))'))
        for c in (2, 6, 7, 8, 12, 13):
            ws.cell(row=r, column=c).number_format = "0.0000"
        ws.cell(row=r, column=5).number_format = "0.0%"
    n_m = r0 + len(meses) - 1
    # bounds de winsorizacao sobre a amostra na janela
    _n(ws, 4, 20, "lim_inf", bold=True); _n(ws, 4, 21, "lim_sup", bold=True)
    ws.cell(row=5, column=20, value=f"=PERCENTILE(FILTRO_DESV,WINS_LO)").number_format = "0.0000"
    ws.cell(row=5, column=21, value=f"=PERCENTILE(FILTRO_DESV,WINS_HI)").number_format = "0.0000"
    _n(ws, 6, 20, "amostra usada"); ws.cell(row=6, column=21, value=f"=SUM(K{r0}:K{n_m})")
    # coluna auxiliar contigua para percentil (texto e ignorado por PERCENTILE)
    _n(ws, 4, 18, "desvio_na_janela", bold=True)
    for i in range(len(meses)):
        r = r0 + i
        ws.cell(row=r, column=18, value=f'=IF(K{r}=1,H{r},"")')
    _nome(wb, "FILTRO_DESV", f"CALC_MENSAL!$R${r0}:$R${n_m}")
    _nome(wb, "M_MESREF", f"CALC_MENSAL!$A${r0}:$A${n_m}")
    _nome(wb, "M_FLAT", f"CALC_MENSAL!$B${r0}:$B${n_m}")
    _nome(wb, "M_IDADE", f"CALC_MENSAL!$I${r0}:$I${n_m}")
    _nome(wb, "M_MESCAL", f"CALC_MENSAL!$J${r0}:$J${n_m}")
    _nome(wb, "M_JAN", f"CALC_MENSAL!$K${r0}:$K${n_m}")
    _nome(wb, "M_DESVW", f"CALC_MENSAL!$L${r0}:$L${n_m}")
    _nome(wb, "M_RESID", f"CALC_MENSAL!$M${r0}:$M${n_m}")
    _nome(wb, "M_NIVEL", f"CALC_MENSAL!$G${r0}:$G${n_m}")
    _nome(wb, "M_NIVEL_EWMA", f"CALC_MENSAL!$P${r0}:$P${n_m}")
    _nome(wb, "M_IDX", f"CALC_MENSAL!$Q${r0}:$Q${n_m}")
    _nome(wb, "M_DESV_NUM", f"CALC_MENSAL!$S${r0}:$S${n_m}")

    # =================================================== 4. CALC_SAZONAL
    ws = wb.create_sheet("CALC_SAZONAL")
    _titulo(ws, "FATORES SAZONAIS — EWMA DE RECENCIA x KERNEL DE CALENDARIO",
            "peso(i,m) = EXP(-LN(2)*idade_i/MEIA_VIDA) * EXP(-0.5*(dist_circular(mes_i,m)/SIGMA_K)^2)")
    _hdr(ws, ["mes", "num_ponderado", "den_ponderado", "log_fator_bruto", "log_fator_centrado",
              "log_fator", "FATOR", "n_efetivo"], linha=5, larg=17)

    def peso_arr(m_cell: str) -> str:
        """Peso de recencia x kernel circular de calendario, SEM IF().

        IF() dentro de SUMPRODUCT nao e avaliado como matriz no Excel (funciona no
        LibreOffice, o que mascarou o bug). O resultado era numerador e denominador
        em #VALUE!, capturados pelo IFERROR -> log_fator 0 -> TODOS os fatores 1,00 e
        a curva chapada. A distancia circular vira aritmetica pura:
            circ = (|d|<=6)*|d| + (|d|>6)*(12-|d|)
        """
        d = f"ABS(M_MESCAL-{m_cell})"
        circ = f"(({d}<=6)*{d}+({d}>6)*(12-{d}))"
        return (f"M_JAN*EXP(-LN(2)*M_IDADE/MEIA_VIDA)"
                f"*EXP(-0.5*({circ}/SIGMA_K)^2)")

    for k in range(12):
        r = 6 + k
        _n(ws, r, 1, k + 1)
        w = peso_arr(f"$A{r}")
        ws.cell(row=r, column=2, value=f"=SUMPRODUCT({w}*M_DESVW)")
        ws.cell(row=r, column=3, value=f"=SUMPRODUCT({w})")
        ws.cell(row=r, column=4, value=f'=IFERROR(B{r}/C{r},0)')
        ws.cell(row=r, column=5, value=f"=D{r}-AVERAGE($D$6:$D$17)")
        ws.cell(row=r, column=6, value=f"=E{r}")
        ws.cell(row=r, column=7, value=f"=EXP(F{r})/AVERAGE_FATOR")
        ws.cell(row=r, column=8, value=f"=IFERROR(SUMPRODUCT({w})^2/SUMPRODUCT(({w})^2),0)")
        for c in (2, 3, 4, 5, 6, 7, 8):
            ws.cell(row=r, column=c).number_format = "0.0000"
    _n(ws, 19, 1, "media dos EXP(log_fator) — normalizador", bold=True)
    ws.cell(row=19, column=6, value="=AVERAGE(EXP_AUX)").number_format = "0.000000"
    _n(ws, 5, 10, "EXP(log_fator)", bold=True)
    for k in range(12):
        ws.cell(row=6 + k, column=10, value=f"=EXP(F{6+k})").number_format = "0.0000"
    _nome(wb, "EXP_AUX", "CALC_SAZONAL!$J$6:$J$17")
    _nome(wb, "AVERAGE_FATOR", "CALC_SAZONAL!$F$19")
    _nome(wb, "FATOR_MES", "CALC_SAZONAL!$G$6:$G$17")
    _n(ws, 21, 1, "Soma ponderada dos fatores (controle: deve ser 12)", bold=True)
    ws.cell(row=21, column=6, value="=SUM(FATOR_MES)").number_format = "0.000000"

    ch = BarChart(); ch.title = "Fator sazonal por mes"; ch.height, ch.width = 7, 14
    ch.add_data(Reference(ws, min_col=7, min_row=5, max_row=17), titles_from_data=True)
    ch.set_categories(Reference(ws, min_col=1, min_row=6, max_row=17))
    ws.add_chart(ch, "L5")

    # =================================================== 5. BACKTEST
    ws = wb.create_sheet("BACKTEST")
    _titulo(ws, "SELECAO DA MEIA-VIDA — WALK-FORWARD 1 MES A FRENTE",
            "Em cada origem, os fatores usam APENAS observacoes ate a origem. Sem vazamento.")
    hls = list(prem.meias_vidas_teste)
    _hdr(ws, ["origem", "mes_alvo", "real", "nivel_na_origem"] +
         [f"prev_{h}d" for h in hls] + [f"erro_{h}d" for h in hls], linha=4, larg=15)
    n_orig = 12
    linhas_alvo = list(range(n_m - n_orig + 1, n_m + 1))
    for i, ra in enumerate(linhas_alvo):
        r = 5 + i
        ro = ra - 1
        ws.cell(row=r, column=1, value=f"=CALC_MENSAL!A{ro}")
        ws.cell(row=r, column=2, value=f"=CALC_MENSAL!A{ra}")
        ws.cell(row=r, column=3, value=f"=CALC_MENSAL!B{ra}")
        ws.cell(row=r, column=4, value=f"=CALC_MENSAL!G{ro}")
        for j, h in enumerate(hls):
            d = f"ABS(M_MESCAL-MONTH($B{r}))"
            circ = f"(({d}<=6)*{d}+({d}>6)*(12-{d}))"
            jan = (f"(M_MESREF<=$A{r})*(M_MESREF>EDATE($A{r},-JANELA_M))*(M_DESVW<>0)")
            w = f"{jan}*EXP(-LN(2)*($A{r}-M_MESREF)/{h})*EXP(-0.5*({circ}/SIGMA_K)^2)"
            s = f"IFERROR(SUMPRODUCT({w}*M_DESVW)/SUMPRODUCT({w}),0)"
            ws.cell(row=r, column=5 + j, value=f'=IFERROR(EXP($D{r}+{s}),"")')
            ws.cell(row=r, column=5 + len(hls) + j,
                    value=f'=IF(OR($C{r}="",{get_column_letter(5+j)}{r}=""),"",'
                          f'{get_column_letter(5+j)}{r}-$C{r})')
        for c in range(3, 5 + 2 * len(hls)):
            ws.cell(row=r, column=c).number_format = "#,##0.00"
    rf = 5 + n_orig + 1
    _n(ws, rf, 1, "RMSE", bold=True)
    _n(ws, rf + 1, 1, "MAE", bold=True)
    _n(ws, rf + 2, 1, "vies", bold=True)
    for j, h in enumerate(hls):
        col = get_column_letter(5 + len(hls) + j)
        ws.cell(row=rf, column=5 + j,
                value=f"=SQRT(SUMSQ({col}5:{col}{4+n_orig})/COUNT({col}5:{col}{4+n_orig}))")
        ws.cell(row=rf + 1, column=5 + j,
                value=f"=SUMPRODUCT(ABS({col}5:{col}{4+n_orig}))/COUNT({col}5:{col}{4+n_orig})")
        ws.cell(row=rf + 2, column=5 + j, value=f"=AVERAGE({col}5:{col}{4+n_orig})")
        for rr in (rf, rf + 1, rf + 2):
            ws.cell(row=rr, column=5 + j).number_format = "#,##0.00"
        _n(ws, rf - 1, 5 + j, f"{h} dias", bold=True)
    _n(ws, rf + 4, 1, "MEIA-VIDA ESCOLHIDA (menor RMSE)", bold=True)
    fc = get_column_letter(5); lc = get_column_letter(4 + len(hls))
    ws.cell(row=rf + 4, column=5,
            value=f"=INDEX({{{','.join(str(h) for h in hls)}}},MATCH(MIN({fc}{rf}:{lc}{rf}),{fc}{rf}:{lc}{rf},0))")
    ws.cell(row=rf + 4, column=5).fill = PatternFill("solid", fgColor=F_CALC)
    _n(ws, rf + 5, 1, "Cole este valor em PARAMETROS!MEIA_VIDA para usar a escolha do backtest.")

    # =================================================== 6. CALC_REGIME
    ws = wb.create_sheet("CALC_REGIME")
    _titulo(ws, "REGIME HIDROLOGICO E MULTIPLICADORES DE CENARIO",
            "indice = 0,65*z(ENA %MLT) + 0,35*z(EAR %max). Efeito = media dos residuos por regime.")
    _hdr(ws, ["mes_ref", "ENA_%MLT", "EAR_%max", "z_ENA", "z_EAR", "indice_hidro",
              "regime", "resid_do_mes", "na_amostra"], linha=4, larg=15)
    for i in range(len(hid)):
        r = 5 + i
        ws.cell(row=r, column=1, value=f"=DADOS_HIDRO_M!A{i+2}")
        ws.cell(row=r, column=1).number_format = "yyyy-mm"
        ws.cell(row=r, column=2, value=f"=DADOS_HIDRO_M!B{i+2}")
        ws.cell(row=r, column=3, value=f"=DADOS_HIDRO_M!C{i+2}")
        ws.cell(row=r, column=4, value=f'=IF(B{r}="","",(B{r}-AVERAGE(ENA_COL))/STDEVP(ENA_COL))')
        ws.cell(row=r, column=5, value=f'=IF(C{r}="","",(C{r}-AVERAGE(EAR_COL))/STDEVP(EAR_COL))')
        ws.cell(row=r, column=6, value=f'=IF(D{r}="","",0.65*D{r}+0.35*IF(E{r}="",0,E{r}))')
        ws.cell(row=r, column=7, value=(
            f'=IF(F{r}="","",IF(F{r}<=LIM_SECO,"seco",IF(F{r}>=LIM_UMIDO,"umido","base")))'))
        ws.cell(row=r, column=8, value=(
            f'=IFERROR(INDEX(M_RESID,MATCH(A{r},M_MESREF,0)),0)'))
        ws.cell(row=r, column=9, value=(
            f'=IF(AND(A{r}>EDATE(REF_MES,-12*ANOS_REG),A{r}<=REF_MES,H{r}<>0),1,0)'))
        for c in (4, 5, 6, 8):
            ws.cell(row=r, column=c).number_format = "0.000"
    n_r = 4 + len(hid)
    _nome(wb, "ENA_COL", f"CALC_REGIME!$B$5:$B${n_r}")
    _nome(wb, "EAR_COL", f"CALC_REGIME!$C$5:$C${n_r}")
    _nome(wb, "IDX_COL", f"CALC_REGIME!$F$5:$F${n_r}")
    _nome(wb, "REG_COL", f"CALC_REGIME!$G$5:$G${n_r}")
    _nome(wb, "RESID_COL", f"CALC_REGIME!$H$5:$H${n_r}")
    _nome(wb, "AMOSTRA_COL", f"CALC_REGIME!$I$5:$I${n_r}")
    lin = n_r + 2
    for k, (rot, form, fmt) in enumerate([
        ("limiar seco (quantil Q_SECO)", "=PERCENTILE(IDX_COL,Q_SECO)", "0.000"),
        ("limiar umido (quantil Q_UMIDO)", "=PERCENTILE(IDX_COL,Q_UMIDO)", "0.000"),
        ("media resid | base", '=AVERAGEIFS(RESID_COL,REG_COL,"base",AMOSTRA_COL,1)', "0.0000"),
        ("media resid | seco", '=AVERAGEIFS(RESID_COL,REG_COL,"seco",AMOSTRA_COL,1)', "0.0000"),
        ("media resid | umido", '=AVERAGEIFS(RESID_COL,REG_COL,"umido",AMOSTRA_COL,1)', "0.0000"),
        ("k_SECO", f"=EXP(B{lin+3}-B{lin+2})", "0.000"),
        ("k_UMIDO", f"=EXP(B{lin+4}-B{lin+2})", "0.000"),
        ("n seco", '=COUNTIFS(REG_COL,"seco",AMOSTRA_COL,1)', "0"),
        ("n base", '=COUNTIFS(REG_COL,"base",AMOSTRA_COL,1)', "0"),
        ("n umido", '=COUNTIFS(REG_COL,"umido",AMOSTRA_COL,1)', "0"),
        ("ordenacao coerente (Seco>1>Umido)", f'=IF(AND(B{lin+5}>1,B{lin+6}<1),"SIM","INVERTIDA - INVESTIGAR")', None),
    ]):
        _n(ws, lin + k, 1, rot, bold=True)
        c = ws.cell(row=lin + k, column=2, value=form)
        if fmt:
            c.number_format = fmt
        c.fill = PatternFill("solid", fgColor=F_CALC)
    _nome(wb, "LIM_SECO", f"CALC_REGIME!$B${lin}")
    _nome(wb, "LIM_UMIDO", f"CALC_REGIME!$B${lin+1}")
    _nome(wb, "K_SECO", f"CALC_REGIME!$B${lin+5}")
    _nome(wb, "K_UMIDO", f"CALC_REGIME!$B${lin+6}")
    ws.conditional_formatting.add(f"G5:G{n_r}", ColorScaleRule(
        start_type="min", start_color="FFCDD2", end_type="max", end_color="C8E6C9"))

    # =================================================== 7. CALC_VAR
    ws = wb.create_sheet("CALC_VAR_SPOT")
    _titulo(ws, "CHALLENGER — VaR SOBRE O PLD SPOT (NAO E O METODO PRINCIPAL)",
            "Serve de TETO. Spot nao e o fator de risco de um contrato a termo: reverte a "
            "media e tem piso/teto administrativos. O metodo principal esta em CALC_FWD.")
    _hdr(ws, ["data", "pld", "var_log", "var_abs", "ewma_var", "vol_1d", "z_1d",
              "soma_h", "ewma_var_h", "vol_h", "z_h", "z_1d^2"], linha=4, larg=13)
    lim_linhas = min(len(pld_d), 3000)
    ini = len(pld_d) - lim_linhas
    for i in range(lim_linhas):
        r = 5 + i
        src = ini + i + 2
        ws.cell(row=r, column=1, value=f"=DADOS_PLD_D!A{src}")
        ws.cell(row=r, column=1).number_format = "yyyy-mm-dd"
        ws.cell(row=r, column=2, value=f"=DADOS_PLD_D!B{src}")
        if i == 0:
            ws.cell(row=r, column=3, value='=""'); ws.cell(row=r, column=4, value='=""')
            ws.cell(row=r, column=5, value=f"=VARP(B5:B{min(4+250,4+lim_linhas)})")
        else:
            ws.cell(row=r, column=3, value=f"=LN(B{r}/B{r-1})")
            ws.cell(row=r, column=4, value=f"=B{r}-B{r-1}")
            ws.cell(row=r, column=5, value=f"=LAMBDA_*E{r-1}+(1-LAMBDA_)*C{r}^2")
        ws.cell(row=r, column=6, value=f"=SQRT(E{r})")
        ws.cell(row=r, column=7, value=f'=IF(OR(C{r}="",F{r}=0),"",C{r}/F{r})')
        if i >= prem.var_horizonte_du:
            ws.cell(row=r, column=8, value=f"=SUM(C{r-prem.var_horizonte_du+1}:C{r})")
        else:
            ws.cell(row=r, column=8, value='=""')
        if i == prem.var_horizonte_du:
            ws.cell(row=r, column=9, value=f"=H{r}^2")
        elif i > prem.var_horizonte_du:
            ws.cell(row=r, column=9, value=f'=IF(H{r}="","",LAMBDA_*I{r-1}+(1-LAMBDA_)*H{r}^2)')
        else:
            ws.cell(row=r, column=9, value='=""')
        # vol EX-ANTE: padroniza x_t pela variancia estimada ATE t-1 (sem olhar x_t)
        ws.cell(row=r, column=10, value=(f'=IF(OR(I{r}="",I{r-1}=""),"",SQRT(I{r-1}))'
                                         if i > prem.var_horizonte_du + 1 else '=""'))
        ws.cell(row=r, column=11, value=f'=IF(OR(H{r}="",J{r}="",J{r}=0),"",H{r}/J{r})')
        ws.cell(row=r, column=12, value=f'=IF(G{r}="","",G{r}^2)')
        for c in (2, 3, 4, 6, 7, 8, 10, 11, 12):
            ws.cell(row=r, column=c).number_format = "0.0000"
    n_v = 4 + lim_linhas
    _nome(wb, "LAMBDA_", "PARAMETROS!$B$12")
    _nome(wb, "Z_H", f"CALC_VAR_SPOT!$K$5:$K${n_v}")
    _nome(wb, "VOL_H_ULT", f"CALC_VAR_SPOT!$J${n_v}")
    _nome(wb, "VAR_H_ULT", f"CALC_VAR_SPOT!$I${n_v}")
    _nome(wb, "PLD_ULT", f"CALC_VAR_SPOT!$B${n_v}")
    b = n_v + 2
    itens = [
        ("Observacoes de z_h", "=COUNT(Z_H)", "0"),
        ("Vol EWMA prevista para o proximo horizonte", "=SQRT(VAR_H_ULT)", "0.0000"),
        ("Quantil de z_h em (1-CONF)", "=PERCENTILE(Z_H,1-CONF)", "0.0000"),
        ("Media de z_h abaixo do quantil (ES)", f'=AVERAGEIF(Z_H,"<="&B{b+2})', "0.0000"),
        ("Choque de VaR em log", f"=B{b+2}*B{b+1}", "0.0000"),
        ("Choque de ES em log", f"=B{b+3}*B{b+1}", "0.0000"),
        ("VaR de preco SPOT (challenger)", f"=ABS(FV_FLAT*(EXP(B{b+4})-1))", "#,##0.00"),
        ("ES de preco SPOT (challenger)", f"=ABS(FV_FLAT*(EXP(B{b+5})-1))", "#,##0.00"),
        ("Curtose de z_1d (controle)",
         "=SUMPRODUCT((CALC_VAR_SPOT!$G$7:$G$" + str(n_v) + "-AVERAGE(CALC_VAR_SPOT!$G$7:$G$" + str(n_v)
         + "))^4)/COUNT(CALC_VAR_SPOT!$G$7:$G$" + str(n_v) + ")/STDEVP(CALC_VAR_SPOT!$G$7:$G$"
         + str(n_v) + ")^4", "0.00"),
        ("Autocorr de z_1d^2 (controle)",
         "=CORREL(CALC_VAR_SPOT!$L$8:$L$" + str(n_v) + ",CALC_VAR_SPOT!$L$7:$L$" + str(n_v - 1) + ")", "0.000"),
    ]
    for k, (rot, form, fmt) in enumerate(itens):
        _n(ws, b + k, 1, rot, bold=True)
        if form:
            c = ws.cell(row=b + k, column=2, value=form)
            c.number_format = fmt
            c.fill = PatternFill("solid", fgColor=F_CALC)
    _nome(wb, "VAR_SPOT", f"CALC_VAR_SPOT!$B${b+6}")

    # =================================================== 7b. CALC_FWD
    ws = wb.create_sheet("CALC_FWD")
    _titulo(ws, "BACKCAST DO FORWARD — MARCACAO A MERCADO DA MESMA POSICAO NO PASSADO",
            "Em cada origem, o MESMO strip de entrega e avaliado so com dados ate ali. "
            "E esta serie, e nao o PLD spot, que define o risco de um contrato a termo.")
    meses_prod = [int(pd.Timestamp(m).month) for m in alvo]
    horas_prod = [pd.Period(pd.Timestamp(m), freq="M").days_in_month * 24 for m in alvo]
    cabec = (["origem", "nivel_ewma"] + [f"s_mes_{m}" for m in range(1, 13)] +
             ["norm=media(EXP(s))", "fator_pond", "FV_termo", "ln(FV)", "ret_log",
              "ewma_var", "vol_exante", "z", "idx_origem"])
    _hdr(ws, cabec, linha=5, larg=12)
    # As origens param no MES DE REFERENCIA. Incluir o mes corrente parcial
    # (cobertura < 100%) acrescentava uma origem espuria no fim e deslocava toda a
    # serie de retornos em uma posicao frente a implementacao de referencia.
    _ref_ts = pd.Timestamp(meta["ref_mes"]).to_period("M").to_timestamp()
    idx_ref = list(meses).index(_ref_ts)
    linha_ref = r0 + idx_ref
    n_or = max(2, min(prem.n_origens_forward, idx_ref - 17))
    linhas_or = list(range(linha_ref - n_or + 1, linha_ref + 1))
    prim = 6
    for i, ro in enumerate(linhas_or):
        r = prim + i
        ws.cell(row=r, column=1, value=f"=CALC_MENSAL!A{ro}")
        ws.cell(row=r, column=1).number_format = "yyyy-mm"
        ws.cell(row=r, column=2, value=f"=CALC_MENSAL!P{ro}")
        # Idade e janela por INDICE INTEIRO de mes, nunca por aritmetica de datas:
        # data dentro de SUMPRODUCT devolve #VALUE! no Excel.
        ws.cell(row=r, column=23, value=f"=INDEX(M_IDX,MATCH($A{r},M_MESREF,0))")
        for m in range(1, 13):
            d = f"ABS(M_MESCAL-{m})"
            circ = f"(({d}<=6)*{d}+({d}>6)*(12-{d}))"
            jan = f"(M_IDX<=$W{r})*(M_IDX>$W{r}-JANELA_M)*(M_DESV_NUM<>0)"
            idade = f"($W{r}-M_IDX)*30.4375"
            w = f"{jan}*EXP(-LN(2)*{idade}/MEIA_VIDA)*EXP(-0.5*({circ}/SIGMA_K)^2)"
            ws.cell(row=r, column=2 + m,
                    value=f"=IFERROR(SUMPRODUCT({w}*M_DESV_NUM)/SUMPRODUCT({w}),\"\")")
        # AVERAGE(EXP(intervalo)) exige matriz; SUMPRODUCT forca contexto de matriz
        ws.cell(row=r, column=15, value=f"=SUMPRODUCT(EXP(C{r}:N{r}))/12")
        num = "+".join(f"{h}*EXP({get_column_letter(2+m)}{r})"
                       for m, h in zip(meses_prod, horas_prod))
        ws.cell(row=r, column=16, value=f"=({num})/{sum(horas_prod)}/O{r}")
        ws.cell(row=r, column=17, value=f"=EXP(B{r})*P{r}")
        ws.cell(row=r, column=18, value=f"=LN(Q{r})")
        if i == 0:
            ws.cell(row=r, column=19, value='=""')
            ws.cell(row=r, column=20, value='=""')
        elif i == 1:
            ws.cell(row=r, column=19, value=f"=R{r}-R{r-1}")
            ws.cell(row=r, column=20, value=f"=S{r}^2")
        else:
            ws.cell(row=r, column=19, value=f"=R{r}-R{r-1}")
            ws.cell(row=r, column=20, value=f"=LAMBDA_FWD*T{r-1}+(1-LAMBDA_FWD)*S{r}^2")
        ws.cell(row=r, column=21,
                value=(f'=IF(T{r-1}="","",SQRT(T{r-1}))' if i > 1 else '=""'))
        ws.cell(row=r, column=22, value=f'=IF(OR(U{r}="",U{r}=0,S{r}=""),"",S{r}/U{r})')
        for c in (2, 15, 16, 17, 18, 19, 21, 22):
            ws.cell(row=r, column=c).number_format = "0.0000"
        ws.cell(row=r, column=17).number_format = "#,##0.00"
    n_f = prim + len(linhas_or) - 1
    _nome(wb, "FWD_RET", f"CALC_FWD!$S${prim+1}:$S${n_f}")
    _nome(wb, "FWD_VAR_ULT", f"CALC_FWD!$T${n_f}")
    _nome(wb, "FWD_FV", f"CALC_FWD!$Q${n_f}")
    b = n_f + 2
    L_OBS, L_VOL, L_VOLH, L_Z, L_VAR = b, b + 1, b + 2, b + 3, b + 4
    L_ZES, L_ES, L_CH1, L_CH2, L_RAZ, L_KURT = b + 5, b + 6, b + 7, b + 8, b + 9, b + 10
    blocos = [
        (L_OBS, "Observacoes de retorno do termo", "=COUNT(FWD_RET)", "0"),
        (L_VOL, "Vol EWMA mensal prevista", "=SQRT(FWD_VAR_ULT)", "0.0000"),
        (L_VOLH, "Vol no horizonte", f"=B{L_VOL}*SQRT(HORIZ_M)", "0.0000"),
        (L_Z, "Quantil normal (1-CONF)", "=NORMSINV(1-CONF)", "0.0000"),
        (L_VAR, "VaR de preco (delta-normal)", f"=ABS(FV_FLAT*(EXP(B{L_Z}*B{L_VOLH})-1))", "#,##0.00"),
        (L_ZES, "z do Expected Shortfall", f"=-EXP(-(B{L_Z}^2)/2)/SQRT(2*PI())/(1-CONF)", "0.0000"),
        (L_ES, "ES de preco (delta-normal)", f"=ABS(FV_FLAT*(EXP(B{L_ZES}*B{L_VOLH})-1))", "#,##0.00"),
        (L_CH1, "CHALLENGER 1 - quantil historico do termo",
         "=ABS(FV_FLAT*(EXP(PERCENTILE(FWD_RET,1-CONF)*SQRT(HORIZ_M))-1))", "#,##0.00"),
        (L_CH2, "CHALLENGER 2 - FHS sobre PLD spot (teto)", "=VAR_SPOT", "#,##0.00"),
        (L_RAZ, "Razao spot / termo", f"=IFERROR(B{L_CH2}/B{L_VAR},\"\")", "0.0"),
        (L_KURT, "Curtose dos retornos do termo",
         "=SUMPRODUCT((FWD_RET-AVERAGE(FWD_RET))^4)/COUNT(FWD_RET)/STDEVP(FWD_RET)^4", "0.00"),
    ]
    for lin_, rot, form, fmt in blocos:
        _n(ws, lin_, 1, rot, bold=True)
        cel = ws.cell(row=lin_, column=2, value=form)
        cel.number_format = fmt
        cel.fill = PatternFill("solid", fgColor=F_CALC)
    _nome(wb, "VOL_FWD", f"CALC_FWD!$B${L_VOL}")
    _nome(wb, "VAR_PRECO", f"CALC_FWD!$B${L_VAR}")
    _nome(wb, "ES_PRECO", f"CALC_FWD!$B${L_ES}")
    _n(ws, L_KURT + 2, 1,
       "PISO, NAO MEDIDA EXATA: o nivel do modelo e uma media exponencial do spot "
       "realizado, enquanto o termo de verdade reprecifica com expectativa e se move "
       "mais. A vol aqui e um limite inferior. Por isso o tamanho e cruzado com o teto "
       "de liquidez e com a perda no cenario adverso, e o VaR sobre spot aparece como teto.",
       bold=True)
    _n(ws, L_KURT + 3, 1,
       "Por que delta-normal e nao historico: a serie do termo e mensal e curta "
       "(dezenas de observacoes). Estimar uma volatilidade e mais confiavel do que "
       "estimar diretamente um quantil de 5%. O historico fica como challenger e a "
       "diferenca entre os dois mede o custo da hipotese de normalidade.", bold=True)
    ch = LineChart(); ch.title = "Forward reconstruido (R$/MWh)"; ch.height, ch.width = 7, 16
    ch.add_data(Reference(ws, min_col=17, min_row=5, max_row=n_f), titles_from_data=True)
    ch.set_categories(Reference(ws, min_col=1, min_row=prim, max_row=n_f))
    ws.add_chart(ch, "X6")

    # =================================================== 8. CURVA
    ws = wb.create_sheet("CURVA")
    _titulo(ws, "CURVA PUBLICA DE REFERENCIA — FAIR VALUE MENSAL",
            "fair_value = nivel_na_referencia * fator_sazonal, truncado por piso e teto da ANEEL.")
    tem_anc = bool(meta.get("tem_ancora"))
    seco_est = bool(meta.get("seco_estatistico"))
    _hdr(ws, ["mes_ref", "horas_mes", "mes_cal", "fator_sazonal", "nivel_ref",
              "estatistico", "ANCORA InfoPLD", "peso ancora", "nowcast %",
              "base combinada", "FAIR_VALUE", "trava", "SECO", "UMIDO", "ESPERADO"],
         linha=4, larg=14)
    alvo = meta["alvo"]
    for i, m in enumerate(alvo):
        r = 5 + i
        k = i + 1                       # posicao no vetor de meses do InfoPLD
        _n(ws, r, 1, pd.Timestamp(m).date(), "yyyy-mm")
        ws.cell(row=r, column=2, value=f"=DAY(EOMONTH(A{r},0))*24")
        ws.cell(row=r, column=3, value=f"=MONTH(A{r})")
        ws.cell(row=r, column=4, value=f"=INDEX(FATOR_MES,C{r})")
        ws.cell(row=r, column=5, value="=NIVEL_REF")
        ws.cell(row=r, column=6, value=f"=EXP(E{r})*D{r}")            # estatistico
        if tem_anc:
            ws.cell(row=r, column=7, value=f"=INDEX(IP_ESPERADO,{k})")
            ws.cell(row=r, column=8, value="=PESO_ANCORA")
            ws.cell(row=r, column=9, value=f"=INDEX(NOWCAST_AJUSTE,{k})")
            # O premio a termo entra AQUI, e o modo e chaveavel na propria planilha:
            #   visao_propria       -> so o NIVEL (a forma da curva continua do modelo)
            #   mercado_consistente -> o premio de cada mes (curva replica o mercado)
            #   sem_premio          -> fair value fundamental puro
            prem = ('IF(PREMIO_MODO="visao_propria",PREMIO_NIVEL,'
                    f'IF(PREMIO_MODO="mercado_consistente",INDEX(PREMIO_MES,{k}),0))'
                    ) if meta.get("tem_premio") else "PREMIO_BASIS"
            ws.cell(row=r, column=10,
                    value=f"=(H{r}*G{r}+(1-H{r})*F{r})*(1+I{r}/100)+{prem}")
        else:
            ws.cell(row=r, column=7, value=None)
            ws.cell(row=r, column=8, value=0)
            ws.cell(row=r, column=9, value=0)
            ws.cell(row=r, column=10, value=f"=F{r}")
        ws.cell(row=r, column=11, value=f"=MEDIAN(PISO,J{r},TETO)")
        ws.cell(row=r, column=12, value=f'=IF(K{r}>J{r},"piso",IF(K{r}<J{r},"teto",""))')
        if tem_anc and not seco_est:
            ws.cell(row=r, column=13, value=f"=K{r}*INDEX(IP_SECO,{k})/INDEX(IP_ESPERADO,{k})")
        else:
            ws.cell(row=r, column=13, value=f"=K{r}*K_SECO")
        if tem_anc:
            ws.cell(row=r, column=14, value=f"=K{r}*INDEX(IP_UMIDO,{k})/INDEX(IP_ESPERADO,{k})")
        else:
            ws.cell(row=r, column=14, value=f"=K{r}*K_UMIDO")
        ws.cell(row=r, column=15,
                value=f"=K{r}*PROB_BASE+M{r}*PROB_SECO+N{r}*PROB_UMIDO")
        for c in (4, 6, 7, 10, 11, 13, 14, 15):
            ws.cell(row=r, column=c).number_format = "#,##0.00"
        ws.cell(row=r, column=9).number_format = "0.000"
    n_c = 4 + len(alvo)
    _nome(wb, "C_MES", f"CURVA!$A$5:$A${n_c}")
    _nome(wb, "C_HORAS", f"CURVA!$B$5:$B${n_c}")
    _nome(wb, "C_FV", f"CURVA!$K$5:$K${n_c}")
    _nome(wb, "C_SECO", f"CURVA!$M$5:$M${n_c}")
    _nome(wb, "C_UMIDO", f"CURVA!$N$5:$N${n_c}")
    lr = n_c + 2
    _n(ws, lr, 1, "Nivel na referencia (ln, EWMA dessazonalizado)", bold=True)
    ws.cell(row=lr, column=2,
            value="=INDEX(M_NIVEL_EWMA,MATCH(REF_MES,M_MESREF,0))").number_format = "0.0000"
    _nome(wb, "NIVEL_REF", f"CURVA!$B${lr}")
    for k, (rot, form) in enumerate([
        ("FAIR VALUE FLAT do periodo (pond. por horas)", "=SUMPRODUCT(C_FV,C_HORAS)/SUM(C_HORAS)"),
        ("Preco medio cenario SECO", "=SUMPRODUCT(C_SECO,C_HORAS)/SUM(C_HORAS)"),
        ("Preco medio cenario UMIDO", "=SUMPRODUCT(C_UMIDO,C_HORAS)/SUM(C_HORAS)"),
        ("Horas totais do produto", "=SUM(C_HORAS)"),
    ], start=1):
        _n(ws, lr + k, 1, rot, bold=True)
        c = ws.cell(row=lr + k, column=2, value=form)
        c.number_format = "#,##0.00"; c.fill = PatternFill("solid", fgColor=F_FV)
    _nome(wb, "FV_FLAT", f"CURVA!$B${lr+1}")
    _nome(wb, "PV_SECO", f"CURVA!$B${lr+2}")
    _nome(wb, "PV_UMIDO", f"CURVA!$B${lr+3}")
    _nome(wb, "HORAS_TOT", f"CURVA!$B${lr+4}")

    ch = LineChart(); ch.title = "Curva mensal Base / Seco / Umido"; ch.height, ch.width = 8, 18
    ch.add_data(Reference(ws, min_col=11, max_col=11, min_row=4, max_row=n_c), titles_from_data=True)
    ch.add_data(Reference(ws, min_col=13, max_col=14, min_row=4, max_row=n_c), titles_from_data=True)
    ch.set_categories(Reference(ws, min_col=1, min_row=5, max_row=n_c))
    ws.add_chart(ch, "M5")

    # =================================================== 9. POSICAO
    ws = wb.create_sheet("POSICAO")
    _titulo(ws, "DIRECAO, PRECO LIMITE E DIMENSIONAMENTO",
            "Tudo formula. Mexer no edge minimo ou no orcamento redimensiona a posicao.")
    linhas = [
        ("Proxy publica de transacao (MVE) — preco medio ponderado",
         '=IFERROR(SUMPRODUCT(MVE_PRECO,MVE_MONT)/SUM(MVE_MONT),IFERROR(AVERAGE(MVE_PRECO),""))', "#,##0.00", F_PROXY),
        ("Negocios comparaveis no MVE", "=COUNT(MVE_PRECO)", "0", F_PROXY),
        ("Referencia alternativa observavel (PLD flat dos 3 ultimos meses)",
         "=AVERAGE(INDEX(M_FLAT,MATCH(REF_MES,M_MESREF,0)-2):INDEX(M_FLAT,MATCH(REF_MES,M_MESREF,0)))",
         "#,##0.00", F_OBS),
        ("FAIR VALUE FLAT", "=FV_FLAT", "#,##0.00", F_FV),
        ("Comprar ate", "=FV_FLAT-EDGE_MIN", "#,##0.00", F_CALC),
        ("Vender acima de", "=FV_FLAT+EDGE_MIN", "#,##0.00", F_CALC),
        ("Tipo de recomendacao",
         '=IF(B6=0,"CONDICIONAL",IF(ABS(FV_FLAT-B5)<EDGE_MIN,"SEM_POSICAO","DIRECIONAL"))', None, F_CALC),
        ("Direcao (+1 comprado, -1 vendido, 0 fora)",
         '=IF(B11="SEM_POSICAO",0,IF(B11="DIRECIONAL",IF(FV_FLAT>B5,1,-1),IF(FV_FLAT>B7,1,-1)))', "0", F_CALC),
        ("Preco de entrada de referencia",
         '=IF(B12=0,"",IF(B11="DIRECIONAL",B5,IF(B12>0,B9,B10)))', "#,##0.00", F_CALC),
        ("Edge sobre a entrada", '=IF(B13="","",FV_FLAT-B13)', "#,##0.00", F_CALC),
        ("Orcamento de VaR", "=LIMITE_VAR*FRAC_ORC", "#,##0.00", F_CALC),
        ("[1] MWm que o VaR permite", "=IF(VAR_PRECO=0,0,FLOOR(B15/(VAR_PRECO*HORAS_TOT),1))",
         "#,##0", F_CALC),
        ("[2] MWm que a LIQUIDEZ permite", "=MWM_TETO", "#,##0", F_CALC),
        ("Perda por MWm no cenario adverso (ate a entrega)",
         "=MAX(ABS(PV_UMIDO-(FV_FLAT-EDGE_MIN)),ABS(PV_SECO-(FV_FLAT+EDGE_MIN)))*HORAS_TOT",
         "#,##0", F_CALC),
        ("[3] MWm que o CENARIO ADVERSO permite",
         "=IF(B18=0,0,FLOOR(LIMITE_VAR*FRAC_STRESS/B18,1))", "#,##0", F_CALC),
        ("RESTRICAO VINCULANTE",
         '=INDEX({"VaR","liquidez","cenario adverso"},MATCH(MIN(B16,B17,B19),CHOOSE({1;2;3},B16,B17,B19),0))',
         None, F_CALC),
        ("MWm DA POSICAO", "=IF(B12=0,0,MIN(B16,B17,B19))", "#,##0", F_PREM),
        ("Energia do periodo (GWh)", "=B21*HORAS_TOT/1000", "#,##0", F_CALC),
        ("VaR da posicao (R$)", "=VAR_PRECO*B21*HORAS_TOT", "#,##0", F_CALC),
        ("ES da posicao (R$)", "=ES_PRECO*B21*HORAS_TOT", "#,##0", F_CALC),
        ("% do limite consumido", "=B23/LIMITE_VAR", "0.0%", F_CALC),
        ("Perda no cenario adverso (R$)", "=B18*B21", "#,##0", F_CALC),
        ("% do limite no cenario adverso", "=B25/LIMITE_VAR", "0.0%", F_CALC),
        ("VaR pelo challenger SPOT (R$)", "=VAR_SPOT*B21*HORAS_TOT", "#,##0", F_PROXY),
    ]
    for k, (rot, form, fmt, fill) in enumerate(linhas, start=5):
        _n(ws, k, 1, rot, bold=True)
        c = ws.cell(row=k, column=2, value=form)
        if fmt:
            c.number_format = fmt
        c.fill = PatternFill("solid", fgColor=fill)
    ws.column_dimensions["A"].width = 58
    ws.column_dimensions["B"].width = 20
    ws.conditional_formatting.add("B26", CellIsRule(
        operator="greaterThan", formula=["1"], fill=PatternFill("solid", fgColor="FFCDD2")))
    _nome(wb, "DIRECAO", "POSICAO!$B$12")
    _nome(wb, "P_ENTRADA", "POSICAO!$B$13")
    _nome(wb, "MWM", "POSICAO!$B$21")
    _nome(wb, "VAR_BRL", "POSICAO!$B$23")
    _nome(wb, "ES_BRL", "POSICAO!$B$24")
    _nome(wb, "PERDA_CEN", "POSICAO!$B$26")
    _nome(wb, "RESTRICAO", "POSICAO!$B$20")

    # =================================================== 10. PNL_VPL
    ws = wb.create_sheet("PNL_VPL")
    _titulo(ws, "PnL POR MES, CENARIOS E VPL",
            "PnL = direcao x MWm x horas x (preco_saida - preco_entrada). Desconto ate 31/12.")
    _hdr(ws, ["mes_ref", "horas", "MWm", "direcao", "P_entrada", "P_saida_BASE",
              "PnL_BASE", "PnL_SECO", "PnL_UMIDO", "fator_desc", "VP_BASE"], linha=4, larg=15)
    for i in range(len(alvo)):
        r, rc = 5 + i, 5 + i
        ws.cell(row=r, column=1, value=f"=CURVA!A{rc}"); ws.cell(row=r, column=1).number_format = "yyyy-mm"
        ws.cell(row=r, column=2, value=f"=CURVA!B{rc}")
        ws.cell(row=r, column=3, value="=MWM")
        ws.cell(row=r, column=4, value="=DIRECAO")
        ws.cell(row=r, column=5, value="=IF(P_ENTRADA=\"\",0,P_ENTRADA)")
        ws.cell(row=r, column=6, value=f"=CURVA!K{rc}")
        ws.cell(row=r, column=7, value=f"=D{r}*C{r}*B{r}*(F{r}-E{r})")
        ws.cell(row=r, column=8, value=f"=D{r}*C{r}*B{r}*(CURVA!M{rc}-E{r})")
        ws.cell(row=r, column=9, value=f"=D{r}*C{r}*B{r}*(CURVA!N{rc}-E{r})")
        ws.cell(row=r, column=10, value=f"=1/(1+TX_DESC)^((EOMONTH(A{r},0)-DATA_CORTE)/365)")
        ws.cell(row=r, column=11, value=f"=G{r}*J{r}")
        for c in (5, 6, 7, 8, 9, 11):
            ws.cell(row=r, column=c).number_format = "#,##0"
        ws.cell(row=r, column=10).number_format = "0.0000"
    n_p = 4 + len(alvo)
    t = n_p + 2
    for k, (rot, form, fmt) in enumerate([
        ("PnL total BASE", f"=SUM(G5:G{n_p})", "#,##0"),
        ("PnL total SECO", f"=SUM(H5:H{n_p})", "#,##0"),
        ("PnL total UMIDO", f"=SUM(I5:I{n_p})", "#,##0"),
        ("PnL ESPERADO (ponderado)", f"=B{t}*PROB_BASE+B{t+1}*PROB_SECO+B{t+2}*PROB_UMIDO", "#,##0"),
        ("VPL BASE ate 31/12", f"=SUM(K5:K{n_p})", "#,##0"),
        ("Retorno / VaR", f"=IFERROR(B{t+3}/VAR_BRL,\"\")", "0.00"),
        ("Retorno / ES", f"=IFERROR(B{t+3}/ES_BRL,\"\")", "0.00"),
        ("Pior cenario", f"=MIN(B{t}:B{t+2})", "#,##0"),
        ("Pior cenario cabe no VaR?", f'=IF(ABS(MIN(0,B{t+7}))<=VAR_BRL,"SIM","NAO - descasamento de horizonte")', None),
    ]):
        _n(ws, t + k, 1, rot, bold=True)
        c = ws.cell(row=t + k, column=2, value=form)
        if fmt:
            c.number_format = fmt
        c.fill = PatternFill("solid", fgColor=F_CALC)
    ws.column_dimensions["A"].width = 34
    _nome(wb, "PNL_ESP", f"PNL_VPL!$B${t+3}")
    _nome(wb, "VPL_BASE", f"PNL_VPL!$B${t+4}")

    # =================================================== 11. CHECKS
    ws = wb.create_sheet("CHECKS")
    _titulo(ws, "VALIDACOES AUTOMATICAS", "Recalculam junto com a planilha. Vermelho = investigar.")
    _hdr(ws, ["check", "resultado", "esperado", "status"], linha=4, larg=52)
    chk = [
        ("Media dos fatores sazonais = 1", "=AVERAGE(FATOR_MES)", "1,000", "=IF(ABS(B5-1)<0.0001,\"OK\",\"FALHA\")"),
        ("Soma dos fatores = 12", "=SUM(FATOR_MES)", "12,000", "=IF(ABS(B6-12)<0.001,\"OK\",\"FALHA\")"),
        ("Fatores NAO sao todos iguais (curva nao chapada)",
         "=MAX(FATOR_MES)-MIN(FATOR_MES)", "> 0,05",
         '=IF(MAX(FATOR_MES)-MIN(FATOR_MES)>0.05,"OK","FALHA - sazonalidade degenerada")'),
        ("Fair value varia entre os meses do horizonte",
         "=MAX(C_FV)-MIN(C_FV)", "> 0",
         '=IF(MAX(C_FV)-MIN(C_FV)>0.01,"OK","FALHA - curva chapada")'),
        ("Nenhum mes do horizonte fora de piso/teto",
         '=SUMPRODUCT(--(CURVA!L5:L' + str(n_c) + '<>""))', "0", '=IF(B7=0,"OK","ALERTA")'),
        ("Ordenacao Seco > Base > Umido", "=IF(AND(K_SECO>1,K_UMIDO<1),1,0)", "1",
         '=IF(B8=1,"OK","FALHA")'),
        ("VaR abaixo do teto de R$ 50 mi", "=VAR_BRL", "<= LIMITE_VAR",
         '=IF(VAR_BRL<=LIMITE_VAR,"OK","FALHA")'),
        ("ES pior ou igual ao VaR", "=ES_PRECO-VAR_PRECO", ">= 0",
         '=IF(ES_PRECO>=VAR_PRECO,"OK","FALHA")'),
        # A perda de cenario deixou de dimensionar quando o risco passou a ser
        # so o VaR. Ela continua monitorada, mas como ALERTA de stress: carregar
        # ate a entrega no cenario adverso e um evento de horizonte diferente do
        # limite de VaR, e tratar como FALHA travaria o book por uma metrica que
        # nao e a restricao declarada do mandato.
        ("Stress de carrego x limite (informativo)", "=PERDA_CEN/LIMITE_VAR",
         "<= FRAC_STRESS",
         '=IF(PERDA_CEN<=LIMITE_VAR*FRAC_STRESS,"OK","ALERTA - stress acima do '
         'confortavel; o limite vinculante continua sendo o VaR")'),
        ("VaR do challenger SPOT tambem cabe no limite", "=VAR_SPOT*MWM*HORAS_TOT/LIMITE_VAR",
         "<= 1", '=IF(VAR_SPOT*MWM*HORAS_TOT<=LIMITE_VAR,"OK","ALERTA - so passa no metodo do termo")'),
        ("Observacoes do backcast do forward", "=COUNT(FWD_RET)", ">= 24",
         '=IF(COUNT(FWD_RET)>=24,"OK","ALERTA")'),
        # trava de congruencia: o risco canonico tem de dominar os dois componentes
        # A invariante mudou junto com a definicao de risco. Antes o risco era
        # MAX(VaR, cenario) e o teste era de dominancia; agora o risco E o VaR,
        # entao o teste e de IGUALDADE. Manter o teste antigo faria a pasta
        # acusar FALHA justamente por estar correta.
        ("Risco do book = VaR de marcacao", "=RISCO_BOOK-VAR_MTM_BOOK", "= 0",
         '=IF(ABS(RISCO_BOOK-VAR_MTM_BOOK)<1,"OK","FALHA")'),
        ("Stress de carrego sobre o VaR (informativo)",
         "=IFERROR(PERDA_CEN_BOOK/VAR_MTM_BOOK,0)", "referencia",
         '=IF(IFERROR(PERDA_CEN_BOOK/VAR_MTM_BOOK,0)<=3,"OK",'
         '"ALERTA - cenario adverso muito acima do VaR de 1 mes")'),
        ("Risco do book dentro do limite", "=RISCO_BOOK/LIMITE_VAR", "<= 100%",
         '=IF(RISCO_BOOK<=LIMITE_VAR,"OK","FALHA")'),
        ("Pior cenario coberto pelo limite", "=BOOK_PIOR", ">= -limite",
         '=IF(ABS(MIN(0,BOOK_PIOR))<=LIMITE_VAR,"OK","ALERTA")'),
        ("Amostra sazonal na janela", "=SUM(M_JAN)", ">= 18",
         '=IF(SUM(M_JAN)>=18,"OK","ALERTA")'),
        ("Observacoes de z_h para o VaR", "=COUNT(Z_H)", ">= 250",
         '=IF(COUNT(Z_H)>=250,"OK","ALERTA")'),
        ("Cobertura minima nos meses fechados", "=MIN(CALC_MENSAL!N5:N" + str(n_m) + ")",
         ">= 0,98", '=IF(MIN(CALC_MENSAL!N5:N' + str(n_m) + ')>=0.98,"OK","ALERTA")'),
        ("Consistencia: PnL base = dir*MWm*horas*(FV-entrada)",
         f"=BOOK_PNL_CONV-SUM(BOOK_RISCO!I5:I{4+len(alvo)})", "~0",
         f'=IF(ABS(B15)<1,"OK","FALHA")'),
    ]
    for k, (rot, form, esp, st) in enumerate(chk, start=5):
        # Referencia de linha fixa dentro do status quebrava em silencio toda vez que
        # um check novo era inserido no meio da lista. Aqui ela e reescrita para a
        # linha corrente; referencias com nome de aba (PNL_VPL!B120) sao preservadas.
        st = re.sub(r"(?<![!$A-Za-z0-9_])B\d+\b", f"B{k}", st)
        _n(ws, k, 1, rot)
        ws.cell(row=k, column=2, value=form).number_format = "#,##0.0000"
        _n(ws, k, 3, esp)
        ws.cell(row=k, column=4, value=st)
    ws.conditional_formatting.add(f"D5:D{4+len(chk)}", CellIsRule(
        operator="equal", formula=['"OK"'], fill=PatternFill("solid", fgColor="C8E6C9")))
    ws.conditional_formatting.add(f"D5:D{4+len(chk)}", CellIsRule(
        operator="equal", formula=['"FALHA"'], fill=PatternFill("solid", fgColor="FFCDD2")))
    ws.conditional_formatting.add(f"D5:D{4+len(chk)}", CellIsRule(
        operator="equal", formula=['"ALERTA"'], fill=PatternFill("solid", fgColor="FFF9C4")))

    # =================================================== 12. MANIFESTO / FONTES
    ws = wb.create_sheet("MANIFESTO")
    _titulo(ws, "LINHAGEM DOS DADOS", "Proveniencia, hash e limitacoes de cada arquivo de origem.")
    cols = ["instituicao", "conjunto", "url_origem", "url_recurso", "baixado_em_utc",
            "sha256", "bytes", "frequencia_atualizacao", "qualidade_limitacoes"]
    _hdr(ws, cols, linha=4, larg=28)
    for i, r in enumerate(man.itertuples(index=False), start=5):
        for j, cn in enumerate(cols, 1):
            v = getattr(r, cn, "")
            _n(ws, i, j, str(v)[:250], fill=F_OBS)

    # =================================================== 12b. CROSSCHECK
    ws = wb.create_sheet("CROSSCHECK")
    _titulo(ws, "CONFERENCIA CRUZADA — PLANILHA x IMPLEMENTACAO DE REFERENCIA EM PYTHON",
            "A coluna Python e valor de referencia para auditoria. O calculo oficial e a formula.")
    _hdr(ws, ["grandeza", "Python (referencia)", "Excel (formula)", "delta", "delta %",
              "tolerancia", "status", "por que esta tolerancia"], linha=4, larg=22)
    for i, item in enumerate(meta.get("crosscheck", []), start=5):
        rot, val_py, form = item[0], item[1], item[2]
        tol = item[3] if len(item) > 3 else 0.005
        just = item[4] if len(item) > 4 else ""
        _n(ws, i, 1, rot, bold=True)
        _n(ws, i, 2, float(val_py), "#,##0.0000", F_CALC)
        ws.cell(row=i, column=3, value=form).number_format = "#,##0.0000"
        ws.cell(row=i, column=4, value=f"=C{i}-B{i}").number_format = "#,##0.0000"
        ws.cell(row=i, column=5, value=f'=IFERROR(ABS(D{i}/B{i}),"")').number_format = "0.00%"
        _n(ws, i, 6, tol, "0.00%", F_PREM)
        ws.cell(row=i, column=7, value=f'=IF(IFERROR(ABS(D{i}/B{i}),1)<=F{i},"OK","DIVERGENCIA")')
        _n(ws, i, 8, just)
    ncc = 4 + len(meta.get("crosscheck", []))
    ws.conditional_formatting.add(f"G5:G{ncc}", CellIsRule(
        operator="equal", formula=['"OK"'], fill=PatternFill("solid", fgColor="C8E6C9")))
    ws.conditional_formatting.add(f"G5:G{ncc}", CellIsRule(
        operator="equal", formula=['"DIVERGENCIA"'], fill=PatternFill("solid", fgColor="FFCDD2")))
    _n(ws, ncc + 2, 1,
       "Tolerancia de 0,5%. Divergencia acima disso significa que a formula e o Python "
       "deixaram de concordar — investigar antes de usar qualquer numero.", bold=True)
    ws.column_dimensions["A"].width = 40

    # =================================================== 12c. TESE
    ws = wb.create_sheet("TESE")
    _titulo(ws, "TESE — TEXTO MONTADO POR FORMULA",
            "Nenhuma frase tem numero digitado. Cada valor vem de celula calculada; "
            "mudar PARAMETROS reescreve a tese.")
    ws.column_dimensions["A"].width = 150
    linhas_tese = [
        ("1. TESE", None),
        (None, '="Curva publica de referencia para "&TEXT(MIN(C_MES),"mm/yyyy")&"-"&'
               'TEXT(MAX(C_MES),"mm/yyyy")&" em "&"' + meta["submercado"] + '"&" aponta fair value de R$ "&'
               'TEXT(FV_FLAT,"#,##0.00")&"/MWh, contra R$ "&'
               'TEXT(BOOK_ENTRADA,"#,##0.00")&"/MWh de entrada media do book. O lado nao e '
               'unico: o book esta "&BOOK_LADO&", com "&TEXT(COUNTIF(LADO_VERT,"VENDIDO"),"0")&'
               '" vertice(s) vendido(s) e "&TEXT(COUNTIF(LADO_VERT,"COMPRADO"),"0")&'
               '" comprado(s). Cada mes e decidido contra a propria projecao de PLD: '
               'forward acima da projecao vende, abaixo compra."'),
        ("2. DIMENSIONAMENTO", None),
        (None, '="Book de "&TEXT(BOOK_ENERGIA_MWH/1000,"#,##0.0")&" GWh e R$ "&TEXT(BOOK_NOTIONAL/1000000,"#,##0.0")&" mi de notional, em produtos MENSAIS com quantidade propria por mes — nao ha um MWm unico que descreva a posicao, e o shape faz parte da tese. Dimensionados vertice a '
               'vertice: MWm(mes) = orcamento de risco x conviccao(mes) / risco(mes). A '
               'restricao vinculante predominante foi "&BOOK_RESTRICAO&", e o risco de cada '
               'vertice e o maior entre a perda no cenario adverso do SEU lado e o VaR de '
               'marcacao. O detalhe por vertice esta em POSICAO_BOOK."'),
        ("3. RISCO", None),
        (None, '="VaR de "&TEXT(VAR_BRL/1000000,"#,##0.0")&" mi, ou "&TEXT(VAR_BRL/LIMITE_VAR,"0.0%")&'
               '" do teto de R$ 50 mi, a "&TEXT(CONF,"0%")&" em "&TEXT(HORIZ_M,"0")&'
               '" mes. Expected Shortfall de "&TEXT(ES_BRL/1000000,"#,##0.0")&" mi. O fator de risco '
               'e o preco A TERMO do proprio strip, remarcado a mercado em "&TEXT(COUNT(FWD_RET),"0")&'
               '" datas passadas, e nao o PLD spot: pelo spot o VaR seria "&'
               'TEXT(VAR_SPOT*MWM*HORAS_TOT/1000000,"#,##0.0")&" mi, "&'
               'TEXT(VAR_SPOT/VAR_PRECO,"0.0")&"x maior, porque spot reverte a media e tem piso e '
               'teto administrativos."'),
        ("4. RESULTADO ESPERADO (INTERVALO)", None),
        (None, '="Carrego ate a entrega: esperado R$ "&TEXT(BOOK_PNL_ESP,"#,##0")&'
               '", seco R$ "&TEXT(BOOK_PNL_SECO,"#,##0")&", umido R$ "&'
               'TEXT(BOOK_PNL_UMIDO,"#,##0")&". Convergencia do premio R$ "&'
               'TEXT(BOOK_PNL_CONV,"#,##0")&", com VPL ate 31/12 de R$ "&'
               'TEXT(BOOK_VPL,"#,##0")&". Pior cenario R$ "&TEXT(BOOK_PIOR,"#,##0")&'
               '", retorno sobre VaR de "&TEXT(BOOK_PNL_ESP/BOOK_VAR,"0.00")&"x."'),
        ("5. CENARIOS HIDROLOGICOS", None),
        (None, '="Regimes classificados por indice de ENA %MLT e EAR %max, cortes nos quantis "&'
               'TEXT(Q_SECO,"0%")&" e "&TEXT(Q_UMIDO,"0%")&". Multiplicadores estimados: seco "&'
               'TEXT(K_SECO,"0.000")&" ("&TEXT(CALC_REGIME!B' + "{L_NSECO}" + ',"0")&" meses) e umido "&'
               'TEXT(K_UMIDO,"0.000")&" ("&TEXT(CALC_REGIME!B' + "{L_NUMIDO}" + ',"0")&" meses). '
               'Ordenacao Seco > Base > Umido: "&IF(AND(K_SECO>1,K_UMIDO<1),"coerente",'
               '"INVERTIDA - investigar")&"."'),
        ("6. GATILHOS E O QUE INVALIDA", None),
        (None, '="Sair se o preco a termo se mover mais que R$ "&TEXT(VAR_PRECO,"#,##0.00")&'
               '"/MWh contra a posicao em um mes, ou se o indice hidrologico cruzar "&'
               'TEXT(LIM_SECO,"0.00")&" (seco) ou "&TEXT(LIM_UMIDO,"0.00")&'
               '" (umido). Invalida a tese: ordenacao de cenarios deixar de valer, '
               'surgir preco publico comparavel que contradiga o fair value, ou a '
               'sazonalidade degenerar (amplitude atual de "&'
               'TEXT(MAX(FATOR_MES)-MIN(FATOR_MES),"0.00")&")."'),
        ("7. LIMITACOES", None),
        (None, '="Referencia PUBLICA de fair value, nao curva de mercado observada. '
               'Amostra sazonal de "&TEXT(SUM(M_JAN),"0")&" meses e "&TEXT(COUNT(FWD_RET),"0")&'
               '" remarcacoes de forward. O premio de risco a termo nao e observavel em '
               'fonte publica: risco de base declarado. Status do run: "&PAINEL!B2&"."'),
    ]
    _L = {"L_PNL_BASE": t, "L_PNL_SECO": t + 1, "L_PNL_UMIDO": t + 2,
          "L_NSECO": lin + 7, "L_NUMIDO": lin + 9}
    rr = 4
    for tit, form in linhas_tese:
        if tit:
            _n(ws, rr, 1, tit, bold=True)
            ws.cell(row=rr, column=1).font = Font(bold=True, size=11, color=ROSA)
        else:
            ws.cell(row=rr, column=1, value=form.format(**_L))
            ws.cell(row=rr, column=1).alignment = Alignment(wrap_text=True, vertical="top")
            ws.row_dimensions[rr].height = 46
        rr += 1

    # =================================================== 13. PAINEL
    ws = wb.create_sheet("PAINEL")
    wb.move_sheet("PAINEL", offset=-len(wb.sheetnames) + 1)
    ws.sheet_view.showGridLines = False
    ws["A1"] = meta["nome_curva"]
    ws["A1"].font = Font(bold=True, size=16, color=ROXO)
    ws["A2"] = f"Status: {meta['status']}   |   Data de corte: {meta['data_corte']}   |   Submercado: {meta['submercado']}"
    ws["A2"].font = Font(size=10, color=ROSA, bold=True)
    ws["A3"] = meta.get("status_motivo", "")
    ws["A3"].font = Font(size=9, italic=True, color="B00020")

    ws["A5"] = "COMO ATUALIZAR"
    ws["A5"].font = Font(bold=True, size=12, color=ROSA)
    for k, txt in enumerate([
        "1.  Botao ATUALIZAR DADOS (macro) — roda o Python, baixa CCEE/ONS e reescreve as abas DADOS_*.",
        "2.  Sem macro: rode  make sync && make excel  no terminal e reabra o arquivo.",
        "3.  Para reprecificar SEM baixar dado novo, basta editar PARAMETROS. Tudo recalcula sozinho.",
    ], start=6):
        ws.cell(row=k, column=1, value=txt).font = Font(size=10)

    ws["A10"] = "RESULTADO"
    ws["A10"].font = Font(bold=True, size=12, color=ROSA)
    painel = [
        ("Fair value flat do periodo", "=FV_FLAT", "#,##0.00", "R$/MWh"),
        # O painel resume o BOOK, nao uma posicao unica: nao existe mais um
        # "comprar ate" flat, porque cada vertice tem o seu lado e o seu nivel.
        ("Recomendacao",
         '=IF(BOOK_LADO="MISTO","book misto — lado por vertice em POSICAO_BOOK",'
         'BOOK_LADO&" em todos os vertices com sinal")', None, ""),
        ("Lado do book", "=BOOK_LADO", None, ""),
        ("Vertices vendidos", '=COUNTIF(LADO_VERT,"VENDIDO")', "0", "vertices"),
        ("Vertices comprados", '=COUNTIF(LADO_VERT,"COMPRADO")', "0", "vertices"),
        # strip de produtos mensais com quantidade diferente por mes nao tem
        # um MWm unico: o tamanho e o ladder (POSICAO_BOOK), resumido aqui em
        # energia e notional. As leituras em MWm ficam abaixo, rotuladas.
        ("TAMANHO — energia", "=BOOK_ENERGIA_MWH/1000", "#,##0.0", "GWh"),
        ("TAMANHO — notional", "=BOOK_NOTIONAL", "#,##0", "R$"),
        ("  equivalente flat", "=BOOK_MWM_FLAT", "#,##0.0",
         "MWm — so p/ comparar com produto flat"),
        ("  soma de MWm das pernas", "=BOOK_SOMA_MWM", "#,##0",
         "MWm — diagnostico, nao e tamanho"),
        ("Preco medio de entrada", "=BOOK_ENTRADA", "#,##0.00", "R$/MWh"),
        ("Restricao que dimensionou", "=BOOK_RESTRICAO", None, ""),
        # UMA hierarquia de risco. As duas primeiras linhas sao COMPONENTES; a
        # terceira e a unica que consome o limite. Antes o painel mostrava a
        # marcacao como "VaR do book" e o % do limite calculado sobre ela,
        # enquanto POSICAO_BOOK usava o risco cheio — dois numeros, mesmo rotulo.
        ("  componente A: VaR de marcacao 1 mes", "=VAR_MTM_BOOK", "#,##0", "R$"),
        ("  componente B: perda no cenario adverso", "=PERDA_CEN_BOOK", "#,##0", "R$"),
        ("RISCO DO BOOK = (A), o VaR", "=RISCO_BOOK", "#,##0", "R$"),
        ("% do limite de R$ 50 mi", "=RISCO_BOOK/LIMITE_VAR", "0.0%", ""),
        ("Folga ate o stop", "=LIMITE_VAR-RISCO_BOOK", "#,##0", "R$"),
        ("Expected Shortfall da marcacao", "=BOOK_ES", "#,##0", "R$"),
        ("PnL esperado (carrego)", "=BOOK_PNL_ESP", "#,##0", "R$"),
        ("PnL da convergencia do premio", "=BOOK_PNL_CONV", "#,##0", "R$"),
        ("Pior cenario", "=BOOK_PIOR", "#,##0", "R$"),
        ("VPL ate 31/12", "=BOOK_VPL", "#,##0", "R$"),
        ("Retorno / risco (carrego esperado)",
         "=IFERROR(BOOK_PNL_ESP/RISCO_BOOK,\"\")", "0.00", "x"),
        ("k_seco  /  k_umido", '=TEXT(K_SECO,"0.000")&"  /  "&TEXT(K_UMIDO,"0.000")', None, ""),
        ("Meia-vida em uso", "=MEIA_VIDA", "0", "dias"),
        ("Checks com FALHA", '=COUNTIF(CHECKS!D5:D25,"FALHA")', "0", ""),
        ("Checks com ALERTA", '=COUNTIF(CHECKS!D5:D25,"ALERTA")', "-", ""),
        ("Divergencias planilha x Python", '=COUNTIF(CROSSCHECK!G5:G20,"DIVERGENCIA")', "0", ""),
    ]
    for k, (rot, form, fmt, un) in enumerate(painel, start=11):
        _n(ws, k, 1, rot, bold=True)
        c = ws.cell(row=k, column=2, value=form)
        if fmt:
            c.number_format = fmt
        c.fill = PatternFill("solid", fgColor=F_FV)
        c.border = BORDA
        _n(ws, k, 3, un)
    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 10
    leg = 11 + len(painel) + 1
    _n(ws, leg, 1, "LEGENDA DE ROTULOS", bold=True)
    ws.cell(row=leg, column=1).font = Font(bold=True, size=11, color=ROSA)
    for k, (rot, cor, desc) in enumerate([
        ("OBSERVADO", F_OBS, "publicado por fonte oficial, sem transformacao de valor"),
        ("CALCULADO", F_CALC, "formula deterministica sobre observado"),
        ("PREMISSA", F_PREM, "escolha declarada — editavel em PARAMETROS"),
        ("PROXY", F_PROXY, "preco publico de transacao (MVE), com risco de base"),
        ("FAIR_VALUE", F_FV, "estimativa do modelo"),
    ], start=leg + 1):
        _n(ws, k, 1, rot, fill=cor, bold=True)
        _n(ws, k, 2, desc)
    nota = leg + 7
    ws.cell(row=nota, column=1, value=(
        "Nenhum resultado desta pasta e valor colado. As unicas celulas com valor sao as "
        "abas DADOS_* (serie observada) e o MANIFESTO. Todo o resto e formula nativa."
    )).font = Font(italic=True, size=9)

    # VaR a termo pelo NIVEL — criado por ultimo porque reponta VAR_PRECO/ES_PRECO
    # do backcast suavizado para o metodo principal.
    # So meses COMPLETOS entram na estimacao: o mes corrente tem cobertura parcial
    # e o retorno Jul->Ago(parcial) e artificial. Como a EWMA da o maior peso a
    # ultima observacao, incluir o mes parcial inflava a vol de 16,0% para 17,2%.
    ult = pd.Timestamp(pld_d.data.max())
    n_completo = len(meses) - (1 if ult.date() != (ult + pd.offsets.MonthEnd(0)).date() else 0)
    aba_var_nivel(wb, alvo, n_completo, prem)

    caminho = Path(caminho)
    wb.save(caminho)
    return {"arquivo": str(caminho), "abas": wb.sheetnames, "linhas_pld": len(pld_d),
            "meses": len(meses), "horizonte": len(alvo)}


def aba_var_nivel(wb, alvo, n_flat: int, corte) -> dict:
    """CALC_VAR_NIVEL — o VaR a termo, calculado na planilha, sem valor colado.

    Substitui o VaR medido sobre o backcast do proprio modelo. Aquele passava o
    nivel por uma EWMA de 6 meses antes de medir a volatilidade, entao media o
    suavizador e nao o mercado: dava R$ 6/MWh sobre um preco de R$ 244, com
    cenarios publicos abrindo de R$ 70 a R$ 302 no mesmo vertice.
    """
    ws = wb.create_sheet("CALC_VAR_NIVEL")
    ws["A1"] = "VaR A TERMO — VOLATILIDADE DO NIVEL COM AMORTECIMENTO ATE A ENTREGA"
    ws["A1"].font = Font(bold=True, size=12, color=ROSA)
    ws["A2"] = ("nivel(m) = flat(m) / fator sazonal do mes. O comprador de Out/26 nao corre "
                "risco da sazonalidade de outubro — ela e conhecida e ja esta no preco. "
                "Corre risco de o NIVEL se deslocar.")
    ws["A3"] = ("Amortecimento de Samuelson: um choque de hoje reverte parcialmente ate a "
                "entrega, entao vertice longo balanca menos que vertice curto. "
                "vol_fwd(T) = vol_nivel x EXP(-kappa x T), kappa do AR(1) do proprio nivel.")
    fim = 4 + n_flat
    _hdr(ws, ["mes_ref", "flat", "mes", "fator sazonal", "nivel dessaz",
              "ln(nivel)", "retorno", "ret^2", "EWMA var", "lag ln(nivel)"], 4, 15)
    for i in range(n_flat):
        r, rm = 5 + i, 5 + i
        ws.cell(row=r, column=1, value=f"=CALC_MENSAL!A{rm}").number_format = "yyyy-mm"
        ws.cell(row=r, column=2, value=f"=CALC_MENSAL!B{rm}").number_format = "#,##0.00"
        ws.cell(row=r, column=3, value=f"=MONTH(A{r})")
        ws.cell(row=r, column=4, value=f"=INDEX(FATOR_MES,C{r})").number_format = "0.0000"
        ws.cell(row=r, column=5, value=f"=IFERROR(B{r}/D{r},\"\")").number_format = "#,##0.00"
        ws.cell(row=r, column=6, value=f"=IFERROR(LN(E{r}),\"\")").number_format = "0.0000"
        if i:
            ws.cell(row=r, column=7, value=f"=IFERROR(F{r}-F{r-1},\"\")").number_format = "0.0000"
            ws.cell(row=r, column=8, value=f"=IFERROR(G{r}^2,\"\")").number_format = "0.000000"
            # recursao EWMA identica a do Python: var_i = lambda*var_(i-1) + (1-lambda)*r_i^2
            ws.cell(row=r, column=9, value=(f"=H{r}" if i == 1 else
                                            f"=LAMBDA_FWD*I{r-1}+(1-LAMBDA_FWD)*H{r}")
                    ).number_format = "0.000000"
            ws.cell(row=r, column=10, value=f"=F{r-1}").number_format = "0.0000"

    b = fim + 2
    itens = [
        ("Vol EWMA mensal do nivel", f"=SQRT(I{fim})", "0.00%"),
        ("Vol amostral mensal do nivel", f"=STDEV(G6:G{fim})", "0.00%"),
        ("Vol usada (maior das duas)", f"=MAX(C{b},C{b+1})", "0.00%"),
        ("AR(1) phi do nivel", f"=SLOPE(F6:F{fim},J6:J{fim})", "0.000"),
        ("kappa = -LN(phi)", f"=-LN(MIN(MAX(C{b+3},0.000001),0.999999))", "0.000"),
        ("Meia-vida do nivel (meses)", f"=LN(2)/C{b+4}", "0.0"),
        ("z do quantil", "=NORMSINV(CONF)", "0.000"),
        # -C^2 no Excel e lido como (-C)^2. Sem o parenteses o z do ES sai 30,86
        # em vez de 2,06 — o mesmo erro de precedencia ja corrigido em CALC_FWD.
        ("z do Expected Shortfall", f"=EXP(-(C{b+6}^2)/2)/SQRT(2*PI())/(1-CONF)", "0.000"),
    ]
    for k, (rot_i, f, fmt) in enumerate(itens):
        ws.cell(row=b + k, column=1, value=rot_i).font = Font(bold=True)
        c = ws.cell(row=b + k, column=3, value=f)
        c.number_format = fmt
        c.fill = PatternFill("solid", fgColor=F_CALC)
    _nome(wb, "VOL_NIVEL", f"CALC_VAR_NIVEL!$C${b+2}")
    _nome(wb, "KAPPA_NIVEL", f"CALC_VAR_NIVEL!$C${b+4}")
    _nome(wb, "Z_CONF", f"CALC_VAR_NIVEL!$C${b+6}")
    _nome(wb, "Z_ES", f"CALC_VAR_NIVEL!$C${b+7}")

    v0 = b + 10
    ws.cell(row=v0 - 1, column=1, value="ESTRUTURA A TERMO DE VOLATILIDADE E VaR POR VERTICE"
            ).font = Font(bold=True, size=11, color=ROSA)
    _hdr(ws, ["vertice", "fim da entrega", "T (meses)", "amortecimento",
              "vol do forward", "preco", "VaR R$/MWh", "ES R$/MWh", "horas"], v0, 15)
    n = len(alvo)
    for i, m in enumerate(alvo):
        r = v0 + 1 + i
        ws.cell(row=r, column=1, value=pd.Timestamp(m).strftime("%b/%y"))
        ws.cell(row=r, column=2, value=f"=EOMONTH(INDEX(C_MES,{i+1}),0)").number_format = "dd/mm/yyyy"
        ws.cell(row=r, column=3, value=f"=MAX((B{r}-DATA_CORTE)/30.44,0.5)").number_format = "0.0"
        ws.cell(row=r, column=4, value=f"=EXP(-KAPPA_NIVEL*C{r})").number_format = "0.00"
        ws.cell(row=r, column=5, value=f"=VOL_NIVEL*D{r}").number_format = "0.0%"
        ws.cell(row=r, column=6, value=f"=INDEX(C_FV,{i+1})").number_format = "#,##0.00"
        ws.cell(row=r, column=7, value=f"=F{r}*(1-EXP(-Z_CONF*E{r}))").number_format = "#,##0.00"
        ws.cell(row=r, column=8, value=f"=F{r}*(1-EXP(-Z_ES*E{r}))").number_format = "#,##0.00"
        ws.cell(row=r, column=9, value=f"=INDEX(C_HORAS,{i+1})").number_format = "#,##0"
    uv = v0 + n
    t = uv + 2
    ws.cell(row=t, column=1, value="VaR medio ponderado por horas (R$/MWh)").font = Font(bold=True)
    cv = ws.cell(row=t, column=3, value=f"=SUMPRODUCT(G{v0+1}:G{uv},I{v0+1}:I{uv})/SUM(I{v0+1}:I{uv})")
    cv.number_format = "#,##0.00"; cv.fill = PatternFill("solid", fgColor=F_CALC)
    ws.cell(row=t + 1, column=1, value="ES medio ponderado por horas (R$/MWh)").font = Font(bold=True)
    ce = ws.cell(row=t + 1, column=3, value=f"=SUMPRODUCT(H{v0+1}:H{uv},I{v0+1}:I{uv})/SUM(I{v0+1}:I{uv})")
    ce.number_format = "#,##0.00"; ce.fill = PatternFill("solid", fgColor=F_CALC)
    ws.cell(row=t + 3, column=1, value=(
        "PISO DECLARADO: o backcast do proprio modelo (CALC_FWD) mede a volatilidade do "
        "suavizador, nao a do mercado, e por isso e reportado como piso, nao como VaR."))
    ws.cell(row=t + 4, column=1, value=(
        "TETO DECLARADO: o VaR sobre PLD spot (CALC_VAR_SPOT) superestima o termo, "
        "porque spot reverte a media e tem piso e teto administrativos."))
    _nome(wb, "VAR_VERT", f"CALC_VAR_NIVEL!$G${v0+1}:$G${uv}")
    _nome(wb, "ES_VERT", f"CALC_VAR_NIVEL!$H${v0+1}:$H${uv}")
    # VAR_PRECO e ES_PRECO passam a apontar para o metodo principal
    _nome(wb, "VAR_PRECO", f"CALC_VAR_NIVEL!$C${t}")
    _nome(wb, "ES_PRECO", f"CALC_VAR_NIVEL!$C${t+1}")
    for col, w in (("A", 40), ("B", 15), ("C", 16), ("D", 15), ("E", 14),
                   ("F", 12), ("G", 13), ("H", 12), ("I", 10), ("J", 12)):
        ws.column_dimensions[col].width = w
    return {"aba": "CALC_VAR_NIVEL", "linha_var": t}
