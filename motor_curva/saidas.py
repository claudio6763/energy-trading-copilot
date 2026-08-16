"""Planilha aberta com formulas, graficos e texto da Entrega 2."""
from __future__ import annotations
import json
from pathlib import Path

import numpy as np
import pandas as pd

from .config import DIR_OUT, NOME_CURVA, Rotulo

VERDE = "E8F5E9"; AZUL = "E3F2FD"; AMAR = "FFF8E1"; ROSA = "FCE4EC"; CINZA = "ECEFF1"
LEGENDA = {
    "OBSERVADO": (VERDE, "dado publicado por fonte oficial"),
    "CALCULADO": (AZUL, "formula deterministica sobre observado"),
    "PREMISSA": (AMAR, "escolha declarada do analista"),
    "PROXY": (ROSA, "preco publico de transacao (MVE) - risco de base"),
    "FAIR_VALUE": (CINZA, "estimativa do modelo"),
}


def _cab(ws, cols, larg=18):
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    for j, c in enumerate(cols, 1):
        cell = ws.cell(row=1, column=j, value=str(c))
        cell.fill = PatternFill("solid", fgColor="1F1235")
        cell.font = Font(color="FFFFFF", bold=True, size=10)
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(j)].width = larg
    ws.freeze_panes = "A2"


def _tabela(ws, df, larg=18, rotulo=None):
    from openpyxl.styles import PatternFill
    _cab(ws, list(df.columns), larg)
    fill = PatternFill("solid", fgColor=LEGENDA[rotulo][0]) if rotulo else None
    for i, row in enumerate(df.itertuples(index=False), start=2):
        for j, v in enumerate(row, start=1):
            if isinstance(v, (np.integer,)):
                v = int(v)
            elif isinstance(v, (np.floating,)):
                v = None if not np.isfinite(v) else float(v)
            elif isinstance(v, pd.Timestamp):
                v = v.date()
            elif isinstance(v, (list, dict)):
                v = json.dumps(v, ensure_ascii=False)[:500]
            c = ws.cell(row=i, column=j, value=v)
            if fill:
                c.fill = fill


def gerar_planilha(caminho: Path, ctx: dict):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from openpyxl.chart import LineChart, Reference

    wb = Workbook()

    # ---------------- Fontes
    ws = wb.active; ws.title = "Fontes"
    ws["A1"] = NOME_CURVA
    ws["A1"].font = Font(bold=True, size=13, color="C4007A")
    ws["A2"] = f"Status: {ctx['status']}  |  Data de corte: {ctx['data_corte']}  |  Gerado: {ctx['gerado_em']}"
    ws["A4"] = "LEGENDA DE CORES"; ws["A4"].font = Font(bold=True)
    r = 5
    for k, (cor, desc) in LEGENDA.items():
        ws.cell(row=r, column=1, value=k).fill = PatternFill("solid", fgColor=cor)
        ws.cell(row=r, column=2, value=desc)
        r += 1
    ws.column_dimensions["A"].width = 22; ws.column_dimensions["B"].width = 60
    ws2 = wb.create_sheet("Fontes_Manifesto")
    _tabela(ws2, ctx["manifesto"], 26, "OBSERVADO")

    # ---------------- Dados brutos resumo
    _tabela(wb.create_sheet("Dados_Brutos_Resumo"), ctx["resumo_dados"], 20, "OBSERVADO")

    # ---------------- Sazonalidade (inputs do calculo)
    _tabela(wb.create_sheet("Sazonalidade"), ctx["fatores"], 16, "CALCULADO")

    # ---------------- Curva_Base com FORMULAS
    ws = wb.create_sheet("Curva_Base")
    cols = ["mes_ref", "horas_mes", "fundamental_CMO", "sazonal_EWMA", "peso_w",
            "FAIR_VALUE (formula)", "limite_aplicado"]
    _cab(ws, cols, 20)
    cb = ctx["curva"]
    for i, row in enumerate(cb.itertuples(index=False), start=2):
        ws.cell(row=i, column=1, value=pd.Timestamp(row.mes_ref).date())
        ws.cell(row=i, column=2, value=int(row.horas_mes))
        ws.cell(row=i, column=3, value=None if not np.isfinite(row.fundamental) else float(row.fundamental))
        ws.cell(row=i, column=4, value=float(row.sazonal))
        ws.cell(row=i, column=5, value=float(row.peso))
        ws.cell(row=i, column=6, value=f"=IF(ISBLANK(C{i}),D{i},E{i}*C{i}+(1-E{i})*D{i})")
        ws.cell(row=i, column=7, value=row.limite or "")
        for j in (3, 4, 6):
            ws.cell(row=i, column=j).number_format = "#,##0.00"
    n = len(cb) + 1
    ch = LineChart(); ch.title = "Fair value mensal (R$/MWh)"
    ch.add_data(Reference(ws, min_col=6, min_row=1, max_row=n), titles_from_data=True)
    ch.set_categories(Reference(ws, min_col=1, min_row=2, max_row=n))
    ws.add_chart(ch, "I2")

    # ---------------- Cenarios com FORMULAS
    ws = wb.create_sheet("Cenarios")
    _cab(ws, ["mes_ref", "BASE (ref Curva_Base)", "k_seco", "k_umido",
              "SECO (formula)", "UMIDO (formula)", "prob_base", "prob_seco", "prob_umido",
              "ESPERADO (formula)"], 20)
    ks, ku = ctx["efeitos"]["k_seco"], ctx["efeitos"]["k_umido"]
    pb, ps, pu = (ctx["prob"]["base"], ctx["prob"]["seco"], ctx["prob"]["umido"])
    for i in range(2, len(cb) + 2):
        ws.cell(row=i, column=1, value=f"=Curva_Base!A{i}")
        ws.cell(row=i, column=2, value=f"=Curva_Base!F{i}")
        ws.cell(row=i, column=3, value=ks); ws.cell(row=i, column=4, value=ku)
        ws.cell(row=i, column=5, value=f"=B{i}*C{i}")
        ws.cell(row=i, column=6, value=f"=B{i}*D{i}")
        ws.cell(row=i, column=7, value=pb); ws.cell(row=i, column=8, value=ps); ws.cell(row=i, column=9, value=pu)
        ws.cell(row=i, column=10, value=f"=B{i}*G{i}+E{i}*H{i}+F{i}*I{i}")

    # ---------------- Posicao
    ws = wb.create_sheet("Posicao")
    _cab(ws, ["campo", "valor", "rotulo"], 30)
    for i, (k, v) in enumerate(ctx["posicao_tab"], start=2):
        ws.cell(row=i, column=1, value=k)
        ws.cell(row=i, column=2, value=v if not isinstance(v, (list, dict)) else json.dumps(v)[:300])
        ws.cell(row=i, column=3, value=ctx["posicao_rotulo"].get(k, ""))

    # ---------------- PnL_VPL com FORMULAS
    ws = wb.create_sheet("PnL_VPL")
    _cab(ws, ["mes_ref", "horas", "MWm", "direcao", "P_entrada", "P_saida (cenario base)",
              "PnL (formula)", "fator_desconto (formula)", "VP (formula)"], 20)
    mwm = ctx["posicao"].get("mwm", 0.0) or 0.0
    dirc = ctx["posicao"].get("direcao", 0) or 0
    pent = ctx["posicao"].get("preco_entrada", ctx["posicao"].get("comprar_ate", 0.0)) or 0.0
    taxa = ctx["taxa_desconto"]
    ws["L1"] = "taxa_desconto_aa"; ws["L2"] = taxa
    ws["L3"] = "hoje"; ws["L4"] = ctx["hoje"]
    for i in range(2, len(cb) + 2):
        ws.cell(row=i, column=1, value=f"=Curva_Base!A{i}")
        ws.cell(row=i, column=2, value=f"=Curva_Base!B{i}")
        ws.cell(row=i, column=3, value=mwm)
        ws.cell(row=i, column=4, value=dirc)
        ws.cell(row=i, column=5, value=pent)
        ws.cell(row=i, column=6, value=f"=Curva_Base!F{i}")
        ws.cell(row=i, column=7, value=f"=D{i}*C{i}*B{i}*(F{i}-E{i})")
        ws.cell(row=i, column=8, value=f"=1/(1+$L$2)^((EOMONTH(A{i},0)-$L$4)/365)")
        ws.cell(row=i, column=9, value=f"=G{i}*H{i}")
    ws.cell(row=len(cb) + 3, column=6, value="TOTAL PnL / VPL")
    ws.cell(row=len(cb) + 3, column=7, value=f"=SUM(G2:G{len(cb)+1})")
    ws.cell(row=len(cb) + 3, column=9, value=f"=SUM(I2:I{len(cb)+1})")

    # ---------------- VaR
    ws = wb.create_sheet("VaR")
    _cab(ws, ["campo", "valor", "unidade", "rotulo"], 26)
    for i, (k, v, u, rot) in enumerate(ctx["var_tab"], start=2):
        ws.cell(row=i, column=1, value=k)
        ws.cell(row=i, column=2, value=v)
        ws.cell(row=i, column=3, value=u)
        ws.cell(row=i, column=4, value=rot)
    b = len(ctx["var_tab"]) + 3
    ws.cell(row=b, column=1, value="VaR posicao (formula)")
    ws.cell(row=b, column=2, value=f"=Posicao!B{ctx['linha_mwm']}*{ctx['horas_total']}*B{ctx['linha_var_preco']}")
    ws.cell(row=b + 1, column=1, value="% do limite de R$ 50 mi (formula)")
    ws.cell(row=b + 1, column=2, value=f"=B{b}/50000000")
    ws.cell(row=b + 1, column=2).number_format = "0.0%"

    # ---------------- Backtests
    _tabela(wb.create_sheet("Backtests"), ctx["backtests"], 16, "CALCULADO")

    # ---------------- Premissas
    ws = wb.create_sheet("Premissas")
    _cab(ws, ["premissa", "valor", "justificativa"], 40)
    for i, (k, v, j) in enumerate(ctx["premissas_tab"], start=2):
        ws.cell(row=i, column=1, value=k)
        ws.cell(row=i, column=2, value=str(v))
        ws.cell(row=i, column=3, value=j)
        ws.cell(row=i, column=1).fill = PatternFill("solid", fgColor=AMAR)

    # ---------------- Checks
    _tabela(wb.create_sheet("Checks"), ctx["checks"], 30)

    wb.save(caminho)


# ------------------------------------------------------------------ graficos
def gerar_graficos(dirg: Path, ctx: dict):
    import matplotlib
    matplotlib.use("Agg")
    import matplotlib.pyplot as plt
    dirg.mkdir(parents=True, exist_ok=True)
    P = "#C4007A"; D = "#1F1235"
    cb, hist = ctx["curva"], ctx["flat"]

    fig, ax = plt.subplots(figsize=(9, 4.2))
    ax.plot(hist.mes_ref, hist.flat, color=D, lw=1.2, label="PLD flat mensal observado")
    ax.plot(cb.mes_ref, cb.fair_value, color=P, lw=2.2, label="Fair value (Base)")
    ax.fill_between(cb.mes_ref, cb.umido, cb.seco, color=P, alpha=.15, label="Umido-Seco")
    ax.set_ylabel("R\\$/MWh"); ax.legend(fontsize=8); ax.grid(alpha=.25)
    ax.set_title("Historico de PLD e curva projetada")
    fig.tight_layout(); fig.savefig(dirg / "01_curva_vs_historico.png", dpi=130); plt.close(fig)

    fig, ax = plt.subplots(1, 2, figsize=(11, 3.8))
    f = ctx["fatores"]
    ax[0].bar(f.mes, f.fator, color=P); ax[0].axhline(1, ls="--", color=D)
    ax[0].set_title("Fator sazonal EWMA de calendario"); ax[0].set_xticks(range(1, 13)); ax[0].grid(alpha=.25)
    s = ctx["sens_meia_vida"]
    for hl in s.meia_vida.unique():
        g = s[s.meia_vida == hl]
        ax[1].plot(g.mes, g.fator, marker="o", ms=3, label=f"{hl}d")
    ax[1].axhline(1, ls="--", color=D); ax[1].legend(fontsize=8, title="meia-vida")
    ax[1].set_title("Sensibilidade a meia-vida"); ax[1].set_xticks(range(1, 13)); ax[1].grid(alpha=.25)
    fig.tight_layout(); fig.savefig(dirg / "02_sazonalidade.png", dpi=130); plt.close(fig)

    if ctx.get("hidro") is not None and len(ctx["hidro"]):
        h = ctx["hidro"]
        fig, ax = plt.subplots(figsize=(9, 3.6))
        ax.plot(h.mes_ref, h.ena_pct, color=P, label="ENA %MLT")
        if h.ear_pct.notna().any():
            ax.plot(h.mes_ref, h.ear_pct, color=D, label="EAR %max")
        ax.axhline(100, ls="--", color="grey", lw=.8)
        ax.legend(fontsize=8); ax.grid(alpha=.25); ax.set_title("ENA e EAR contra referencia historica")
        fig.tight_layout(); fig.savefig(dirg / "03_hidrologia.png", dpi=130); plt.close(fig)

    if ctx.get("amostra_pnl") is not None:
        a = ctx["amostra_pnl"]
        fig, ax = plt.subplots(figsize=(8, 3.8))
        ax.hist(a / 1e6, bins=60, color=P, alpha=.75)
        ax.axvline(-ctx["var_brl"] / 1e6, color=D, ls="--", label=f"VaR 95% = R$ {ctx['var_brl']/1e6:.1f} mi")
        ax.axvline(-ctx["es_brl"] / 1e6, color="black", ls=":", label=f"ES = R$ {ctx['es_brl']/1e6:.1f} mi")
        ax.set_xlabel("PnL (R$ milhoes)"); ax.legend(fontsize=8); ax.grid(alpha=.25)
        ax.set_title("Distribuicao simulada de PnL")
        fig.tight_layout(); fig.savefig(dirg / "04_pnl_var.png", dpi=130); plt.close(fig)

    if ctx.get("sens_tamanho") is not None:
        st = ctx["sens_tamanho"]
        fig, ax = plt.subplots(figsize=(7, 3.6))
        ax.plot(st.mwm, st.var_brl / 1e6, color=P, lw=2)
        ax.axhline(50, color=D, ls="--", label="limite R$ 50 mi")
        ax.axhline(ctx["orcamento"] / 1e6, color="grey", ls=":", label="orcamento com buffer")
        ax.set_xlabel("MWm"); ax.set_ylabel("VaR (R$ mi)"); ax.legend(fontsize=8); ax.grid(alpha=.25)
        ax.set_title("Sensibilidade do tamanho da posicao")
        fig.tight_layout(); fig.savefig(dirg / "05_sensibilidade_tamanho.png", dpi=130); plt.close(fig)


# ------------------------------------------------------- graficos do book
def graficos_book(dirg, cb, sinal, pos, vpl_df, pnl_mes, ref_mkt, book_res):
    """Posicao, PnL por cenario e VPL acumulado do book."""
    import matplotlib
    matplotlib.use("Agg")
    import matplotlib.pyplot as plt
    import matplotlib.ticker as mtick
    from pathlib import Path
    dirg = Path(dirg); dirg.mkdir(parents=True, exist_ok=True)
    P, D, V, A = "#C4007A", "#1F1235", "#2E7D32", "#B00020"
    mes = [pd.Timestamp(m).strftime("%b/%y") for m in cb.mes_ref]

    # ---- 1. curva x mercado x sinal
    fig, ax = plt.subplots(2, 1, figsize=(11, 8), height_ratios=[2, 1], sharex=True)
    ax[0].plot(mes, cb.fair_value, "o-", color=P, lw=2.4, ms=6, label="Fair value do modelo")
    ax[0].plot(mes, ref_mkt.to_numpy(), "s--", color=D, lw=1.8, ms=5,
               label="Referência de mercado")
    ax[0].fill_between(mes, cb.umido, cb.seco, color=P, alpha=.12, label="Úmido–Seco")
    ax[0].set_ylabel("R\\$/MWh"); ax[0].legend(fontsize=9); ax[0].grid(alpha=.25)
    ax[0].set_title("Curva do modelo contra a referência de mercado", fontsize=12, color=D)

    res = sinal.residuo_vs_mercado.to_numpy()
    cores = [V if x > 0 else A for x in res]
    ax[1].bar(mes, res, color=cores, alpha=.85)
    lim = float(sinal.limiar.iloc[0])
    ax[1].axhline(lim, ls="--", lw=1, color=D); ax[1].axhline(-lim, ls="--", lw=1, color=D)
    ax[1].axhline(0, color="black", lw=.8)
    ax[1].set_ylabel("Resíduo R\\$/MWh"); ax[1].grid(alpha=.25)
    ax[1].set_title(f"Sinal: acima de ±{lim:.0f} R\\$/MWh vira ação", fontsize=10, color=D)
    for i, (x, a_) in enumerate(zip(res, sinal.acao)):
        if a_ != "FORA":
            ax[1].annotate(a_, (i, x), ha="center",
                           va="bottom" if x > 0 else "top", fontsize=8, weight="bold")
    fig.tight_layout(); fig.savefig(dirg / "06_curva_vs_mercado_e_sinal.png", dpi=140)
    plt.close(fig)

    # ---- 2. posicao e PnL
    fig, ax = plt.subplots(1, 2, figsize=(13, 4.6))
    m2 = [pd.Timestamp(m).strftime("%b/%y") for m in pos.mes_ref]
    cores = [V if x > 0 else A for x in pos.mwmed_liquido]
    ax[0].bar(m2, pos.mwmed_liquido, color=cores, alpha=.85)
    ax[0].axhline(0, color="black", lw=.8)
    for i, (x, p_) in enumerate(zip(pos.mwmed_liquido, pos.preco_entrada_medio)):
        ax[0].annotate(f"{x:.0f} MWm\n@ {p_:.2f}", (i, x), ha="center",
                       va="bottom" if x > 0 else "top", fontsize=8)
    ax[0].set_ylabel("MWmed líquido"); ax[0].grid(alpha=.25)
    ax[0].set_title("Posição líquida por mês, com preço de entrada", fontsize=11, color=D)

    larg = 0.26
    idx = np.arange(len(m2))
    for k, (cen, cor) in enumerate((("Esperado", P), ("Seco", "#EF6C00"), ("Umido", "#0277BD"))):
        ax[1].bar(idx + (k - 1) * larg, pnl_mes[f"pnl_{cen}"].to_numpy(), larg,
                  label=cen, color=cor, alpha=.9)
    ax[1].set_xticks(idx); ax[1].set_xticklabels(m2)
    ax[1].axhline(0, color="black", lw=.8)
    ax[1].yaxis.set_major_formatter(mtick.FuncFormatter(lambda v, p: f"{v/1e6:.1f}"))
    ax[1].set_ylabel("PnL (R\\$ milhões)"); ax[1].legend(fontsize=9); ax[1].grid(alpha=.25)
    ax[1].set_title("PnL por mês e cenário", fontsize=11, color=D)
    fig.tight_layout(); fig.savefig(dirg / "07_posicao_e_pnl.png", dpi=140); plt.close(fig)

    # ---- 3. VPL acumulado
    fig, ax = plt.subplots(figsize=(9, 4.4))
    m3 = [pd.Timestamp(m).strftime("%b/%y") for m in vpl_df.mes_ref]
    ax.bar(m3, vpl_df.vp, color=P, alpha=.55, label="VP do mês")
    ax.plot(m3, vpl_df.vp_acumulado, "o-", color=D, lw=2.2, ms=6, label="VPL acumulado")
    for i, v_ in enumerate(vpl_df.vp_acumulado):
        ax.annotate(f"{v_/1e6:.2f}", (i, v_), ha="center", va="bottom", fontsize=8.5)
    ax.yaxis.set_major_formatter(mtick.FuncFormatter(lambda v, p: f"{v/1e6:.1f}"))
    ax.set_ylabel("R\\$ milhões"); ax.legend(fontsize=9); ax.grid(alpha=.25)
    ax.axhline(0, color="black", lw=.8)
    tot = book_res["vpl"] / 1e6
    ax.set_title(f"VPL do book até 31/12  —  total R\\$ {tot:.2f} mi "
                 f"(VaR R\\$ {book_res['var_total']/1e6:.2f} mi, "
                 f"{book_res['consumo_limite']:.1%} do limite)", fontsize=11, color=D)
    fig.tight_layout(); fig.savefig(dirg / "08_vpl_book.png", dpi=140); plt.close(fig)
    return ["06_curva_vs_mercado_e_sinal.png", "07_posicao_e_pnl.png", "08_vpl_book.png"]

