# -*- coding: utf-8 -*-
"""Abas de referencia de mercado, premio, sinal e book multi-perna.

Tudo formula. Mudar o preco de entrada ou o MWmed de uma perna reprecifica o
book inteiro: exposicao liquida, PnL nos tres cenarios, VPL, VaR e consumo do
limite.

A matriz PRODUTO x MES e o que permite misturar mensais e trimestrais sem dupla
contagem: a exposicao de cada perna em cada mes e
    sinal x MWmed x pertence(produto, mes)
e a posicao liquida do mes e a soma da coluna.
"""
from __future__ import annotations

import pandas as pd
from openpyxl.chart import BarChart, LineChart, Reference, Series
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName

ROSA, ROXO = "C4007A", "1F1235"
F_IN, F_CALC, F_REF, F_OUT = "FFF8E1", "E3F2FD", "FCE4EC", "ECEFF1"


def _hdr(ws, cols, linha, larg=13):
    for j, c in enumerate(cols, 1):
        cel = ws.cell(row=linha, column=j, value=c)
        cel.fill = PatternFill("solid", fgColor=ROXO)
        cel.font = Font(color="FFFFFF", bold=True, size=9)
        cel.alignment = Alignment(horizontal="center", wrap_text=True, vertical="center")
        ws.column_dimensions[get_column_letter(j)].width = larg


def _tit(ws, t, sub=""):
    ws["A1"] = t
    ws["A1"].font = Font(bold=True, size=13, color=ROSA)
    if sub:
        ws["A2"] = sub
        ws["A2"].font = Font(italic=True, size=9, color="666677")


def _nome(wb, n, ref):
    if n in wb.defined_names:
        del wb.defined_names[n]
    wb.defined_names.add(DefinedName(n, attr_text=ref))


def adicionar(wb, alvo, ref_mercado: pd.Series, pernas: list, produtos: list[str],
              premio_modo: str, limiar_sinal: float, modo_sinal: str = "valor_relativo",
              k_premio: float = 0.0) -> dict:
    from .book import rotulo_mes, meses_do_produto, horas_do_mes
    n = len(alvo)
    rot = [rotulo_mes(m) for m in alvo]

    # ============================================== 1. REFERENCIA_MERCADO
    ws = wb.create_sheet("REFERENCIA_MERCADO")
    _tit(ws, "REFERENCIA DE MERCADO E PREMIO A TERMO",
         "PREMISSA declarada, sem identificacao de fonte. Amarelo e editavel. "
         "A curva sem premio continua disponivel na aba CURVA para auditoria.")
    _hdr(ws, ["mes_ref", "horas", "ancora InfoPLD", "REFERENCIA MERCADO (editavel)",
              "premio R$/MWh", "premio %", "premio NIVEL", "premio FORMA",
              "dispersao (Seco-Umido)/Esp", "k", "PREMIO JUSTO", "EDGE = obs - justo",
              "saida CONVERGENCIA", "saida ENTREGA esperado", "saida ENTREGA seco",
              "saida ENTREGA umido"], 4, 15)
    r0 = 5
    for i, m in enumerate(alvo):
        r = r0 + i
        ws.cell(row=r, column=1, value=pd.Timestamp(m).date()).number_format = "yyyy-mm"
        ws.cell(row=r, column=2, value=horas_do_mes(m))
        ws.cell(row=r, column=3, value=f"=INDEX(IP_ESPERADO,{i+1})").number_format = "#,##0.00"
        c = ws.cell(row=r, column=4, value=float(ref_mercado.iloc[i]))
        c.fill = PatternFill("solid", fgColor=F_IN); c.number_format = "#,##0.00"
        ws.cell(row=r, column=5, value=f"=D{r}-C{r}").number_format = "#,##0.00"
        ws.cell(row=r, column=6, value=f"=IF(C{r}=0,\"\",E{r}/C{r})").number_format = "0.0%"
        ws.cell(row=r, column=7, value="=PREMIO_NIVEL").number_format = "#,##0.00"
        ws.cell(row=r, column=8, value=f"=E{r}-G{r}").number_format = "#,##0.00"
        # dispersao = abertura do leque daquele mes; e ela que define quanto de
        # premio o mes MERECE. Mes quase realizado tem leque estreito e premio baixo.
        ws.cell(row=r, column=9, value=(
            f"=(INDEX(IP_SECO,{i+1})-INDEX(IP_UMIDO,{i+1}))/INDEX(IP_ESPERADO,{i+1})"
        )).number_format = "0.000"
        ws.cell(row=r, column=10, value="=K_PREMIO").number_format = "#,##0.0"
        ws.cell(row=r, column=11, value=f"=J{r}*I{r}").number_format = "#,##0.00"
        ws.cell(row=r, column=12, value=f"=E{r}-K{r}").number_format = "#,##0.00"
        ws.cell(row=r, column=13, value=f"=C{r}+K{r}").number_format = "#,##0.00"
        ws.cell(row=r, column=14, value=f"=C{r}").number_format = "#,##0.00"
        ws.cell(row=r, column=15, value=f"=INDEX(IP_SECO,{i+1})").number_format = "#,##0.00"
        ws.cell(row=r, column=16, value=f"=INDEX(IP_UMIDO,{i+1})").number_format = "#,##0.00"
    nf = r0 + n - 1
    b = nf + 2
    ws.cell(row=b, column=1, value="PREMIO NIVEL (media ponderada por horas)").font = Font(bold=True)
    ws.cell(row=b, column=4,
            value=f"=SUMPRODUCT(E{r0}:E{nf},B{r0}:B{nf})/SUM(B{r0}:B{nf})"
            ).number_format = "#,##0.00"
    _nome(wb, "PREMIO_NIVEL", f"REFERENCIA_MERCADO!$D${b}")
    ws.cell(row=b + 1, column=1, value="PREMIO NIVEL em %").font = Font(bold=True)
    ws.cell(row=b + 1, column=4,
            value=f"=SUMPRODUCT(F{r0}:F{nf},B{r0}:B{nf})/SUM(B{r0}:B{nf})"
            ).number_format = "0.0%"
    ws.cell(row=b + 4, column=1, value="k (R$/MWh por unidade de dispersao)").font = Font(bold=True)
    ck = ws.cell(row=b + 4, column=4, value=k_premio)
    ck.fill = PatternFill("solid", fgColor=F_IN); ck.number_format = "#,##0.0"
    _nome(wb, "K_PREMIO", f"REFERENCIA_MERCADO!$D${b+4}")
    ws.cell(row=b + 5, column=1, value=(
        "k e calibrado para que a media ponderada de k*dispersao reproduza o premio "
        "observado. O NIVEL vem do mercado; a FORMA vem do risco de cada mes.")
    ).font = Font(italic=True, size=9)
    ws.cell(row=b + 2, column=1, value="MODO DO PREMIO").font = Font(bold=True)
    cm = ws.cell(row=b + 2, column=4, value=premio_modo)
    cm.fill = PatternFill("solid", fgColor=F_IN)
    ws.cell(row=b + 3, column=1,
            value='visao_propria = so o NIVEL entra na curva; a forma por mes vira SINAL. '
                  'mercado_consistente = premio por mes; a curva reproduz o mercado e nao gera sinal.')
    _nome(wb, "PREMIO_MODO", f"REFERENCIA_MERCADO!$D${b+2}")
    _nome(wb, "REF_MERCADO", f"REFERENCIA_MERCADO!$D${r0}:$D${nf}")
    _nome(wb, "PREMIO_MES", f"REFERENCIA_MERCADO!$E${r0}:$E${nf}")
    _nome(wb, "PREMIO_JUSTO", f"REFERENCIA_MERCADO!$K${r0}:$K${nf}")
    _nome(wb, "EDGE_MES", f"REFERENCIA_MERCADO!$L${r0}:$L${nf}")
    _nome(wb, "SAIDA_CONV", f"REFERENCIA_MERCADO!$M${r0}:$M${nf}")
    _nome(wb, "SAIDA_ESP", f"REFERENCIA_MERCADO!$N${r0}:$N${nf}")
    _nome(wb, "SAIDA_SECO", f"REFERENCIA_MERCADO!$O${r0}:$O${nf}")
    _nome(wb, "SAIDA_UMIDO", f"REFERENCIA_MERCADO!$P${r0}:$P${nf}")

    # ============================================== 2. SINAL
    ws = wb.create_sheet("SINAL")
    _tit(ws, "ONDE O MODELO DISCORDA DO MERCADO",
         "Residuo = fair value do modelo menos referencia de mercado. "
         "Acima do limiar vira acao; abaixo e ruido.")
    _hdr(ws, ["mes_ref", "projecao InfoPLD (E[spot])", "referencia mercado",
              "premio observado", "dispersao do mes", "PREMIO JUSTO",
              "EDGE = obs - justo", "z do edge", "limiar", "ACAO", "conviccao"], 4, 15)
    for i, m in enumerate(alvo):
        r = 5 + i
        ws.cell(row=r, column=1, value=pd.Timestamp(m).date()).number_format = "yyyy-mm"
        ws.cell(row=r, column=2, value=f"=INDEX(IP_ESPERADO,{i+1})").number_format = "#,##0.00"
        ws.cell(row=r, column=3, value=f"=INDEX(REF_MERCADO,{i+1})").number_format = "#,##0.00"
        ws.cell(row=r, column=4, value=f"=C{r}-B{r}").number_format = "#,##0.00"
        ws.cell(row=r, column=5, value=f"=INDEX(DISPERSAO,{i+1})").number_format = "0.000"
        ws.cell(row=r, column=6, value=f"=INDEX(PREMIO_JUSTO,{i+1})").number_format = "#,##0.00"
        # No modo DIRECIONAL o edge e o proprio premio (forward acima da projecao
        # = preco deve cair = VENDER). No modo VALOR RELATIVO o edge e o quanto o
        # premio se afasta do premio justo daquele mes.
        ws.cell(row=r, column=7, value=(
            f'=IF(MODO_SINAL="direcional_premio",D{r},D{r}-F{r})')).number_format = "#,##0.00"
        ws.cell(row=r, column=8, value=f"=IF(SIGMA_EDGE=0,\"\",G{r}/SIGMA_EDGE)").number_format = "0.00"
        ws.cell(row=r, column=9, value="=LIMIAR_SINAL").number_format = "#,##0.00"
        # edge POSITIVO = mercado paga premio demais por esse mes = forward caro = VENDER
        ws.cell(row=r, column=10,
                value=f'=IF(G{r}>=I{r},"VENDER",IF(G{r}<=-I{r},"COMPRAR","FORA"))')
        ws.cell(row=r, column=11, value=f"=ABS(G{r})").number_format = "#,##0.00"
    ns = 4 + n
    _nome(wb, "DISPERSAO", f"REFERENCIA_MERCADO!$I$5:$I${4+n}")
    _nome(wb, "EDGE_SINAL", f"SINAL!$G$5:$G${4+n}")
    _nome(wb, "ACAO_SINAL", f"SINAL!$J$5:$J${4+n}")
    ws.cell(row=ns + 1, column=1, value="MODO DO SINAL").font = Font(bold=True)
    cmd = ws.cell(row=ns + 1, column=4, value=modo_sinal)
    cmd.fill = PatternFill("solid", fgColor=F_IN)
    _nome(wb, "MODO_SINAL", f"SINAL!$D${ns+1}")
    ws.cell(row=ns + 1, column=5, value=(
        'direcional_premio = vende onde o forward esta acima da projecao de PLD '
        '(colhe o premio). valor_relativo = compara o premio com o premio justo '
        'do mes (spread, neutro em nivel).')).font = Font(italic=True, size=9)
    ws.cell(row=ns + 2, column=1, value="Desvio-padrao do edge (sigma)").font = Font(bold=True)
    ws.cell(row=ns + 2, column=4, value=f"=STDEVP(G5:G{ns})").number_format = "#,##0.00"
    _nome(wb, "SIGMA_EDGE", f"SINAL!$D${ns+2}")
    ws.cell(row=ns + 3, column=1, value="k_sigma (quantos sigmas viram sinal)").font = Font(bold=True)
    ck = ws.cell(row=ns + 3, column=4, value=1.0); ck.fill = PatternFill("solid", fgColor=F_IN)
    ws.cell(row=ns + 4, column=1, value="Custo de execucao (piso do limiar)").font = Font(bold=True)
    cc = ws.cell(row=ns + 4, column=4, value=5.0); cc.fill = PatternFill("solid", fgColor=F_IN)
    cc.number_format = "#,##0.00"
    ws.cell(row=ns + 6, column=1, value="Limiar fixo (modo direcional)").font = Font(bold=True)
    cf = ws.cell(row=ns + 6, column=4, value=limiar_sinal)
    cf.fill = PatternFill("solid", fgColor=F_IN); cf.number_format = "#,##0.00"
    ws.cell(row=ns + 5, column=1, value="LIMIAR DE SINAL (R$/MWh)").font = Font(bold=True)
    # O limiar autocalibrado (1 sigma) so vale para o spread de valor relativo,
    # onde o risco de nivel esta neutralizado. No modo direcional o limiar e fixo
    # e declarado — usar sigma ali deixava agosto dentro no Excel e fora no Python.
    ws.cell(row=ns + 5, column=4,
            value=f'=IF(MODO_SINAL="valor_relativo",MAX(D{ns+4},D{ns+3}*D{ns+2}),D{ns+6})'
            ).number_format = "#,##0.00"
    _nome(wb, "LIMIAR_SINAL", f"SINAL!$D${ns+5}")
    ws.cell(row=ns + 7, column=1, value=(
        "COMO LER: o edge NAO e a diferenca entre o forward e a projecao de PLD. "
        "Essa diferenca e o PREMIO, e ela e positiva em todo mes — comparar direto "
        "daria VENDER em tudo, que nao e sinal, e vender o premio de risco. "
        "O edge e o quanto o premio de cada mes se afasta do premio JUSTO daquele mes, "
        "e o premio justo e proporcional a abertura do leque de cenarios. "
        "Assim o mes corrente, que quase nao tem risco restante, nao ganha edge "
        "artificial por receber o premio medio.")).font = Font(italic=True, size=9)
    for palavra, cor in (("COMPRAR", "C8E6C9"), ("VENDER", "FFCDD2")):
        ws.conditional_formatting.add(f"J5:J{ns}", CellIsRule(
            operator="equal", formula=[f'"{palavra}"'],
            fill=PatternFill("solid", fgColor=cor)))

    # ============================================== 3. BOOK
    ws = wb.create_sheet("BOOK")
    _tit(ws, "BOOK MULTI-PERNA — UMA LINHA POR OPERACAO",
         "Cada perna tem produto, lado, MWmed e PRECO DE ENTRADA proprios. "
         "Amarelo e editavel; o resto e formula.")

    # matriz de pertencimento produto x mes
    lin_map = 4
    ws.cell(row=lin_map, column=1, value="MATRIZ PRODUTO x MES (1 = o produto entrega naquele mes)"
            ).font = Font(bold=True, size=10, color=ROSA)
    _hdr(ws, ["produto"] + rot, lin_map + 1, 12)
    for k, p in enumerate(produtos):
        r = lin_map + 2 + k
        ws.cell(row=r, column=1, value=p).font = Font(bold=True)
        ms = set(meses_do_produto(p, alvo))
        for j, m in enumerate(alvo, start=2):
            ws.cell(row=r, column=j, value=1 if m in ms else 0)
    lm0, lm1 = lin_map + 2, lin_map + 1 + len(produtos)
    _nome(wb, "MAPA_PROD", f"BOOK!$A${lm0}:$A${lm1}")
    _nome(wb, "MAPA_MAT", f"BOOK!$B${lm0}:${get_column_letter(1+n)}${lm1}")

    lb = lm1 + 3
    ws.cell(row=lb - 1, column=1, value="PERNAS DO BOOK").font = Font(
        bold=True, size=11, color=ROSA)
    cols = (["#", "produto", "lado (C/V)", "MWmed", "preco entrada", "sinal"]
            + [f"expos {x}" for x in rot]
            + ["PnL convergencia", "PnL entrega esperado", "PnL entrega seco",
               "PnL entrega umido", "energia GWh"])
    _hdr(ws, cols, lb, 12)
    n_linhas = max(len(pernas), 12)      # linhas em branco para o usuario adicionar
    for i in range(n_linhas):
        r = lb + 1 + i
        pn = pernas[i] if i < len(pernas) else None
        ws.cell(row=r, column=1, value=i + 1)
        for col, val, fmt in ((2, pn.produto if pn else None, None),
                              (3, pn.lado if pn else None, None),
                              (4, pn.mwmed if pn else None, "#,##0"),
                              (5, pn.preco_entrada if pn else None, "#,##0.00")):
            c = ws.cell(row=r, column=col, value=val)
            c.fill = PatternFill("solid", fgColor=F_IN)
            if fmt:
                c.number_format = fmt
        ws.cell(row=r, column=6,
                value=f'=IF($B{r}="","",IF(UPPER($C{r})="C",1,IF(UPPER($C{r})="V",-1,"?")))')
        for j in range(n):
            col = 7 + j
            ws.cell(row=r, column=col, value=(
                f'=IF($B{r}="",0,$F{r}*$D{r}*INDEX(MAPA_MAT,MATCH($B{r},MAPA_PROD,0),{j+1}))'
            )).number_format = "#,##0"
        # A exposicao da perna e um vetor HORIZONTAL (G..K) e os precos sao
        # VERTICAIS (SAIDA_*, C_HORAS). SUMPRODUCT exige a mesma orientacao e
        # devolvia #VALOR!. Somatorio explicito por mes resolve e fica legivel.
        for k, cen in enumerate(("SAIDA_CONV", "SAIDA_ESP", "SAIDA_SECO", "SAIDA_UMIDO")):
            termos = "+".join(
                f'{get_column_letter(7+j)}{r}*(INDEX({cen},{j+1})-$E{r})'
                f'*INDEX(C_HORAS,{j+1})' for j in range(n))
            ws.cell(row=r, column=7 + n + k,
                    value=f'=IF($B{r}="",0,{termos})').number_format = "#,##0"
        termos_e = "+".join(f'ABS({get_column_letter(7+j)}{r})*INDEX(C_HORAS,{j+1})'
                            for j in range(n))
        ws.cell(row=r, column=11 + n,
                value=f'=IF($B{r}="",0,({termos_e})/1000)').number_format = "#,##0"
    lb1 = lb + n_linhas
    tot = lb1 + 1
    ws.cell(row=tot, column=2, value="POSICAO LIQUIDA (MWmed)").font = Font(bold=True)
    for j in range(n):
        col = get_column_letter(7 + j)
        ws.cell(row=tot, column=7 + j,
                value=f"=SUM({col}{lb+1}:{col}{lb1})").number_format = "#,##0"
    ws.cell(row=tot + 1, column=2, value="POSICAO BRUTA (MWmed)").font = Font(bold=True)
    for j in range(n):
        col = get_column_letter(7 + j)
        ws.cell(row=tot + 1, column=7 + j,
                value=f"=SUMPRODUCT(ABS({col}{lb+1}:{col}{lb1}))").number_format = "#,##0"
    for k, rot_t in enumerate(("PnL convergencia", "PnL entrega esperado",
                               "PnL entrega seco", "PnL entrega umido")):
        c = get_column_letter(7 + n + k)
        ws.cell(row=tot, column=7 + n + k,
                value=f"=SUM({c}{lb+1}:{c}{lb1})").number_format = "#,##0"
        ws.cell(row=tot, column=7 + n + k).font = Font(bold=True)
    _nome(wb, "BOOK_LIQ", f"BOOK!$G${tot}:${get_column_letter(6+n)}${tot}")
    _nome(wb, "BOOK_PNL_CONV", f"BOOK!${get_column_letter(7+n)}${tot}")
    _nome(wb, "BOOK_PNL_ESP", f"BOOK!${get_column_letter(8+n)}${tot}")
    _nome(wb, "BOOK_PNL_SECO", f"BOOK!${get_column_letter(9+n)}${tot}")
    _nome(wb, "BOOK_PNL_UMIDO", f"BOOK!${get_column_letter(10+n)}${tot}")

    ws.cell(row=tot + 3, column=1,
            value="Comprar Q4/26 e vender OUT/26 zera outubro e preserva novembro e "
                  "dezembro. A posicao liquida acima ja reflete isso: nao ha dupla contagem."
            ).font = Font(italic=True, size=9)

    # ============================================== 4. BOOK_RISCO
    ws = wb.create_sheet("BOOK_RISCO")
    _tit(ws, "RISCO, PnL E VPL DO BOOK",
         "Posicao liquida por mes. O VaR e somado entre meses (correlacao 1): "
         "assumir diversificacao entre meses do mesmo submercado seria otimista.")
    _hdr(ws, ["mes_ref", "MWmed liquido", "horas", "energia GWh", "preco entrada medio",
              "saida CONVERGENCIA", "saida entrega SECO", "saida entrega UMIDO",
              "PnL convergencia", "PnL entrega seco", "PnL entrega umido", "VaR mes",
              "contrib VaR %", "fator desconto", "VP convergencia",
              "VP acumulado", "saida entrega ESPERADO",
              "PnL entrega esperado"], 4, 14)
    for i, m in enumerate(alvo):
        r = 5 + i
        col_b = get_column_letter(7 + i)
        ws.cell(row=r, column=1, value=pd.Timestamp(m).date()).number_format = "yyyy-mm"
        ws.cell(row=r, column=2, value=f"=BOOK!{col_b}${tot}").number_format = "#,##0"
        ws.cell(row=r, column=3, value=horas_do_mes(m))
        ws.cell(row=r, column=4, value=f"=B{r}*C{r}/1000").number_format = "#,##0"
        ws.cell(row=r, column=5, value=(
            f'=IFERROR(SUMPRODUCT(BOOK!{col_b}{lb+1}:{col_b}{lb1},'
            f'BOOK!$E${lb+1}:$E${lb1})/BOOK!{col_b}${tot},"")')).number_format = "#,##0.00"
        for k, nm in enumerate(("SAIDA_CONV", "SAIDA_SECO", "SAIDA_UMIDO")):
            ws.cell(row=r, column=6 + k, value=f"=INDEX({nm},{i+1})").number_format = "#,##0.00"
        for k in range(3):
            ws.cell(row=r, column=9 + k, value=(
                f'=IF(B{r}=0,0,B{r}*C{r}*({get_column_letter(6+k)}{r}-E{r}))')
            ).number_format = "#,##0"
        # VaR do VERTICE, nao a media do strip. Usar VAR_PRECO aqui dava
        # R$ 19,10 mi contra R$ 18,88 mi em POSICAO_BOOK — a mesma grandeza com
        # dois valores, porque a estrutura a termo de vol nao e plana.
        ws.cell(row=r, column=12,
                value=f"=ABS(B{r})*C{r}*INDEX(VAR_VERT,{i+1})").number_format = "#,##0"
        ws.cell(row=r, column=14, value=(
            f"=1/(1+TX_DESC)^((EOMONTH(A{r},0)-DATA_CORTE)/365)")).number_format = "0.0000"
        ws.cell(row=r, column=15, value=f"=I{r}*N{r}").number_format = "#,##0"
        ws.cell(row=r, column=16, value=(f"=O{r}" if i == 0 else f"=P{r-1}+O{r}")
                ).number_format = "#,##0"
        # Cenario ESPERADO: e o caso central do carrego e o que alimenta o PAINEL.
        # Fica em Q/R, no fim, para nao deslocar as colunas ja referenciadas.
        ws.cell(row=r, column=17, value=f"=INDEX(SAIDA_ESP,{i+1})").number_format = "#,##0.00"
        ws.cell(row=r, column=18, value=f'=IF(B{r}=0,0,B{r}*C{r}*(Q{r}-E{r}))'
                ).number_format = "#,##0"
    nr = 4 + n
    for i in range(n):
        ws.cell(row=5 + i, column=13,
                value=f"=IF($L${nr+1}=0,0,L{5+i}/$L${nr+1})").number_format = "0.0%"
    t2 = nr + 1
    ws.cell(row=t2, column=1, value="TOTAL").font = Font(bold=True)
    for col in (9, 10, 11, 12, 15, 18):
        c = get_column_letter(col)
        ws.cell(row=t2, column=col, value=f"=SUM({c}5:{c}{nr})").number_format = "#,##0"
        ws.cell(row=t2, column=col).font = Font(bold=True)

    r2 = t2 + 2
    itens = [
        # COMPONENTE, nao o risco do book. O risco que consome o limite e o
        # maior entre esta marcacao e a perda no cenario adverso, vertice a
        # vertice, e vive em POSICAO_BOOK. Ter dois numeros com o mesmo rotulo
        # "VaR do book" era a origem da divergencia entre as abas.
        ("VaR de marcacao 1 mes — COMPONENTE (R$)", f"=L{t2}", "#,##0"),
        ("  ... quanto e do RISCO DO BOOK", f"=IFERROR(L{t2}/RISCO_BOOK,\"\")", "0.0%"),
        ("RISCO DO BOOK — consome o limite (R$)", "=RISCO_BOOK", "#,##0"),
        ("% do limite de R$ 50 mi (risco do book)", "=RISCO_BOOK/LIMITE_VAR", "0.0%"),
        ("Expected Shortfall da marcacao (R$)", f"=L{t2}/VAR_PRECO*ES_PRECO", "#,##0"),
        ("PnL convergencia do premio (R$)", f"=I{t2}", "#,##0"),
        ("PnL carrego, cenario ESPERADO (R$)", f"=R{t2}", "#,##0"),
        ("PnL carrego, cenario seco (R$)", f"=J{t2}", "#,##0"),
        ("PnL carrego, cenario umido (R$)", f"=K{t2}", "#,##0"),
        ("Pior cenario (R$)", f"=MIN(I{t2},J{t2},K{t2},R{t2})", "#,##0"),
        ("VPL da convergencia ate 31/12 (R$)", f"=O{t2}", "#,##0"),
        # Retorno sobre risco: UMA definicao so em toda a pasta —
        # PnL do carrego esperado dividido pelo RISCO DO BOOK.
        ("Retorno / risco (carrego esperado)", f"=IFERROR(R{t2}/RISCO_BOOK,\"\")", "0.00"),
        ("Retorno / risco (so a convergencia)", f"=IFERROR(I{t2}/RISCO_BOOK,\"\")", "0.00"),
        ("Pior cenario cabe no limite?",
         f'=IF(ABS(MIN(0,MIN(I{t2},J{t2},K{t2},R{t2})))<=LIMITE_VAR,"SIM","NAO")', None),
        ("MWmed liquido absoluto", f"=SUMPRODUCT(ABS(B5:B{nr}))", "#,##0"),
    ]
    for k, (rot_i, f, fmt) in enumerate(itens):
        ws.cell(row=r2 + k, column=1, value=rot_i).font = Font(bold=True)
        c = ws.cell(row=r2 + k, column=3, value=f)
        if fmt:
            c.number_format = fmt
        c.fill = PatternFill("solid", fgColor=F_OUT)
    _nome(wb, "VAR_MTM_BOOK", f"BOOK_RISCO!$C${r2}")
    _nome(wb, "BOOK_VAR", f"BOOK_RISCO!$C${r2+2}")      # = RISCO_BOOK, canonico
    _nome(wb, "BOOK_ES", f"BOOK_RISCO!$C${r2+4}")
    _nome(wb, "BOOK_PNL_CONV", f"BOOK_RISCO!$C${r2+5}")
    _nome(wb, "BOOK_PNL_ESP", f"BOOK_RISCO!$C${r2+6}")
    _nome(wb, "BOOK_PIOR", f"BOOK_RISCO!$C${r2+9}")
    _nome(wb, "BOOK_VPL", f"BOOK_RISCO!$C${r2+10}")
    _nome(wb, "BOOK_MWM_LIQ", f"BOOK_RISCO!$C${r2+14}")

    # --- graficos nativos
    ch = BarChart(); ch.type = "col"; ch.title = "Posicao liquida por mes (MWmed)"
    ch.height, ch.width = 8, 16
    ch.add_data(Reference(ws, min_col=2, min_row=4, max_row=nr), titles_from_data=True)
    ch.set_categories(Reference(ws, min_col=1, min_row=5, max_row=nr))
    ws.add_chart(ch, "R4")

    ch2 = BarChart(); ch2.type = "col"
    ch2.title = "PnL por mes: convergencia do premio x carrego ate a entrega (R$)"
    ch2.height, ch2.width = 8, 16
    ch2.add_data(Reference(ws, min_col=9, max_col=11, min_row=4, max_row=nr),
                 titles_from_data=True)
    ch2.set_categories(Reference(ws, min_col=1, min_row=5, max_row=nr))
    ws.add_chart(ch2, "R22")

    ch3 = LineChart(); ch3.title = "VPL acumulado do book (R$)"; ch3.height, ch3.width = 8, 16
    ch3.add_data(Reference(ws, min_col=16, min_row=4, max_row=nr), titles_from_data=True)
    ch3.set_categories(Reference(ws, min_col=1, min_row=5, max_row=nr))
    ws.add_chart(ch3, "R40")

    return {"abas": ["REFERENCIA_MERCADO", "SINAL", "BOOK", "BOOK_RISCO"],
            "linha_total_book": tot, "primeira_perna": lb + 1, "ultima_perna": lb1}


# =========================================================== organizacao
ORDEM_ABAS = [
    "PAINEL", "POSICAO_BOOK", "EXERCICIO_VAR_100", "BOOK", "BOOK_RISCO", "SINAL", "CURVA",
    "REFERENCIA_MERCADO",
    "INFOPLD_ENTRADA", "INFOPLD_TRAJETORIAS", "IPDO_NOWCAST", "TESE",
    "PARAMETROS", "CHECKS", "CROSSCHECK",
    "CALC_MENSAL", "CALC_SAZONAL", "BACKTEST", "CALC_REGIME", "CALC_FWD",
    "CALC_VAR_NIVEL", "CALC_VAR_SPOT",
    "DADOS_PLD_D", "DADOS_HIDRO_M", "DADOS_MVE", "BOLETINS_AUDITORIA", "MANIFESTO",
]

INDICE = [
    ("OPERAR", None, None),
    ("POSICAO_BOOK", "o que entrar em cada vertice: lado, MWmed pelo VaR e o grafico", "resultado"),
    ("BOOK", "as pernas: produto, lado, MWmed e PRECO DE ENTRADA de cada uma", "editavel"),
    ("EXERCICIO_VAR_100", "o mesmo book ampliado ate consumir 100% do limite de VaR, "
     "com PnL, VPL e MWm nos tres cenarios", "resultado"),
    ("BOOK_RISCO", "posicao liquida por mes, PnL por cenario, VaR, VPL e graficos", "resultado"),
    ("CALC_VAR_NIVEL", "VaR a termo: vol do nivel amortecida ate a entrega, por vertice", "calculo"),
    ("SINAL", "onde o modelo discorda do mercado e vira acao", "resultado"),
    ("CURVA", "fair value mensal, cenarios Seco e Umido", "resultado"),
    ("REFERENCIA_MERCADO", "referencia de mercado e premio a termo", "editavel"),
    ("ENTENDER A CURVA", None, None),
    ("INFOPLD_ENTRADA", "projecoes da CCEE: extraido x override manual x efetivo", "editavel"),
    ("INFOPLD_TRAJETORIAS", "as 5 trajetorias oficiais e o score hidrologico", "referencia"),
    ("IPDO_NOWCAST", "surpresas do dia e decaimento no horizonte", "referencia"),
    ("TESE", "a tese escrita por formula", "resultado"),
    ("CONTROLE", None, None),
    ("PARAMETROS", "as premissas, todas editaveis", "editavel"),
    ("CHECKS", "validacoes automaticas com semaforo", "controle"),
    ("CROSSCHECK", "planilha contra a implementacao de referencia", "controle"),
    ("MOTOR DE CALCULO", None, None),
    ("CALC_MENSAL", "flat mensal, nivel, desvio, nivel EWMA", "calculo"),
    ("CALC_SAZONAL", "os 12 fatores sazonais", "calculo"),
    ("BACKTEST", "walk-forward que escolhe a meia-vida", "calculo"),
    ("CALC_REGIME", "regime hidrologico e multiplicadores", "calculo"),
    ("CALC_FWD", "backcast do forward e VaR a termo", "calculo"),
    ("CALC_VAR_SPOT", "challenger sobre PLD spot", "calculo"),
    ("DADOS", None, None),
    ("DADOS_PLD_D", "PLD diario observado", "fonte"),
    ("DADOS_HIDRO_M", "ENA e EAR mensais", "fonte"),
    ("BOLETINS_AUDITORIA", "proveniencia de cada numero extraido dos PDFs", "auditoria"),
]


def organizar(wb, linha_book_ini: int, linha_book_fim: int) -> dict:
    """Reordena as abas e escreve um indice no PAINEL.

    Com 26 abas, a barra fica lotada e as abas operacionais acabavam no fim,
    fora da area visivel. Aqui elas vao para a frente e o PAINEL ganha um mapa.
    """
    # As abas POSICAO e PNL_VPL vinham do modelo de POSICAO UNICA e contradiziam
    # o book: mostravam uma posicao comprada de 79 MWm enquanto o book estava
    # vendido em quatro vertices. Duas fontes de verdade no mesmo arquivo e o
    # pior defeito possivel. Ficam removidas; BOOK, POSICAO_BOOK e BOOK_RISCO
    # passam a ser a unica origem de posicao, PnL, VPL e risco.
    for legado in ("POSICAO", "PNL_VPL"):
        if legado in wb.sheetnames:
            del wb[legado]
    # Apagar as abas deixaria PAINEL, TESE, CHECKS e CROSSCHECK com #REF!, porque
    # esses nomes apontavam para celulas de la. Em vez de apagar os nomes e
    # reescrever dezenas de formulas, cada nome passa a apontar para o agregado
    # equivalente do book — as formulas continuam validas e agora leem o book.
    equivalencias = {
        # MWM legado apontava para um MWm agregado. Ele nao existe mais como
        # medida de tamanho; o nome sobrevive apontando para o equivalente
        # flat, que e a unica leitura em MWm que fecha aritmeticamente.
        "DIRECAO": "BOOK_DIRECAO", "MWM": "BOOK_MWM_FLAT", "P_ENTRADA": "BOOK_ENTRADA",
        "RESTRICAO": "BOOK_RESTRICAO", "PERDA_CEN": "BOOK_PERDA_CEN",
        "VAR_BRL": "BOOK_VAR", "ES_BRL": "BOOK_ES",
        "PNL_ESP": "BOOK_PNL_ESP", "VPL_BASE": "BOOK_VPL",
    }
    for antigo, novo in equivalencias.items():
        if novo in wb.defined_names and antigo in wb.defined_names:
            wb.defined_names[antigo].value = wb.defined_names[novo].value
    # AVISO DE PROCEDENCIA no topo do PAINEL. A serie de PLD diario define a
    # sazonalidade E o VaR; se ela for fixture, esses dois blocos nao sustentam
    # tese, por mais correta que seja a formula. Isso precisa estar visivel na
    # primeira aba, nao escondido no MANIFESTO.
    if "PAINEL" in wb.sheetnames:
        pw = wb["PAINEL"]
        pw.insert_rows(1, 3)
        pw["A1"] = ("PROCEDENCIA DOS DADOS — LER ANTES DE USAR QUALQUER NUMERO")
        pw["A1"].font = Font(bold=True, size=12, color="B71C1C")
        pw["A2"] = ('=IF(COUNTIF(MANIFESTO!A:A,"*fixture*")+COUNTIF(MANIFESTO!B:B,"*fixture*")>0,'
                    '"ATENCAO: a serie de PLD diario e FIXTURE sintetica. Ancora, cenarios e '
                    'nowcast vem de InfoPLD/IPDO reais, mas SAZONALIDADE e VaR sao calculados '
                    'sobre serie simulada e NAO sustentam tese. Rodar make sync para substituir '
                    'por dado da CCEE.","Serie de PLD diario proveniente de fonte oficial.")')
        pw["A2"].font = Font(bold=True, size=10, color="B71C1C")
        pw["A3"] = ("REAL: ancora de preco, cenarios Seco/Esperado/Umido e nowcast (InfoPLD e "
                    "IPDO, PDFs oficiais). NAO ATUALIZADO: serie de PLD horario 2019-2026 "
                    "(fixture sintetica) -> contamina fatores sazonais, vol do nivel, VaR, ES e "
                    "tudo que deriva deles. Tambem sinteticos: ENA, EAR, CMO e MVE. "
                    "O metodo e a auditoria valem; estes NUMEROS nao sustentam tese.")
        pw["A3"].font = Font(italic=True, size=9)

    ordem = [s for s in ORDEM_ABAS if s in wb.sheetnames]
    ordem += [s for s in wb.sheetnames if s not in ordem]
    wb._sheets = [wb[s] for s in ordem]

    ws = wb["PAINEL"]
    # Este bloco comecava numa linha fixa (40) escolhida quando o painel era
    # menor, e o insert_rows(1,3) logo acima ainda empurra tudo mais 3 linhas.
    # Resultado: o indice de abas caia em cima do fim do painel e da legenda.
    # Agora ancora no ultimo conteudo real das tres primeiras colunas.
    lin = max((c.row for r_ in ws.iter_rows(min_col=1, max_col=3) for c in r_
               if c.value not in (None, "")), default=3) + 3
    ws.cell(row=lin, column=1, value="ONDE ESTA CADA COISA").font = Font(
        bold=True, size=12, color=ROSA)
    lin += 1
    ws.cell(row=lin, column=1, value=(
        f"O book fica na aba BOOK, linhas {linha_book_ini} a {linha_book_fim}. "
        f"Cada linha e uma perna, com produto, lado, MWmed e preco de entrada "
        f"proprios. Celulas amarelas sao editaveis.")).font = Font(italic=True, size=9)
    lin += 2
    _hdr(ws, ["aba", "o que tem", "tipo"], lin, 26)
    lin += 1
    for nome, desc, tipo in INDICE:
        if desc is None:
            c = ws.cell(row=lin, column=1, value=nome)
            c.font = Font(bold=True, size=10, color=ROXO)
            c.fill = PatternFill("solid", fgColor="EEEEF2")
            ws.cell(row=lin, column=2, value="").fill = PatternFill("solid", fgColor="EEEEF2")
            ws.cell(row=lin, column=3, value="").fill = PatternFill("solid", fgColor="EEEEF2")
        else:
            ws.cell(row=lin, column=1, value=nome).font = Font(bold=True, size=9)
            ws.cell(row=lin, column=2, value=desc).font = Font(size=9)
            c = ws.cell(row=lin, column=3, value=tipo)
            c.font = Font(size=9)
            if tipo == "editavel":
                c.fill = PatternFill("solid", fgColor=F_IN)
        lin += 1
    ws.column_dimensions["B"].width = 62
    ws.column_dimensions["C"].width = 14
    return {"ordem": ordem[:6], "linha_indice": 40}


# ==================================================================
# ABA POSICAO — o que entrar em cada vertice, dimensionado pelo VaR
# ==================================================================
def aba_posicao(wb, alvo, dim, ref_book: dict, limite_var: float,
                frac_orc: float, mwm_teto: float) -> dict:
    """POSICAO POR VERTICE — lado, tamanho e risco, um vertice de cada vez.

    Cada vertice tem o SEU lado. Comprado num mes e vendido em outro e o caso
    normal, nao a excecao: o lado sai da comparacao do mes contra a projecao de
    PLD daquele mes.

    O cenario adverso tambem depende do lado — comprado sofre no UMIDO, vendido
    sofre no SECO — e e por isso que o risco nao pode ser um numero unico para o
    book inteiro.
    """
    from .book import rotulo_mes, horas_do_mes
    ws = wb.create_sheet("POSICAO_BOOK")
    n = len(alvo)
    _tit(ws, "POSICAO POR VERTICE — LADO E TAMANHO DEFINIDOS VERTICE A VERTICE",
         "Lado: forward acima da projecao de PLD -> VENDIDO; abaixo -> COMPRADO. "
         "Tamanho: orcamento de risco / risco do vertice. O stop e o limite de risco.")

    _hdr(ws, ["vertice", "projecao PLD", "forward (entrada)", "premio",
              "LADO", "cenario adverso", "preco adverso", "perda/MWm cenario",
              "VaR marcacao/MWm", "RISCO/MWm = VaR", "cenario / VaR (stress)",
              "peso conviccao", "MWmed", "risco consumido", "% do limite",
              "restricao"], 4, 13)

    d = dim.set_index("mes_ref") if (dim is not None and len(dim)) else None
    r0 = 5
    for i, m in enumerate(alvo):
        r = r0 + i
        ts = pd.Timestamp(m)
        ws.cell(row=r, column=1, value=rotulo_mes(m)).font = Font(bold=True)
        ws.cell(row=r, column=2, value=f"=INDEX(IP_ESPERADO,{i+1})").number_format = "#,##0.00"
        ws.cell(row=r, column=3, value=f"=INDEX(REF_MERCADO,{i+1})").number_format = "#,##0.00"
        ws.cell(row=r, column=4, value=f"=C{r}-B{r}").number_format = "#,##0.00"
        # LADO por vertice: le a acao da aba SINAL, que compara mes a mes
        ws.cell(row=r, column=5, value=(
            f'=IF(INDEX(ACAO_SINAL,{i+1})="VENDER","VENDIDO",'
            f'IF(INDEX(ACAO_SINAL,{i+1})="COMPRAR","COMPRADO","FORA"))'))
        ws.cell(row=r, column=6, value=f'=IF(E{r}="COMPRADO","UMIDO",IF(E{r}="VENDIDO","SECO",""))')
        ws.cell(row=r, column=7, value=(
            f'=IF(E{r}="COMPRADO",INDEX(C_UMIDO,{i+1}),'
            f'IF(E{r}="VENDIDO",INDEX(C_SECO,{i+1}),""))')).number_format = "#,##0.00"
        # perda por MWm no cenario adverso, com o sinal correto do lado
        ws.cell(row=r, column=8, value=(
            f'=IF(E{r}="FORA",0,{horas_do_mes(m)}*MAX(0,'
            f'IF(E{r}="COMPRADO",C{r}-G{r},G{r}-C{r})))')).number_format = "#,##0"
        ws.cell(row=r, column=9, value=f"=INDEX(VAR_VERT,{i+1})*{horas_do_mes(m)}").number_format = "#,##0"
        # RISCO = VaR de marcacao. A coluna H (perda no cenario adverso) fica
        # como informacao de stress e NAO entra aqui: horizontes diferentes —
        # o VaR e de 1 mes de remarcacao, a perda de cenario e de carregar ate
        # a entrega. Maximizar entre as duas nao tem leitura unica de limite.
        ws.cell(row=r, column=10, value=f'=IF(E{r}="FORA",0,I{r})').number_format = "#,##0"
        # quantas vezes a perda de carrego no cenario adverso supera o VaR de 1 mes
        ws.cell(row=r, column=11, value=(
            f'=IF(OR(E{r}="FORA",I{r}=0),"",H{r}/I{r})')).number_format = "0.00\"x\""
        ws.cell(row=r, column=12, value=(
            f'=IFERROR(ABS(INDEX(EDGE_SINAL,{i+1}))*(E{r}<>"FORA")/SOMA_CONV,0)'
        )).number_format = "0.0%"
        mw = float(d.loc[ts, "mwmed"]) if (d is not None and ts in d.index) else 0.0
        cmw = ws.cell(row=r, column=13, value=mw)
        cmw.number_format = "#,##0"; cmw.fill = PatternFill("solid", fgColor=F_IN)
        ws.cell(row=r, column=14, value=f"=M{r}*J{r}").number_format = "#,##0"
        ws.cell(row=r, column=15, value=f"=N{r}/LIMITE_VAR").number_format = "0.0%"
        ws.cell(row=r, column=16, value=(
            f'=IF(E{r}="FORA","sem sinal",IF(M{r}>=MWM_TETO,"liquidez",'
            f'IF(ABS(M{r}-IFERROR(FLOOR(ORC_RISCO*L{r}/J{r},1),0))<=1,"risco","manual")))'))
    nr = 4 + n
    t2 = nr + 1
    ws.cell(row=t2, column=1, value="TOTAL").font = Font(bold=True)
    for col in (13, 14):
        L = get_column_letter(col)
        ws.cell(row=t2, column=col, value=f"=SUM({L}5:{L}{nr})").number_format = "#,##0"
        ws.cell(row=t2, column=col).font = Font(bold=True)
    # componentes agregados, para a reconciliacao logo abaixo
    for col in (8, 9):
        L = get_column_letter(col)
        ws.cell(row=t2, column=col,
                value=f"=SUMPRODUCT({L}5:{L}{nr},$M$5:$M${nr})").number_format = "#,##0"
    ws.cell(row=t2, column=15, value=f"=N{t2}/LIMITE_VAR").number_format = "0.0%"
    # RISCO_BOOK e a UNICA medida que consome o limite em toda a pasta.
    _nome(wb, "RISCO_BOOK", f"POSICAO_BOOK!$N${t2}")
    _nome(wb, "PERDA_CEN_BOOK", f"POSICAO_BOOK!$H${t2}")

    b = t2 + 2
    ws.cell(row=b - 1, column=1, value="CONTROLE DE RISCO E STOP").font = Font(
        bold=True, size=11, color=ROSA)
    itens = [
        ("Limite de risco (stop)", "=LIMITE_VAR", "#,##0"),
        ("Fracao utilizavel (buffer)", "=FRAC_ORC", "0%"),
        ("ORCAMENTO DE RISCO", "=LIMITE_VAR*FRAC_ORC", "#,##0"),
        ("Risco consumido pelo book", f"=N{t2}", "#,##0"),
        ("Utilizacao do orcamento", f"=N{t2}/(LIMITE_VAR*FRAC_ORC)", "0.0%"),
        ("Folga ate o STOP", f"=LIMITE_VAR-N{t2}", "#,##0"),
        ("Situacao", f'=IF(N{t2}>LIMITE_VAR,"ESTOUROU O STOP",'
                     f'IF(N{t2}>LIMITE_VAR*FRAC_ORC,"acima do orcamento","dentro"))', None),
    ]
    for k, (rot_i, f, fmt) in enumerate(itens):
        ws.cell(row=b + k, column=1, value=rot_i).font = Font(bold=True)
        c = ws.cell(row=b + k, column=3, value=f)
        if fmt:
            c.number_format = fmt
        c.fill = PatternFill("solid", fgColor=F_OUT)
    _nome(wb, "ORC_RISCO", f"POSICAO_BOOK!$C${b+2}")
    ws.cell(row=b, column=5, value="=SUMPRODUCT(ABS(EDGE_SINAL)*(ACAO_SINAL<>\"FORA\"))"
            ).number_format = "#,##0.00"
    ws.cell(row=b, column=6, value="soma das conviccoes com sinal (denominador do peso)")
    _nome(wb, "SOMA_CONV", f"POSICAO_BOOK!$E${b}")

    for palavra, cor in (("COMPRADO", "C8E6C9"), ("VENDIDO", "FFCDD2")):
        ws.conditional_formatting.add(f"E5:E{nr}", CellIsRule(
            operator="equal", formula=[f'"{palavra}"'],
            fill=PatternFill("solid", fgColor=cor)))
    ws.conditional_formatting.add(f"C{b+3}", CellIsRule(
        operator="greaterThan", formula=["LIMITE_VAR"],
        fill=PatternFill("solid", fgColor="FFCDD2")))

    # --- RECONCILIACAO: por que existem tres numeros e qual manda.
    rc = b + 8
    ws.cell(row=rc - 1, column=1, value="RECONCILIACAO DAS MEDIDAS DE RISCO"
            ).font = Font(bold=True, size=11, color=ROSA)
    ws.cell(row=rc - 1, column=4, value=(
        "Tres medidas, uma so consome o limite. Vertice a vertice o risco e o MAIOR "
        "entre a marcacao e o cenario adverso; nunca a soma dos dois."))
    rec = [
        ("(A) VaR de marcacao 1 mes — E O RISCO", f"=I{t2}",
         "quanto a posicao perde se o preco a termo andar contra em 1 mes"),
        ("(B) Perda no cenario adverso ate a entrega — informacao", f"=H{t2}",
         "quanto perde se carregar ate o fim e o cenario adverso do LADO se realizar; "
         "horizonte diferente do VaR, por isso nao dimensiona"),
        ("RISCO DO BOOK = (A), o VaR", f"=N{t2}",
         "o case deu o VaR como unica restricao; e ele que consome o limite"),
        ("Confere: risco do book = VaR?",
         f'=IF(ABS(N{t2}-I{t2})<1,"SIM","NAO — revisar")', ""),
        ("Cobertura do stress: (B) / (A)", f"=IFERROR(H{t2}/I{t2},\"\")",
         "quantas vezes a perda de carrego no cenario adverso supera o VaR de 1 mes"),
        ("Vertices com stress acima do VaR",
         f'=SUMPRODUCT(--(H5:H{nr}>I5:I{nr}),--(M5:M{nr}>0))&" de "'
         f'&SUMPRODUCT(--(M5:M{nr}>0))', "informacao, nao restricao"),
    ]
    for k, (rot_i, f, obs) in enumerate(rec):
        ws.cell(row=rc + k, column=1, value=rot_i).font = Font(bold=True)
        c = ws.cell(row=rc + k, column=3, value=f)
        c.number_format = "#,##0"
        c.fill = PatternFill("solid", fgColor=F_OUT if k != 5 else "EEEEEE")
        ws.cell(row=rc + k, column=4, value=obs).font = Font(italic=True, size=9)
    ws.cell(row=rc + 5, column=1).font = Font(bold=True, color="9E9E9E")

    # --- CONSOLIDADO: os nomes que o PAINEL, a TESE e os CHECKS consomem.
    # Antes vinham da aba POSICAO (posicao unica). Agora sao agregados do book,
    # e o lado deixa de ser um numero so: pode ser MISTO, e o painel diz isso.
    k0 = rc + 8
    ws.cell(row=k0 - 1, column=1, value="CONSOLIDADO DO BOOK (origem do PAINEL)"
            ).font = Font(bold=True, size=11, color=ROSA)
    cons = [
        ("Lado do book", f'=IF(COUNTIF(E5:E{nr},"COMPRADO")*COUNTIF(E5:E{nr},"VENDIDO")>0,'
                         f'"MISTO",IF(COUNTIF(E5:E{nr},"VENDIDO")>0,"VENDIDO",'
                         f'IF(COUNTIF(E5:E{nr},"COMPRADO")>0,"COMPRADO","FORA")))', None),
        ("Direcao numerica (1 C / -1 V / 0 misto)",
         f'=IF(C{k0}="VENDIDO",-1,IF(C{k0}="COMPRADO",1,0))', "0"),
        # TAMANHO. Nao existe um MWm unico que descreva um strip de produtos
        # MENSAIS com quantidade diferente em cada mes. A posicao e o ladder
        # da tabela acima; o que agrega sem mentir sobre o shape e energia e
        # notional. As duas leituras em MWm ficam abaixo, rotuladas.
        ("TAMANHO — energia (MWh)", f"=SUMPRODUCT(M5:M{nr},C_HORAS)", "#,##0"),
        ("TAMANHO — notional (R$)",
         f"=SUMPRODUCT(M5:M{nr},C5:C{nr},C_HORAS)", "#,##0"),
        ("Preco medio de entrada (R$/MWh)",
         f"=IFERROR(SUMPRODUCT(M5:M{nr},C5:C{nr},C_HORAS)/SUMPRODUCT(M5:M{nr},C_HORAS),0)",
         "#,##0.00"),
        ("Horas dos meses com posicao",
         f"=SUMPRODUCT(--(M5:M{nr}>0),C_HORAS)", "#,##0"),
        # aritmeticamente correto, mas achata o shape: serve para comparar com
        # um produto FLAT de trimestre ou ano, nao para descrever este book
        ("Equivalente flat (MWm) — so p/ comparacao",
         f"=IFERROR(SUMPRODUCT(M5:M{nr},C_HORAS)/"
         f"SUMPRODUCT(--(M5:M{nr}>0),C_HORAS),0)", "#,##0.0"),
        # soma de potencia de meses distintos: conta a mesma potencia varias
        # vezes. Fica visivel so para nao ser recalculada por engano
        ("Soma de MWm das pernas (NAO e tamanho)", f"=SUM(M5:M{nr})", "#,##0"),
        ("Restricao predominante",
         f'=IF(COUNTIF(P5:P{nr},"liquidez")>0,"liquidez","risco")', None),
        ("Perda no cenario adverso (R$)", f"=H{t2}", "#,##0"),
        ("RISCO DO BOOK (consome o limite)", f"=N{t2}", "#,##0"),
    ]
    for k, (rot_i, f, fmt) in enumerate(cons):
        ws.cell(row=k0 + k, column=1, value=rot_i).font = Font(bold=True)
        c = ws.cell(row=k0 + k, column=3, value=f)
        if fmt:
            c.number_format = fmt
        c.fill = PatternFill("solid", fgColor=F_OUT)
    _nome(wb, "LADO_VERT", f"POSICAO_BOOK!$E$5:$E${nr}")
    _nome(wb, "BOOK_LADO", f"POSICAO_BOOK!$C${k0}")
    _nome(wb, "BOOK_DIRECAO", f"POSICAO_BOOK!$C${k0+1}")
    _nome(wb, "BOOK_ENERGIA_MWH", f"POSICAO_BOOK!$C${k0+2}")
    _nome(wb, "BOOK_NOTIONAL", f"POSICAO_BOOK!$C${k0+3}")
    _nome(wb, "BOOK_ENTRADA", f"POSICAO_BOOK!$C${k0+4}")
    _nome(wb, "BOOK_HORAS", f"POSICAO_BOOK!$C${k0+5}")
    _nome(wb, "BOOK_MWM_FLAT", f"POSICAO_BOOK!$C${k0+6}")
    _nome(wb, "BOOK_SOMA_MWM", f"POSICAO_BOOK!$C${k0+7}")
    _nome(wb, "BOOK_RESTRICAO", f"POSICAO_BOOK!$C${k0+8}")
    _nome(wb, "BOOK_PERDA_CEN", f"POSICAO_BOOK!$C${k0+9}")

    # --- serie assinada para o grafico
    # k0..k0+9 sao as 10 linhas do CONSOLIDADO; comeca depois delas
    g0 = k0 + len(cons) + 2
    ws.cell(row=g0 - 1, column=1, value="POSICAO ASSINADA (comprado +, vendido -)"
            ).font = Font(bold=True, size=10, color=ROSA)
    _hdr(ws, ["vertice", "MWm assinado"], g0, 16)
    for i, m in enumerate(alvo):
        r = g0 + 1 + i
        ws.cell(row=r, column=1, value=rotulo_mes(m))
        ws.cell(row=r, column=2, value=(
            f'=IF(E{5+i}="COMPRADO",M{5+i},IF(E{5+i}="VENDIDO",-M{5+i},0))')
        ).number_format = "#,##0"
    ch = BarChart(); ch.type = "col"
    ch.title = "Posicao MWm por vertice — comprado positivo, vendido negativo"
    ch.height, ch.width = 9, 22
    ch.y_axis.title = "MWmed"
    ch.add_data(Reference(ws, min_col=2, min_row=g0, max_row=g0 + n),
                titles_from_data=True)
    ch.set_categories(Reference(ws, min_col=1, min_row=g0 + 1, max_row=g0 + n))
    ws.add_chart(ch, "R4")

    ch2 = BarChart(); ch2.type = "col"
    ch2.title = "Risco consumido por vertice (R$)"
    ch2.height, ch2.width = 9, 22
    ch2.add_data(Reference(ws, min_col=14, min_row=4, max_row=nr), titles_from_data=True)
    ch2.set_categories(Reference(ws, min_col=1, min_row=5, max_row=nr))
    ws.add_chart(ch2, "R24")
    return {"aba": "POSICAO_BOOK", "linha_total": t2, "primeira": r0, "ultima": nr}


def aba_exercicio_var(wb, alvo, pos: dict) -> dict:
    """EXERCICIO_VAR_100 — o book dimensionado para consumir 100% do limite.

    POR QUE ESTA ABA EXISTE
    -----------------------
    O case deu UMA restricao: o VaR. O book operacional para em 60% do limite
    porque guarda buffer para remarcacao adversa antes do stop — decisao de mesa,
    nao exigencia do case. Esta aba responde a pergunta que a restricao dada de
    fato faz: qual a capacidade do book quando o limite e a unica amarra.

    COMO ESCALA
    -----------
    risco(mes) = MWm(mes) x risco_por_MWm(mes), e risco_por_MWm NAO depende do
    tamanho. Entao o risco total e linear no tamanho e um unico fator resolve:

        fator = LIMITE_VAR / risco consumido pelo book base

    O fator preserva a reparticao por conviccao entre os vertices. Nao e um
    redimensionamento novo: e o MESMO book, ampliado ate encostar no limite.

    O QUE A ABA NAO ESCONDE
    -----------------------
    O teto de liquidez nao escala junto. Vertices que passam de MWM_TETO estao
    marcados: no papel o VaR permite, na mesa o mercado nao absorve.
    """
    n = len(alvo)
    p0, p1, ptot = pos["primeira"], pos["ultima"], pos["linha_total"]
    ws = wb.create_sheet("EXERCICIO_VAR_100")
    _tit(ws, "EXERCICIO — BOOK NO LIMITE CHEIO DE VaR (100% DE R$ 50 MI)",
         "O VaR foi a unica restricao dada. Esta aba mostra o mesmo book, com os "
         "mesmos lados e a mesma reparticao, ampliado ate consumir 100% do limite.")

    ws.cell(row=4, column=1, value="FATOR DE ESCALA").font = Font(bold=True, size=11, color=ROSA)
    esc = [("Limite de VaR (restricao do case)", "=LIMITE_VAR", "#,##0"),
           ("Risco consumido pelo book operacional", "=RISCO_BOOK", "#,##0"),
           ("Utilizacao do limite hoje", "=RISCO_BOOK/LIMITE_VAR", "0.0%"),
           ("FATOR = limite / risco atual", "=IFERROR(LIMITE_VAR/RISCO_BOOK,0)", "0.000")]
    for k, (rot_i, f, fmt) in enumerate(esc):
        ws.cell(row=5 + k, column=1, value=rot_i).font = Font(bold=True)
        c = ws.cell(row=5 + k, column=3, value=f)
        c.number_format = fmt
        c.fill = PatternFill("solid", fgColor=F_OUT)
    _nome(wb, "FATOR_VAR100", "EXERCICIO_VAR_100!$C$8")

    h0 = 11
    _hdr(ws, ["vertice", "LADO", "sinal", "MWm operacional", "MWm no limite cheio",
              "risco/MWm", "risco consumido", "% do limite", "entrada",
              "saida ESPERADO", "saida SECO", "saida UMIDO", "horas",
              "PnL ESPERADO", "PnL SECO", "PnL UMIDO", "PnL convergencia",
              "fator desconto", "VPL esperado", "teto de liquidez",
              "VPL convergencia"], h0, 14)
    for i in range(n):
        r, rp = h0 + 1 + i, p0 + i
        ws.cell(row=r, column=1, value=f"=POSICAO_BOOK!A{rp}")
        ws.cell(row=r, column=2, value=f"=POSICAO_BOOK!E{rp}")
        ws.cell(row=r, column=3, value=f'=IF(B{r}="COMPRADO",1,IF(B{r}="VENDIDO",-1,0))')
        ws.cell(row=r, column=4, value=f"=POSICAO_BOOK!M{rp}").number_format = "#,##0"
        # FLOOR, nao ROUND: arredondar para cima deixava o consumo em 100,02% do
        # limite. Uma aba de risco nunca pode exibir o limite estourado por
        # arredondamento — o lote inteiro sempre desce.
        ws.cell(row=r, column=5, value=f"=FLOOR(D{r}*FATOR_VAR100,1)").number_format = "#,##0"
        ws.cell(row=r, column=6, value=f"=POSICAO_BOOK!J{rp}").number_format = "#,##0"
        ws.cell(row=r, column=7, value=f"=E{r}*F{r}").number_format = "#,##0"
        ws.cell(row=r, column=8, value=f"=G{r}/LIMITE_VAR").number_format = "0.0%"
        ws.cell(row=r, column=9, value=f"=POSICAO_BOOK!C{rp}").number_format = "#,##0.00"
        for k, nm in enumerate(("SAIDA_ESP", "SAIDA_SECO", "SAIDA_UMIDO")):
            ws.cell(row=r, column=10 + k, value=f"=INDEX({nm},{i+1})").number_format = "#,##0.00"
        ws.cell(row=r, column=13, value=f"=INDEX(C_HORAS,{i+1})").number_format = "#,##0"
        # PnL = sinal x MWm x horas x (saida - entrada). Vendido ganha quando cai.
        for k in range(3):
            ws.cell(row=r, column=14 + k, value=(
                f"=$C{r}*$E{r}*$M{r}*({get_column_letter(10+k)}{r}-$I{r})")).number_format = "#,##0"
        ws.cell(row=r, column=17, value=(
            f"=$C{r}*$E{r}*$M{r}*(INDEX(SAIDA_CONV,{i+1})-$I{r})")).number_format = "#,##0"
        ws.cell(row=r, column=18, value=(
            f"=1/(1+TX_DESC)^((EOMONTH(INDEX(C_MES,{i+1}),0)-DATA_CORTE)/365)")
        ).number_format = "0.0000"
        ws.cell(row=r, column=19, value=f"=N{r}*R{r}").number_format = "#,##0"
        ws.cell(row=r, column=20, value=(
            f'=IF(E{r}>MWM_TETO,"ESTOURA em "&TEXT(E{r}-MWM_TETO,"#,##0")&" MWm","cabe")'))
        ws.cell(row=r, column=21, value=f"=Q{r}*R{r}").number_format = "#,##0"
    hf = h0 + n
    tt = hf + 1
    ws.cell(row=tt, column=1, value="TOTAL").font = Font(bold=True)
    for col in (4, 5, 7, 14, 15, 16, 17, 19, 21):
        L = get_column_letter(col)
        c = ws.cell(row=tt, column=col, value=f"=SUM({L}{h0+1}:{L}{hf})")
        c.number_format = "#,##0"; c.font = Font(bold=True)
    ws.cell(row=tt, column=8, value=f"=G{tt}/LIMITE_VAR").number_format = "0.0%"
    ws.cell(row=tt, column=8).font = Font(bold=True)

    b = tt + 3
    ws.cell(row=b - 1, column=1, value="METRICAS DO BOOK NO LIMITE CHEIO"
            ).font = Font(bold=True, size=12, color=ROSA)
    met = [
        ("TAMANHO — energia", f"=SUMPRODUCT(E{h0+1}:E{hf},M{h0+1}:M{hf})", "#,##0", "MWh"),
        ("  equivalente flat",
         f"=IFERROR(SUMPRODUCT(E{h0+1}:E{hf},M{h0+1}:M{hf})/"
         f"SUMPRODUCT(--(E{h0+1}:E{hf}>0),M{h0+1}:M{hf}),0)", "#,##0.0",
         "MWm — so p/ comparar com produto flat"),
        ("  soma de MWm das pernas", f"=E{tt}", "#,##0",
         "MWm — diagnostico; somar meses conta a mesma potencia varias vezes"),
        ("Energia total no periodo", f"=SUMPRODUCT(E{h0+1}:E{hf},M{h0+1}:M{hf})/1000", "#,##0", "GWh"),
        ("RISCO consumido (mesma definicao do book)", f"=G{tt}", "#,##0", "R$"),
        ("% do limite de R$ 50 mi", f"=G{tt}/LIMITE_VAR", "0.0%", ""),
        ("", "", None, ""),
        ("PnL carrego — cenario ESPERADO", f"=N{tt}", "#,##0", "R$"),
        ("PnL carrego — cenario SECO", f"=O{tt}", "#,##0", "R$"),
        ("PnL carrego — cenario UMIDO", f"=P{tt}", "#,##0", "R$"),
        ("PnL convergencia do premio", f"=Q{tt}", "#,##0", "R$"),
        ("VPL do carrego esperado ate 31/12", f"=S{tt}", "#,##0", "R$"),
        ("VPL da convergencia ate 31/12", f"=U{tt}", "#,##0", "R$"),
        ("", "", None, ""),
        ("Pior cenario", f"=MIN(N{tt},O{tt},P{tt},Q{tt})", "#,##0", "R$"),
        ("Melhor cenario", f"=MAX(N{tt},O{tt},P{tt})", "#,##0", "R$"),
        ("Amplitude (melhor - pior)", f"=MAX(N{tt},O{tt},P{tt})-MIN(N{tt},O{tt},P{tt},Q{tt})", "#,##0", "R$"),
        ("Retorno / risco (carrego esperado)", f"=IFERROR(N{tt}/G{tt},\"\")", "0.00", "x"),
        ("Retorno / risco (pior caso)", f"=IFERROR(MIN(N{tt},O{tt},P{tt},Q{tt})/G{tt},\"\")", "0.00", "x"),
        ("Consumo <= limite?",
         f'=IF(G{tt}<=LIMITE_VAR,"SIM","NAO — revisar arredondamento")', None, ""),
        ("Congruente com o book base?",
         f'=IF(ABS(G{tt}/RISCO_BOOK-E{tt}/D{tt})<0.01,"SIM — mesma definicao de risco",'
         f'"NAO — definicoes divergentes")', None, ""),
        ("Pior cenario cabe no limite?",
         f'=IF(ABS(MIN(0,MIN(N{tt},O{tt},P{tt},Q{tt})))<=LIMITE_VAR,"SIM","NAO — estoura o stop")',
         None, ""),
        ("Vertices que estouram a liquidez", f'=COUNTIF(T{h0+1}:T{hf},"ESTOURA*")', "0", "de "+str(n)),
    ]
    for k, (rot_i, f, fmt, un) in enumerate(met):
        if not rot_i:
            continue
        ws.cell(row=b + k, column=1, value=rot_i).font = Font(bold=True)
        c = ws.cell(row=b + k, column=3, value=f)
        if fmt:
            c.number_format = fmt
        c.fill = PatternFill("solid", fgColor=F_OUT)
        ws.cell(row=b + k, column=4, value=un)
    fim_met = b + len(met)

    # ---- comparacao lado a lado
    c0 = fim_met + 2
    ws.cell(row=c0 - 1, column=1, value="OPERACIONAL (60% DO LIMITE) x LIMITE CHEIO (100%)"
            ).font = Font(bold=True, size=11, color=ROSA)
    _hdr(ws, ["metrica", "operacional", "limite cheio", "diferenca", "multiplo"], c0, 22)
    comp = [("Soma de MWm das pernas", f"=D{tt}", f"=E{tt}", "#,##0"),
            ("RISCO consumido (R$)", "=RISCO_BOOK", f"=G{tt}", "#,##0"),
            ("  do qual, marcacao 1 mes", "=VAR_MTM_BOOK",
             f"=VAR_MTM_BOOK*FATOR_VAR100", "#,##0"),
            ("PnL esperado (R$)", "=BOOK_PNL_ESP", f"=N{tt}", "#,##0"),
            ("PnL seco (R$)", "=BOOK_PNL_SECO", f"=O{tt}", "#,##0"),
            ("PnL umido (R$)", "=BOOK_PNL_UMIDO", f"=P{tt}", "#,##0"),
            ("PnL convergencia (R$)", "=BOOK_PNL_CONV", f"=Q{tt}", "#,##0"),
            # BOOK_VPL desconta a CONVERGENCIA. Comparar contra o VPL do carrego
            # daria um multiplo de 190x, que nao mede nada — sao grandezas
            # diferentes. A comparacao e convergencia contra convergencia.
            ("VPL convergencia (R$)", "=BOOK_VPL", f"=U{tt}", "#,##0"),
            ("VPL carrego esperado (R$)", f'=IFERROR(S{tt}/FATOR_VAR100,"")', f"=S{tt}", "#,##0")]
    for k, (rot_i, fa, fb, fmt) in enumerate(comp):
        r = c0 + 1 + k
        ws.cell(row=r, column=1, value=rot_i).font = Font(bold=True)
        ws.cell(row=r, column=2, value=fa).number_format = fmt
        ws.cell(row=r, column=3, value=fb).number_format = fmt
        ws.cell(row=r, column=4, value=f"=C{r}-B{r}").number_format = fmt
        ws.cell(row=r, column=5, value=f'=IFERROR(C{r}/B{r},"")').number_format = "0.00"

    ws.cell(row=c0 + len(comp) + 3, column=1, value=(
        "LEITURA: o retorno sobre VaR nao melhora ao ampliar — o book escala "
        "linearmente, entao PnL e VaR crescem no mesmo fator. O que o exercicio "
        "mostra e CAPACIDADE, nao eficiencia. O limite operacional de 60% existe "
        "para absorver remarcacao adversa antes do stop, e o teto de liquidez "
        "morde antes do VaR em parte dos vertices."))
    ws.cell(row=c0 + len(comp) + 3, column=1).font = Font(italic=True, size=9)

    ws.conditional_formatting.add(f"T{h0+1}:T{hf}", CellIsRule(
        operator="beginsWith", formula=['"ESTOURA"'],
        fill=PatternFill("solid", fgColor="FFCDD2")))
    for col, w in (("A", 36), ("B", 12), ("C", 16), ("D", 16), ("E", 18)):
        ws.column_dimensions[col].width = w

    ch = BarChart(); ch.type = "col"; ch.grouping = "clustered"
    ch.title = "PnL por cenario no limite cheio de VaR (R$)"
    ch.height, ch.width = 9, 24
    ch.add_data(Reference(ws, min_col=14, max_col=16, min_row=h0, max_row=hf),
                titles_from_data=True)
    ch.set_categories(Reference(ws, min_col=1, min_row=h0 + 1, max_row=hf))
    ws.add_chart(ch, "V4")

    ch2 = BarChart(); ch2.type = "col"
    ch2.title = "MWm por vertice — operacional x limite cheio"
    ch2.height, ch2.width = 9, 24
    ch2.add_data(Reference(ws, min_col=4, max_col=5, min_row=h0, max_row=hf),
                 titles_from_data=True)
    ch2.set_categories(Reference(ws, min_col=1, min_row=h0 + 1, max_row=hf))
    ws.add_chart(ch2, "V24")
    return {"aba": "EXERCICIO_VAR_100", "total": tt}
