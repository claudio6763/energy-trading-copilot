# -*- coding: utf-8 -*-
"""Abas do Excel alimentadas por InfoPLD e IPDO.

DUAS VIAS DE ENTRADA, e as duas convivem:

    AUTOMATICA  o parser preenche as colunas EXTRAIDO (fundo verde)
    MANUAL      o usuario digita nas colunas OVERRIDE (fundo amarelo)
    EFETIVO     formula: =SE(override<>""; override; extraido)

O valor manual NUNCA apaga o extraido. Os dois ficam lado a lado, a coluna
"origem" diz qual venceu, e a aba de auditoria registra fonte, pagina e
justificativa de cada override. Isso e o que permite defender o numero.
"""
from __future__ import annotations

import pandas as pd
from openpyxl.chart import LineChart, Reference
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName

ROSA, ROXO = "C4007A", "1F1235"
F_EXTR, F_MAN, F_EFET, F_AUD = "E8F5E9", "FFF8E1", "E3F2FD", "F3E5F5"


def _hdr(ws, cols, linha, larg=14):
    for j, c in enumerate(cols, 1):
        cel = ws.cell(row=linha, column=j, value=c)
        cel.fill = PatternFill("solid", fgColor=ROXO)
        cel.font = Font(color="FFFFFF", bold=True, size=9)
        cel.alignment = Alignment(horizontal="center", wrap_text=True, vertical="center")
        ws.column_dimensions[get_column_letter(j)].width = larg
    ws.freeze_panes = ws.cell(row=linha + 1, column=1)


def _tit(ws, t, sub=""):
    ws["A1"] = t
    ws["A1"].font = Font(bold=True, size=13, color=ROSA)
    if sub:
        ws["A2"] = sub
        ws["A2"].font = Font(italic=True, size=9, color="666677")


def _nome(wb, n, ref):
    if n in wb.defined_names:
        del wb.defined_names[n]
    wb.defined_names.add(DefinedName(n, attr_text=ref))


def adicionar_abas(wb, bol: dict, alvo, submercado: str = "SE") -> dict:
    """bol: dict com chaves infopld_df, ipdo_df, cenarios, diag_cen, nowcast_det,
    diag_nowcast, docs (lista de Documento.dict())."""
    info = bol["infopld_df"]
    ipdo = bol["ipdo_df"]
    cen = bol["cenarios"]
    dcen = bol["diag_cen"]
    n = len(alvo)

    # ================================================ 1. INFOPLD_ENTRADA
    ws = wb.create_sheet("INFOPLD_ENTRADA")
    _tit(ws, "PROJECOES DO INFOPLD — ENTRADA AUTOMATICA E MANUAL",
         "Verde = extraido do PDF. Amarelo = override manual (deixe vazio para usar o extraido). "
         "Azul = efetivo, calculado por formula.")
    cols = ["mes_ref",
            "PLD Esperado (extraido)", "PLD Seco (extraido)", "PLD Umido (extraido)",
            "ENA %MLT SIN (extraido)", "EARM % SE/CO (extraido)",
            "PLD Esp (override)", "PLD Seco (override)", "PLD Umido (override)",
            "ENA %MLT (override)", "EARM % (override)",
            "PLD ESPERADO", "PLD SECO", "PLD UMIDO", "ENA %MLT", "EARM %",
            "origem PLD Esp"]
    _hdr(ws, cols, 4, 15)

    from .ancora import matriz_trajetorias
    ena_sin = matriz_trajetorias(info, "ENA_PROJ_TRAJ", "SIN", alvo)
    earm_se = matriz_trajetorias(info, "EARM_PROJ_TRAJ", submercado, alvo)
    tj_esp = dcen["trajetoria_esperado"]

    r0 = 5
    for i, m in enumerate(alvo):
        r = r0 + i
        ws.cell(row=r, column=1, value=pd.Timestamp(m).date()).number_format = "yyyy-mm"
        for j, v in enumerate([cen.Esperado.iloc[i], cen.Seco.iloc[i], cen.Umido.iloc[i],
                               ena_sin[tj_esp].iloc[i] if tj_esp in ena_sin.columns else None,
                               earm_se[tj_esp].iloc[i] if tj_esp in earm_se.columns else None],
                              start=2):
            c = ws.cell(row=r, column=j,
                        value=None if v is None or pd.isna(v) else float(v))
            c.fill = PatternFill("solid", fgColor=F_EXTR)
            c.number_format = "#,##0.00"
        for j in range(7, 12):                      # overrides, vazios
            ws.cell(row=r, column=j).fill = PatternFill("solid", fgColor=F_MAN)
            ws.cell(row=r, column=j).number_format = "#,##0.00"
        for k in range(5):                          # efetivos, por formula
            ex = get_column_letter(2 + k)
            ov = get_column_letter(7 + k)
            c = ws.cell(row=r, column=12 + k, value=f'=IF({ov}{r}<>"",{ov}{r},{ex}{r})')
            c.fill = PatternFill("solid", fgColor=F_EFET)
            c.number_format = "#,##0.00"
        ws.cell(row=r, column=17, value=f'=IF(G{r}<>"","MANUAL","EXTRAIDO")')
    nf = r0 + n - 1
    _nome(wb, "IP_MES", f"INFOPLD_ENTRADA!$A${r0}:$A${nf}")
    _nome(wb, "IP_ESPERADO", f"INFOPLD_ENTRADA!$L${r0}:$L${nf}")
    _nome(wb, "IP_SECO", f"INFOPLD_ENTRADA!$M${r0}:$M${nf}")
    _nome(wb, "IP_UMIDO", f"INFOPLD_ENTRADA!$N${r0}:$N${nf}")
    _nome(wb, "IP_ENA", f"INFOPLD_ENTRADA!$O${r0}:$O${nf}")
    _nome(wb, "IP_EARM", f"INFOPLD_ENTRADA!$P${r0}:$P${nf}")

    b = nf + 2
    for k, (rot, val) in enumerate([
        ("Documento", dcen.get("documento", bol.get("doc_infopld", ""))),
        ("Trajetoria Esperado", dcen["trajetoria_esperado"]),
        ("Trajetoria Seco", dcen["trajetoria_seco"]),
        ("Trajetoria Umido", dcen["trajetoria_umido"]),
        ("Mes com valor unico do resumo", ", ".join(dcen.get("meses_valor_unico_resumo", []))),
        ("Posicao do central no leque", dcen.get("posicao_central_no_leque", "")),
        ("Ordenacao Seco>=Esperado>=Umido", "SIM" if dcen["ordenacao_coerente"] else "NAO"),
    ]):
        ws.cell(row=b + k, column=1, value=rot).font = Font(bold=True)
        ws.cell(row=b + k, column=2, value=str(val))
    if dcen.get("alerta"):
        c = ws.cell(row=b + 8, column=1, value="ACHADO: " + dcen["alerta"])
        c.font = Font(bold=True, color="B00020")
        ws.merge_cells(start_row=b + 8, start_column=1, end_row=b + 8, end_column=12)

    ch = LineChart(); ch.title = "Cenarios do InfoPLD (R$/MWh)"; ch.height, ch.width = 8, 16
    ch.add_data(Reference(ws, min_col=12, max_col=14, min_row=4, max_row=nf),
                titles_from_data=True)
    ch.set_categories(Reference(ws, min_col=1, min_row=r0, max_row=nf))
    ws.add_chart(ch, "S5")

    # ================================================ 2. TRAJETORIAS oficiais
    ws = wb.create_sheet("INFOPLD_TRAJETORIAS")
    _tit(ws, "AS CINCO TRAJETORIAS OFICIAIS — SENSIBILIDADES",
         "Cada trajetoria e um cenario coerente e e preservada inteira. "
         "Combinar meses de modelos diferentes criaria uma curva que nenhum modelo gerou.")
    lin = 4
    for var, rot, un in (("PLD_PROJ_TRAJ", "PLD", "R$/MWh"),
                         ("ENA_PROJ_TRAJ", "ENA", "%MLT"),
                         ("EARM_PROJ_TRAJ", "EARM", "%EARMmax")):
        sub = "SIN" if var == "ENA_PROJ_TRAJ" else submercado
        piv = matriz_trajetorias(info, var, sub, alvo)
        ws.cell(row=lin, column=1, value=f"{rot} — {sub} ({un})").font = Font(
            bold=True, size=11, color=ROSA)
        lin += 1
        _hdr(ws, ["trajetoria"] + [pd.Timestamp(m).strftime("%b/%y") for m in alvo], lin, 13)
        lin += 1
        for t in piv.columns:
            ws.cell(row=lin, column=1, value=t).font = Font(bold=True)
            for j, m in enumerate(alvo, start=2):
                v = piv[t].iloc[j - 2]
                ws.cell(row=lin, column=j,
                        value=None if pd.isna(v) else float(v)).number_format = "#,##0"
            lin += 1
        lin += 2

    sc = pd.DataFrame(dcen["score"])
    ws.cell(row=lin, column=1, value="SCORE HIDROLOGICO (score alto = umido)").font = Font(
        bold=True, size=11, color=ROSA)
    lin += 1
    _hdr(ws, ["trajetoria", "ENA media %MLT", "EARM media %", "cobertura", "score"], lin, 16)
    lin += 1
    for _, rr in sc.iterrows():
        ws.cell(row=lin, column=1, value=rr.trajetoria)
        for j, c in enumerate(["ena_media_pct_mlt", "earm_media_pct",
                               "cobertura_horizonte", "score_umidade"], start=2):
            ws.cell(row=lin, column=j, value=None if pd.isna(rr[c]) else float(rr[c])
                    ).number_format = "#,##0.00"
        lin += 1

    # ================================================ 3. IPDO / NOWCAST
    ws = wb.create_sheet("IPDO_NOWCAST")
    _tit(ws, "IPDO — ESTADO OBSERVADO E SURPRESAS DO DIA",
         "Programado x verificado. A surpresa e CALCULADA, nunca extraida como dado.")
    _hdr(ws, ["variavel", "submercado", "programado", "verificado", "surpresa %",
              "z", "sensibilidade", "contribuicao %"], 4, 16)
    det = bol["nowcast_det"]
    for i, rr in det.iterrows():
        r = 5 + i
        ws.cell(row=r, column=1, value=rr.driver)
        ws.cell(row=r, column=2, value="SIN")
        ws.cell(row=r, column=3, value=float(rr.programado)).number_format = "#,##0"
        ws.cell(row=r, column=4, value=float(rr.realizado)).number_format = "#,##0"
        ws.cell(row=r, column=5, value=f"=IF(C{r}=0,\"\",D{r}/C{r}-1)").number_format = "0.00%"
        ws.cell(row=r, column=6, value=float(rr.z)).number_format = "0.000"
        ws.cell(row=r, column=7, value=float(rr.sensibilidade)).number_format = "0.00"
        ws.cell(row=r, column=8, value=f"=F{r}*G{r}").number_format = "0.000"
    nd = 4 + len(det)
    tot = nd + 2
    ws.cell(row=tot, column=1, value="Contribuicao total (%)").font = Font(bold=True)
    ws.cell(row=tot, column=8, value=f"=SUM(H5:H{nd})").number_format = "0.000"
    _nome(wb, "NOWCAST_TOTAL", f"IPDO_NOWCAST!$H${tot}")
    ws.cell(row=tot + 1, column=1, value="Meia-vida do decaimento (meses)").font = Font(bold=True)
    ws.cell(row=tot + 1, column=8, value=bol["diag_nowcast"]["meia_vida_meses"])
    _nome(wb, "NOWCAST_MV", f"IPDO_NOWCAST!$H${tot+1}")

    ws.cell(row=tot + 3, column=1,
            value="DECAIMENTO POR MES — o que aconteceu ontem move o mes corrente, "
                  "quase nao move dezembro").font = Font(bold=True, color=ROSA)
    _hdr(ws, ["mes_ref", "h (meses a frente)", "fator 0,5^(h/meia-vida)",
              "ajuste % no preco"], tot + 4, 20)
    for i, m in enumerate(alvo):
        r = tot + 5 + i
        ws.cell(row=r, column=1, value=pd.Timestamp(m).date()).number_format = "yyyy-mm"
        ws.cell(row=r, column=2, value=i)
        ws.cell(row=r, column=3, value=f"=0.5^(B{r}/NOWCAST_MV)").number_format = "0.0000"
        ws.cell(row=r, column=4, value=f"=NOWCAST_TOTAL*C{r}").number_format = "0.000"
    _nome(wb, "NOWCAST_AJUSTE", f"IPDO_NOWCAST!$D${tot+5}:$D${tot+4+n}")

    est = ipdo[ipdo.variavel.isin(["EARM_PCT", "EARM_VAR_DIA", "EARM_VAR_MES"])]
    le = tot + 6 + n
    ws.cell(row=le, column=1, value="ESTADO DOS RESERVATORIOS (IPDO)").font = Font(
        bold=True, size=11, color=ROSA)
    _hdr(ws, ["variavel", "SIN", "S", "SE", "N", "NE", "unidade"], le + 1, 14)
    for k, v in enumerate(["EARM_PCT", "EARM_VAR_DIA", "EARM_VAR_MES"]):
        r = le + 2 + k
        ws.cell(row=r, column=1, value=v).font = Font(bold=True)
        sel = est[est.variavel == v]
        for j, s in enumerate(["SIN", "S", "SE", "N", "NE"], start=2):
            x = sel[sel.submercado == s]
            if len(x):
                ws.cell(row=r, column=j, value=float(x.valor.iloc[0])).number_format = "0.0"
        ws.cell(row=r, column=7, value=sel.unidade.iloc[0] if len(sel) else "")

    # ================================================ 4. AUDITORIA
    ws = wb.create_sheet("BOLETINS_AUDITORIA")
    _tit(ws, "PROVENIENCIA — DE ONDE VEIO CADA NUMERO",
         "Documento, pagina, natureza, metodo e confianca de cada observacao extraida.")
    _hdr(ws, ["documento", "tipo", "data_referencia", "sha256 (12)", "paginas", "bytes"],
         4, 26)
    for i, d in enumerate(bol.get("docs", [])):
        r = 5 + i
        for j, k in enumerate(["caminho", "tipo", "data_referencia"], start=1):
            ws.cell(row=r, column=j, value=str(d.get(k, ""))[-45:])
        ws.cell(row=r, column=4, value=str(d.get("sha256", ""))[:12])
        ws.cell(row=r, column=5, value=d.get("paginas"))
        ws.cell(row=r, column=6, value=d.get("bytes"))
    base = 5 + len(bol.get("docs", [])) + 2

    obs = pd.concat([info.assign(fonte="InfoPLD"), ipdo.assign(fonte="IPDO")],
                    ignore_index=True)
    obs = obs[["fonte", "variavel", "submercado", "periodo", "valor", "unidade",
               "natureza", "cenario", "pagina", "metodo", "confianca",
               "regra_validacao"]]
    _hdr(ws, list(obs.columns), base, 15)
    for i, rr in obs.iterrows():
        r = base + 1 + i
        for j, c in enumerate(obs.columns, start=1):
            v = rr[c]
            ws.cell(row=r, column=j, value=None if pd.isna(v) else
                    (float(v) if isinstance(v, (int, float)) else str(v)))
        if float(rr.confianca) < 0.9:
            for j in range(1, len(obs.columns) + 1):
                ws.cell(row=r, column=j).fill = PatternFill("solid", fgColor="FFCDD2")
    ws.conditional_formatting.add(
        f"K{base+1}:K{base+len(obs)}",
        CellIsRule(operator="lessThan", formula=["0.9"],
                   fill=PatternFill("solid", fgColor="FFCDD2")))

    return {"abas": ["INFOPLD_ENTRADA", "INFOPLD_TRAJETORIAS", "IPDO_NOWCAST",
                     "BOLETINS_AUDITORIA"],
            "n_obs_auditoria": len(obs),
            "linha_final_entrada": nf}
