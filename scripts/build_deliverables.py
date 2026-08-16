#!/usr/bin/env python3
"""Gera os entregaveis: PDFs (ReportLab) e planilha aberta (openpyxl).

Valida com pypdf (limite de paginas) e openpyxl (formulas vivas, sem valor colado).

    python scripts/build_deliverables.py
"""
from __future__ import annotations

import argparse
import re
import sys
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT)); sys.path.insert(0, str(ROOT / "src"))
OUT = ROOT / "deliverables"


def md_to_pdf(origem: Path, destino: Path, *, titulo: str, max_pages: int) -> tuple[bool, str]:
    from reportlab.lib.enums import TA_JUSTIFY
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
    from reportlab.lib.units import cm
    from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer

    estilos = getSampleStyleSheet()
    corpo = ParagraphStyle("corpo", parent=estilos["BodyText"], fontSize=7.6,
                           leading=9.4, alignment=TA_JUSTIFY, spaceAfter=2)
    h1 = ParagraphStyle("h1", parent=estilos["Heading1"], fontSize=13, leading=15, spaceAfter=4)
    h2 = ParagraphStyle("h2", parent=estilos["Heading2"], fontSize=8.6, leading=10,
                        spaceBefore=4, spaceAfter=2)

    doc = SimpleDocTemplate(str(destino), pagesize=A4, topMargin=1.1*cm,
                            bottomMargin=1.1*cm, leftMargin=1.4*cm, rightMargin=1.4*cm,
                            title=titulo)
    fluxo = []
    for linha in origem.read_text(encoding="utf-8").splitlines():
        t = linha.strip()
        if not t:
            fluxo.append(Spacer(1, 2)); continue
        if t.startswith("|") or set(t) <= set("-| "):
            continue
        t = re.sub(r"\*\*(.+?)\*\*", r"<b>\1</b>", t)
        t = re.sub(r"`(.+?)`", r"<font face='Courier'>\1</font>", t)
        t = t.replace("&", "&amp;").replace("<b>", "\x00b\x01").replace("</b>", "\x00/b\x01")
        t = t.replace("<font face='Courier'>", "\x00f\x01").replace("</font>", "\x00/f\x01")
        t = t.replace("<", "&lt;").replace(">", "&gt;")
        t = (t.replace("\x00b\x01", "<b>").replace("\x00/b\x01", "</b>")
              .replace("\x00f\x01", "<font face='Courier'>").replace("\x00/f\x01", "</font>"))
        if t.startswith("## "):
            fluxo.append(Paragraph(t[3:], h2))
        elif t.startswith("# "):
            fluxo.append(Paragraph(t[2:], h1))
        elif t.startswith("- ") or t.startswith("* "):
            fluxo.append(Paragraph("&bull; " + t[2:], corpo))
        else:
            fluxo.append(Paragraph(t, corpo))
    doc.build(fluxo)

    from pypdf import PdfReader
    paginas = len(PdfReader(str(destino)).pages)
    ok = paginas <= max_pages
    return ok, f"{destino.name}: {paginas} pagina(s) (limite {max_pages})"


def build_workbook(destino: Path) -> str:
    """Planilha com formulas ABERTAS. Nenhum valor colado onde deveria haver calculo."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    wb = Workbook()
    entrada = PatternFill("solid", fgColor="FFF2CC")   # amarelo = input
    calculo = PatternFill("solid", fgColor="DDEBF7")   # azul = formula
    cabecalho = Font(bold=True, color="FFFFFF")
    fundo = PatternFill("solid", fgColor="1F3864")

    def titulo(ws, linha, celulas):
        for i, texto in enumerate(celulas, start=1):
            c = ws.cell(row=linha, column=i, value=texto)
            c.font = cabecalho; c.fill = fundo

    ws = wb.active; ws.title = "LEIA_ME"
    for i, linha in enumerate([
        ["Energy Trading Copilot — modelo da Entrega 2"], [],
        ["Amarelo = INPUT (editavel). Azul = FORMULA (nao sobrescrever)."],
        ["Todas as celulas de calculo sao formulas vivas. Nenhum valor foi colado."], [],
        ["Aba", "Conteudo"],
        ["INPUTS", "Premissas, volume, precos, taxa de desconto e data-base"],
        ["FONTES_CURVA", "Origem de cada preco, classificacao e evidence_id"],
        ["POSICAO", "Dimensionamento: MWmed, horas, MWh, notional"],
        ["CENARIOS_PNL", "P&L por cenario hidrologico"],
        ["VAR", "VaR, add-ons, VaR total e consumo do limite"],
        ["MARGEM_NPV", "Resultado descontado ate 31/12"],
        ["CHECKS", "Validacoes de consistencia"], [],
        ["ATENCAO: preencher com dados reais da data-base antes de entregar."],
    ], start=1):
        for j, v in enumerate(linha, start=1):
            ws.cell(row=i, column=j, value=v)
    ws.column_dimensions["A"].width = 22; ws.column_dimensions["B"].width = 74
    ws["A1"].font = Font(bold=True, size=14)

    ws = wb.create_sheet("INPUTS")
    titulo(ws, 1, ["Parametro", "Valor", "Unidade", "Fonte", "Data-base", "Classificacao"])
    for i, (p, v, u, f, d, c) in enumerate([
        ("Volume", 50, "MWmed", "Decisao da mesa", "2026-08-14", "manual"),
        ("Preco de entrada", 195.00, "R$/MWh", "PREENCHER", "2026-08-14", "PREENCHER"),
        ("Preco de mercado", 205.00, "R$/MWh", "PREENCHER", "2026-08-14", "PREENCHER"),
        ("Lado (1=comprado, -1=vendido)", -1, "sinal", "Decisao da mesa", "2026-08-14", "manual"),
        ("Inicio da entrega", "2027-01-01", "data", "Decisao da mesa", "2026-08-14", "manual"),
        ("Fim da entrega", "2027-12-31", "data", "Decisao da mesa", "2026-08-14", "manual"),
        ("Volatilidade diaria", 0.0180, "fracao", "Motor quant", "2026-08-14", "projetado"),
        ("Confianca do VaR", 0.95, "fracao", "Premissa PR-03", "2026-08-14", "manual"),
        ("Horizonte do VaR", 21, "dias uteis", "Premissa PR-03", "2026-08-14", "manual"),
        ("Limite de VaR", 50000000, "R$", "Case", "2026-08-14", "manual"),
        ("Taxa de desconto", 0.10, "a.a.", "Premissa da mesa", "2026-08-14", "manual"),
        ("Add-on de proxy", 0.00, "fracao", "quant/addons.py", "2026-08-14", "manual"),
        ("Add-on de risco de modelo", 0.10, "fracao", "quant/addons.py", "2026-08-14", "manual"),
    ], start=2):
        ws.cell(row=i, column=1, value=p)
        ws.cell(row=i, column=2, value=v).fill = entrada
        for j, val in enumerate((u, f, d, c), start=3):
            ws.cell(row=i, column=j, value=val)
    for col, w in zip("ABCDEF", (30, 16, 12, 22, 13, 15)):
        ws.column_dimensions[col].width = w

    ws = wb.create_sheet("FONTES_CURVA")
    titulo(ws, 1, ["Tenor", "Inicio", "Fim", "Preco", "Unidade", "Tipo de cotacao",
                   "Origem", "Fonte", "Data-base", "evidence_id"])
    for i, t in enumerate(["A+1", "A+2", "A+3"], start=2):
        ws.cell(row=i, column=1, value=t)
        for j in range(2, 11):
            ws.cell(row=i, column=j, value="PREENCHER").fill = entrada
    ws.cell(row=6, column=1, value="PLD e CMO NAO sao curva forward negociada. "
            "Se usados como referencia, marque Origem=PROXY_SPOT e aplique o add-on.")
    for col, w in zip("ABCDEFGHIJ", (8, 12, 12, 10, 10, 16, 14, 18, 12, 28)):
        ws.column_dimensions[col].width = w

    ws = wb.create_sheet("POSICAO")
    titulo(ws, 1, ["Item", "Formula / valor", "Unidade"])
    for i, (rot, form, un) in enumerate([
        ("Volume", "=INPUTS!B2", "MWmed"),
        ("Dias do periodo", "=INPUTS!B6-INPUTS!B5+1", "dias"),
        ("Horas do periodo", "=B3*24", "h"),
        ("Volume em energia", "=B2*B4", "MWh"),
        ("Preco de entrada", "=INPUTS!B3", "R$/MWh"),
        ("Notional", "=B5*B6", "R$"),
        ("Lado", "=INPUTS!B5", "sinal"),
        ("Exposicao assinada", "=B8*B5*INPUTS!B4", "R$"),
    ], start=2):
        ws.cell(row=i, column=1, value=rot)
        c = ws.cell(row=i, column=2, value=form)
        if str(form).startswith("="):
            c.fill = calculo
        ws.cell(row=i, column=3, value=un)
    for col, w in zip("ABC", (26, 30, 12)):
        ws.column_dimensions[col].width = w

    ws = wb.create_sheet("CENARIOS_PNL")
    titulo(ws, 1, ["Cenario", "Multiplicador", "Preco do cenario", "P&L", "Probabilidade",
                   "P&L ponderado", "O que muda na tese"])
    for i, (nome, mult, prob, delta) in enumerate([
        ("BASE", 1.00, 0.50, "Tese mantida; reavaliar na data prevista."),
        ("SECO", 1.35, 0.30, "Vendido perde; checar gatilho de saida."),
        ("UMIDO", 0.75, 0.20, "Vendido ganha; avaliar realizacao parcial."),
        ("EXTREMO", 1.80, 0.00, "Estresse, nao previsao: dimensiona cauda."),
    ], start=2):
        ws.cell(row=i, column=1, value=nome)
        ws.cell(row=i, column=2, value=mult).fill = entrada
        ws.cell(row=i, column=3, value=f"=INPUTS!$B$4*B{i}").fill = calculo
        ws.cell(row=i, column=4, value=f"=POSICAO!$B$5*(INPUTS!$B$3-C{i})*INPUTS!$B$5*-1").fill = calculo
        ws.cell(row=i, column=5, value=prob).fill = entrada
        ws.cell(row=i, column=6, value=f"=D{i}*E{i}").fill = calculo
        ws.cell(row=i, column=7, value=delta)
    ws.cell(row=6, column=1, value="Esperanca (sem estresse)").font = Font(bold=True)
    ws.cell(row=6, column=6, value="=SUM(F2:F4)").fill = calculo
    for col, w in zip("ABCDEFG", (12, 14, 16, 16, 13, 16, 46)):
        ws.column_dimensions[col].width = w

    ws = wb.create_sheet("VAR")
    titulo(ws, 1, ["Item", "Formula", "Unidade"])
    for i, (rot, form, un) in enumerate([
        ("Exposicao", "=ABS(POSICAO!B9)", "R$"),
        ("Volatilidade diaria", "=INPUTS!B8", "fracao"),
        ("z (confianca)", "=NORM.S.INV(INPUTS!B9)", "-"),
        ("Raiz do horizonte", "=SQRT(INPUTS!B10)", "-"),
        ("VaR de mercado", "=B2*B3*B4*B5", "R$"),
        ("Add-on de proxy", "=B2*INPUTS!B13", "R$"),
        ("Add-on de risco de modelo", "=B6*INPUTS!B14", "R$"),
        ("VaR total", "=B6+B7+B8", "R$"),
        ("Limite", "=INPUTS!B11", "R$"),
        ("Consumo do limite", "=B9/B10", "fracao"),
        ("Dentro do limite?", '=IF(B9<=B10,"SIM","NAO — BLOQUEADA")', "-"),
    ], start=2):
        ws.cell(row=i, column=1, value=rot)
        ws.cell(row=i, column=2, value=form).fill = calculo
        ws.cell(row=i, column=3, value=un)
    for col, w in zip("ABC", (28, 34, 12)):
        ws.column_dimensions[col].width = w

    ws = wb.create_sheet("MARGEM_NPV")
    titulo(ws, 1, ["Item", "Formula", "Unidade"])
    for i, (rot, form, un) in enumerate([
        ("P&L esperado", "=CENARIOS_PNL!F6", "R$"),
        ("Data-base", "=INPUTS!B6", "data"),
        ("Dias ate 31/12", '=DATE(YEAR(INPUTS!B6),12,31)-INPUTS!B6', "dias"),
        ("Taxa anual", "=INPUTS!B12", "a.a."),
        ("Fator de desconto", "=1+B5*B4/365", "-"),
        ("NPV ate 31/12", "=B2/B6", "R$"),
    ], start=2):
        ws.cell(row=i, column=1, value=rot)
        ws.cell(row=i, column=2, value=form).fill = calculo
        ws.cell(row=i, column=3, value=un)
    for col, w in zip("ABC", (26, 40, 12)):
        ws.column_dimensions[col].width = w

    ws = wb.create_sheet("CHECKS")
    titulo(ws, 1, ["Verificacao", "Resultado", "Esperado"])
    for i, (rot, form, esp) in enumerate([
        ("Horas do periodo = 8760 ou 8784", '=IF(OR(POSICAO!B4=8760,POSICAO!B4=8784),"OK","ERRO")', "OK"),
        ("MWh > 0", '=IF(POSICAO!B5>0,"OK","ERRO")', "OK"),
        ("VaR total <= limite", '=IF(VAR!B9<=VAR!B10,"OK","ESTOUROU")', "OK"),
        ("Probabilidades somam 1", '=IF(ABS(SUM(CENARIOS_PNL!E2:E4)-1)<0.001,"OK","ERRO")', "OK"),
        ("Ha 2+ cenarios hidrologicos", '=IF(COUNTA(CENARIOS_PNL!A2:A5)>=2,"OK","ERRO")', "OK"),
        ("Fontes da curva preenchidas", '=IF(COUNTIF(FONTES_CURVA!H2:H4,"PREENCHER")=0,"OK","PENDENTE")', "OK"),
    ], start=2):
        ws.cell(row=i, column=1, value=rot)
        ws.cell(row=i, column=2, value=form).fill = calculo
        ws.cell(row=i, column=3, value=esp)
    for col, w in zip("ABC", (38, 46, 12)):
        ws.column_dimensions[col].width = w
    for aba in wb.worksheets:
        aba.freeze_panes = "A2"
    wb.save(destino)

    from openpyxl import load_workbook
    conferencia = load_workbook(destino)
    abas = set(conferencia.sheetnames)
    exigidas = {"LEIA_ME", "INPUTS", "FONTES_CURVA", "POSICAO", "CENARIOS_PNL",
                "VAR", "MARGEM_NPV", "CHECKS"}
    if not exigidas <= abas:
        raise AssertionError(f"abas faltando: {exigidas - abas}")
    formulas = sum(1 for a in conferencia.worksheets for linha in a.iter_rows()
                   for c in linha if isinstance(c.value, str) and c.value.startswith("="))
    if formulas < 30:
        raise AssertionError(f"poucas formulas ({formulas}): valores podem ter sido colados")
    return f"{destino.name}: {len(abas)} abas, {formulas} formulas vivas"


def main(argv: list[str] | None = None) -> int:
    argparse.ArgumentParser(description="Gera os entregaveis.").parse_args(argv)
    OUT.mkdir(exist_ok=True)
    resultados: list[tuple[bool, str]] = []

    resultados.append(md_to_pdf(OUT / "entrega_1_one_pager.md",
                                OUT / "entrega_1_one_pager.pdf",
                                titulo="Energy Trading Copilot — Entrega 1", max_pages=1))
    posicao = OUT / "entrega_2_posicao.md"
    if posicao.exists():
        resultados.append(md_to_pdf(posicao, OUT / "entrega_2_posicao.pdf",
                                    titulo="Entrega 2 — Proposta de posicao", max_pages=2))
    try:
        resultados.append((True, build_workbook(OUT / "entrega_2_modelo.xlsx")))
    except Exception as exc:
        resultados.append((False, f"planilha: {exc}"))

    print("\nEntregaveis:")
    for ok, msg in resultados:
        print(f"  [{'OK  ' if ok else 'FALHA'}] {msg}")
    falhas = [m for ok, m in resultados if not ok]
    if falhas:
        print("\nFALHOU:", "; ".join(falhas))
        return 1
    print("\nTodos os entregaveis gerados e validados.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
