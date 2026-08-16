#!/usr/bin/env python3
"""Verificador do MVP ponta a ponta. Retorna 0 so quando tudo aplicavel passa.

Diferenca para `verify_agent.py`: aquele e a trava de liberacao do agente; este
cobre tambem os entregaveis e a Entrega 2 (quando liberada).
"""
from __future__ import annotations

import subprocess
import sys
import time
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT)); sys.path.insert(0, str(ROOT / "src"))

RESULTS: list[tuple[str, str, float, str]] = []


def etapa(nome: str, fn) -> bool:
    inicio = time.perf_counter()
    try:
        msg = fn() or "ok"
        RESULTS.append((nome, "PASS", time.perf_counter() - inicio, str(msg)))
        return True
    except Exception as exc:
        RESULTS.append((nome, "FAIL", time.perf_counter() - inicio, f"{type(exc).__name__}: {exc}"))
        return False


def skip(nome: str, msg: str) -> None:
    RESULTS.append((nome, "N/A", 0.0, msg))


def rodar(*args: str) -> str:
    p = subprocess.run([sys.executable, *args], cwd=ROOT, capture_output=True, text=True)
    if p.returncode != 0:
        raise AssertionError((p.stdout + p.stderr).strip()[-400:])
    return (p.stdout.strip().splitlines() or ["ok"])[-1][:80]


def main() -> int:
    ok = True
    ok &= etapa("1-19. trava do agente (verify_agent.py)",
                lambda: rodar("scripts/verify_agent.py"))
    ok &= etapa("20. entregaveis (build_deliverables.py)",
                lambda: rodar("scripts/build_deliverables.py"))

    def _um_pager():
        from pypdf import PdfReader
        n = len(PdfReader(str(ROOT / "deliverables/entrega_1_one_pager.pdf")).pages)
        assert n == 1, f"one-pager com {n} paginas"
        return "1 pagina"
    ok &= etapa("21. Entrega 1 tem 1 pagina", _um_pager)

    def _planilha():
        from openpyxl import load_workbook
        wb = load_workbook(ROOT / "deliverables/entrega_2_modelo.xlsx")
        exigidas = {"LEIA_ME", "INPUTS", "FONTES_CURVA", "POSICAO", "CENARIOS_PNL",
                    "VAR", "MARGEM_NPV", "CHECKS"}
        assert exigidas <= set(wb.sheetnames), f"abas faltando: {exigidas - set(wb.sheetnames)}"
        f = sum(1 for a in wb.worksheets for l in a.iter_rows() for c in l
                if isinstance(c.value, str) and c.value.startswith("="))
        assert f >= 30, f"apenas {f} formulas: valores podem ter sido colados"
        return f"{len(wb.sheetnames)} abas, {f} formulas"
    ok &= etapa("22. planilha aberta com formulas", _planilha)

    def _docs():
        import re
        texto = (ROOT / "README.md").read_text(encoding="utf-8")
        quebrados = [a for a in re.findall(r"\]\((docs/[^)]+|deliverables/[^)]+)\)", texto)
                     if not (ROOT / a).exists()]
        assert not quebrados, f"links quebrados: {quebrados}"
        return "sem links internos quebrados"
    ok &= etapa("23. documentacao consistente", _docs)

    liberado = (ROOT / "AGENT_READY.md").exists()
    if not liberado:
        skip("24. Entrega 2", "AGENT_READY.md ausente — Entrega 2 nao liberada")
    elif (ROOT / "READY_FOR_ENTREGA_2.md").exists() and not (
            ROOT / "deliverables/entrega_2_posicao.pdf").exists():
        skip("24. Entrega 2", "aguardando dados reais da data-base (ver READY_FOR_ENTREGA_2.md)")
    else:
        ok &= etapa("24. Entrega 2", lambda: rodar("scripts/verify_entrega_2.py"))

    largura = max(len(n) for n, _, _, _ in RESULTS) + 2
    print("\n" + "=" * 96)
    print(f"{'COMPONENTE':<{largura}} {'RESULTADO':<10} {'TEMPO':>8}  MENSAGEM")
    print("-" * 96)
    for nome, status, dur, msg in RESULTS:
        print(f"{nome:<{largura}} {status:<10} {dur*1000:>7.0f}ms  {msg[:60]}")
    print("=" * 96)
    p = sum(1 for _, s, _, _ in RESULTS if s == "PASS")
    n = sum(1 for _, s, _, _ in RESULTS if s == "N/A")
    print(f"{p} passaram, {n} nao aplicaveis, {len(RESULTS)-p-n} falharam.")
    print("\nMVP VERIFICADO." if ok else "\nMVP COM FALHAS.")
    return 0 if ok else 1


if __name__ == "__main__":
    raise SystemExit(main())
