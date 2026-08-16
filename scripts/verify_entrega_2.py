#!/usr/bin/env python3
"""Verifica a Entrega 2. Retorna 0 so quando ela esta completa e valida."""
from __future__ import annotations

import sys
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT)); sys.path.insert(0, str(ROOT / "src"))
OUT = ROOT / "deliverables"
PROBLEMAS: list[str] = []


def exigir(cond: bool, msg: str) -> None:
    if not cond:
        PROBLEMAS.append(msg)


def main() -> int:
    exigir((ROOT / "AGENT_READY.md").exists(),
           "AGENT_READY.md ausente: a Entrega 2 so pode ser gerada pelo agente aprovado.")
    for nome in ("entrega_2_posicao.md", "entrega_2_posicao.pdf", "entrega_2_modelo.xlsx"):
        exigir((OUT / nome).exists(), f"{nome} ausente")

    if (OUT / "entrega_2_posicao.pdf").exists():
        from pypdf import PdfReader
        n = len(PdfReader(str(OUT / "entrega_2_posicao.pdf")).pages)
        exigir(n <= 2, f"documento com {n} paginas (limite 2)")

    if (OUT / "entrega_2_posicao.md").exists():
        texto = (OUT / "entrega_2_posicao.md").read_text(encoding="utf-8")
        exigir("DEMONSTRA" not in texto.upper(),
               "documento ainda marcado como demonstrativo: nao e posicao oficial")
        for termo in ("Tese", "Dimensionamento", "VaR", "Horizonte", "reavalia",
                      "saída", "invalida", "Premissas", "Fontes", "Cenário", "NPV"):
            exigir(termo.lower() in texto.lower(), f"documento sem seção: {termo}")

    if (OUT / "entrega_2_modelo.xlsx").exists():
        from openpyxl import load_workbook
        wb = load_workbook(OUT / "entrega_2_modelo.xlsx")
        exigir({"LEIA_ME", "INPUTS", "FONTES_CURVA", "POSICAO", "CENARIOS_PNL", "VAR",
                "MARGEM_NPV", "CHECKS"} <= set(wb.sheetnames), "planilha sem todas as abas")
        formulas = sum(1 for a in wb.worksheets for l in a.iter_rows() for c in l
                       if isinstance(c.value, str) and c.value.startswith("="))
        exigir(formulas >= 30, f"apenas {formulas} formulas: valores podem ter sido colados")
        pendentes = sum(1 for a in wb.worksheets for l in a.iter_rows() for c in l
                        if c.value == "PREENCHER")
        exigir(pendentes == 0, f"{pendentes} celulas ainda marcadas PREENCHER")

    # A posicao oficial exige dados reais no banco, nao demonstrativos.
    try:
        from src.database.connection import connect
        conn = connect()
        try:
            demo = conn.execute(
                "SELECT COUNT(*) FROM market_observations WHERE classification='demonstracao'"
            ).fetchone()[0]
            real = conn.execute(
                "SELECT COUNT(*) FROM market_observations WHERE classification IN "
                "('observado','negociado')").fetchone()[0]
            exigir(real > 0, "nenhuma observacao real no banco (so demonstrativas)")
            if demo and not real:
                PROBLEMAS.append(f"{demo} observacoes demonstrativas e nenhuma real")
        finally:
            conn.close()
    except Exception as exc:
        PROBLEMAS.append(f"banco inacessivel: {exc}")

    if PROBLEMAS:
        print("ENTREGA 2 INCOMPLETA:")
        for p in PROBLEMAS:
            print("  -", p)
        print("\nVeja READY_FOR_ENTREGA_2.md para o passo a passo.")
        return 1
    print("ENTREGA 2 VERIFICADA: documento, planilha e dados reais conferidos.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
